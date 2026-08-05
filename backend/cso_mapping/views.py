"""API for the CSO Mapping & Capacity Assessment.

Public (unauthenticated) surface:
  * GET  /api/cso-mapping/schema/  — the form definition the frontend renders.
  * POST /api/cso-mapping/submit/  — accept one questionnaire response (throttled).

Authorised-staff surface (admin role only — submissions hold personal data):
  * GET  /api/cso-mapping/submissions/           — list (filter/search/order).
  * GET  /api/cso-mapping/submissions/<id>/      — retrieve one.
  * GET  /api/cso-mapping/submissions/export/    — CSV of all submissions.
  * GET  /api/cso-mapping/submissions/summary/   — counts by respondent type.
"""
from __future__ import annotations

import re
from io import BytesIO

from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import ScopedRateThrottle
from rest_framework.views import APIView

from audit.recording import record_audit_event

from .models import CsoMappingDraft, CsoMappingSchemaVersion, CsoMappingSubmission
from .permissions import CanUseCsoMapping
from .schema import (
    CORE_BOOL_FIELDS,
    CORE_TEXT_FIELDS,
    SchemaValidationError,
    _bundled_schema,
    field_is_active,
    iter_answerable_fields,
    load_schema,
    validate_schema,
)
from .serializers import (
    DraftWriteSerializer,
    PublicSubmissionSerializer,
    StaffSubmissionSerializer,
)

# Characters Excel forbids in a worksheet title, plus the 31-char cap.
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")

# Leading characters a spreadsheet may interpret as a formula. Respondent text
# beginning with one of these is prefixed with an apostrophe so the cell is a
# literal string (formula / CSV-injection defence). The stored value is unchanged.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _safe_sheet_title(label: str) -> str:
    return _INVALID_SHEET_CHARS.sub(" ", label).strip()[:31] or "Sheet"


def _excel_safe(value):
    """Neutralise formula-injection in a spreadsheet cell (leave DB value as-is)."""
    if isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _submission_receipt(submission, *, created):
    """Public-safe receipt payload — a reference, not the sequential PK."""
    return {
        "reference": submission.public_reference,
        "submitted_at": submission.submitted_at,
        "responding_entity": submission.responding_entity,
        "detail": (
            "Submission received. Thank you."
            if created
            else "This response has already been received. Thank you."
        ),
    }


def _draft_state(draft, *, include_answers=False):
    """Draft state for the client. Never includes the token (only its hash is stored)."""
    state = {
        "current_step": draft.current_step,
        "form_version": draft.form_version,
        "updated_at": draft.updated_at,
        "expires_at": draft.expires_at,
        "client_submission_id": (
            str(draft.client_submission_id) if draft.client_submission_id else None
        ),
    }
    if include_answers:
        state["answers"] = draft.answers
    return state


# The draft resume token travels in this request header — never in the URL, query
# string, body, logs, audit events or error messages.
DRAFT_TOKEN_HEADER = "X-CSO-Draft-Token"


def _request_draft_token(request):
    return (request.headers.get(DRAFT_TOKEN_HEADER) or "").strip()


def _active_draft(token):
    """Resolve a non-expired, non-completed draft by its raw token, else None.

    Lookup is by token *hash*, so the raw token is never stored or logged. Returns
    None uniformly for unknown/expired/completed tokens so responses do not reveal
    whether another person's token exists.
    """
    if not token:
        return None
    return CsoMappingDraft.objects.filter(
        token_hash=CsoMappingDraft.hash_token(token),
        expires_at__gte=timezone.now(),
        completed_at__isnull=True,
    ).first()


def _create_or_replay_submission(serializer):
    """Persist a submission idempotently on client_submission_id. Returns (obj, created)."""
    client_id = serializer.validated_data.get("client_submission_id")
    if client_id:
        existing = CsoMappingSubmission.objects.filter(client_submission_id=client_id).first()
        if existing:
            return existing, False
    try:
        with transaction.atomic():
            return serializer.save(), True
    except IntegrityError:
        existing = CsoMappingSubmission.objects.filter(client_submission_id=client_id).first()
        if existing:
            return existing, False
        raise


def _complete_draft(draft, submission):
    """Mark a draft completed, link the submission, and clear its personal data."""
    draft.completed_at = timezone.now()
    draft.submission = submission
    draft.answers = {}
    draft.save(update_fields=["completed_at", "submission", "answers", "updated_at"])


class FormSchemaView(APIView):
    """Public: the canonical form schema the frontend renders.

    A static form definition (no user input), so it is not throttled — many
    respondents behind one office NAT must be able to load the form freely.
    """

    permission_classes = [AllowAny]

    def get(self, request):
        return Response(load_schema())


class SubmissionCreateView(APIView):
    """Public: accept one questionnaire submission.

    The only unauthenticated write path for this app, so it is rate-limited per
    IP (throttle scope ``cso_mapping``). Validation is schema-driven; consent is
    mandatory.
    """

    permission_classes = [AllowAny]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "cso_mapping"

    def post(self, request):
        serializer = PublicSubmissionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        submission, created = _create_or_replay_submission(serializer)
        # If this submission came from a draft, mark that draft completed.
        draft = _active_draft(str(request.data.get("resume_token", "")).strip())
        if draft is not None:
            _complete_draft(draft, submission)
        return Response(
            _submission_receipt(submission, created=created),
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class DraftCreateView(APIView):
    """Public: create a resumable draft, returning its opaque resume token once.

    Row-creation is bounded by the same per-IP scope as submit. Only the token's
    hash is stored; the raw token is returned here and never again.
    """

    permission_classes = [AllowAny]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "cso_mapping"

    def post(self, request):
        serializer = DraftWriteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        raw_token = CsoMappingDraft.new_raw_token()
        draft = serializer.save(token_hash=CsoMappingDraft.hash_token(raw_token))
        state = _draft_state(draft)
        state["resume_token"] = raw_token  # issued once; DB keeps only the hash
        return Response(state, status=status.HTTP_201_CREATED)


class DraftDetailView(APIView):
    """Public: restore (GET), autosave (PUT), or discard (DELETE) the draft named
    by the ``X-CSO-Draft-Token`` header.

    The token is the sole capability and travels ONLY in the header (never the
    URL). Lookup is by hash. Expired or completed drafts return 404 uniformly
    (never revealing whether a token exists).
    """

    permission_classes = [AllowAny]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "cso_mapping_draft"

    def get(self, request):
        draft = _active_draft(_request_draft_token(request))
        if draft is None:
            return Response(
                {"detail": "Draft not found or expired."}, status=status.HTTP_404_NOT_FOUND
            )
        return Response(_draft_state(draft, include_answers=True))

    def put(self, request):
        draft = _active_draft(_request_draft_token(request))
        if draft is None:
            return Response(
                {"detail": "Draft not found or expired."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = DraftWriteSerializer(draft, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(_draft_state(draft))

    def delete(self, request):
        draft = _active_draft(_request_draft_token(request))
        if draft is not None:
            draft.delete()
        # Uniform 204 whether or not the token existed.
        return Response(status=status.HTTP_204_NO_CONTENT)


class DraftSubmitView(APIView):
    """Public: atomically convert the header-named draft into a completed submission.

    Full validation runs here (unlike draft autosave). Submission creation and the
    draft's completion happen in one transaction. Idempotent on client_submission_id.
    """

    permission_classes = [AllowAny]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "cso_mapping"

    def post(self, request):
        token = _request_draft_token(request)
        draft = CsoMappingDraft.objects.filter(
            token_hash=CsoMappingDraft.hash_token(token)
        ).first() if token else None
        if draft is None:
            return Response(
                {"detail": "Draft not found or expired."}, status=status.HTTP_404_NOT_FOUND
            )
        # Already submitted: replay the original receipt (retry-safe, no duplicate).
        if draft.completed_at is not None:
            if draft.submission_id:
                return Response(
                    _submission_receipt(draft.submission, created=False),
                    status=status.HTTP_200_OK,
                )
            return Response(
                {"detail": "Draft not found or expired."}, status=status.HTTP_404_NOT_FOUND
            )
        if draft.expires_at < timezone.now():
            return Response(
                {"detail": "Draft not found or expired."}, status=status.HTTP_404_NOT_FOUND
            )

        serializer = PublicSubmissionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            submission, created = _create_or_replay_submission(serializer)
            _complete_draft(draft, submission)
        return Response(
            _submission_receipt(submission, created=created),
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class SubmissionViewSet(viewsets.ReadOnlyModelViewSet):
    """Authorised-staff read access to submissions (admin role only)."""

    queryset = CsoMappingSubmission.objects.all()
    serializer_class = StaffSubmissionSerializer
    permission_classes = [IsAuthenticated, CanUseCsoMapping]
    filterset_fields = ["respondent_type"]
    search_fields = ["responding_entity", "respondent_name", "primary_district"]
    ordering_fields = ["submitted_at", "respondent_type", "responding_entity"]
    ordering = ["-submitted_at"]

    def retrieve(self, request, *args, **kwargs):
        response = super().retrieve(request, *args, **kwargs)
        record_audit_event(
            action="view",
            request=request,
            object_type="cso_submission",
            object_id=kwargs.get("pk"),
            description="Viewed a CSO mapping submission",
        )
        return response

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        returned = (
            len(response.data.get("results", []))
            if isinstance(response.data, dict)
            else None
        )
        record_audit_event(
            action="view",
            request=request,
            object_type="cso_submission_list",
            description="Accessed the CSO mapping submissions list",
            metadata={
                "search": request.query_params.get("search", ""),
                "respondent_type": request.query_params.get("respondent_type", ""),
                "returned": returned,
            },
        )
        return response

    @action(detail=False, methods=["get"])
    def summary(self, request):
        qs = self.get_queryset()
        by_type = {
            row["respondent_type"]: row["n"]
            for row in qs.values("respondent_type").annotate(n=Count("id"))
        }
        return Response({"total": qs.count(), "by_respondent_type": by_type})

    # Short, Excel-safe worksheet titles per respondent category.
    SHEET_TITLES = {
        "cso": "Health Service CSOs",
        "coordinating_body": "Coordinating Bodies",
        "strategic_structure": "Strategic Structures",
    }

    @action(detail=False, methods=["get"])
    def export(self, request):
        """Excel workbook with one sheet per respondent category.

        Each sheet carries only the questions relevant to that category (admin
        fields + that annex + final confirmation), one row per submission. The
        respondent-type filter is ignored (the workbook is split by category by
        design); a ``search`` term, if given, is applied across all sheets.
        """
        schema = load_schema()
        base = self.get_queryset()
        search = request.query_params.get("search")
        if search:
            base = base.filter(
                Q(responding_entity__icontains=search)
                | Q(respondent_name__icontains=search)
                | Q(primary_district__icontains=search)
            )

        wb = Workbook()
        wb.remove(wb.active)  # drop the auto-created default sheet
        total_rows = 0

        for choice in schema.get("choices", {}).get("respondent_type", []):
            rtype = choice["name"]
            context = {"consent": "yes", "respondent_type": rtype}
            columns = [
                (field["name"], field.get("label") or field["name"])
                for section, field in iter_answerable_fields(schema)
                if field_is_active(section, field, context)
            ]
            ws = wb.create_sheet(
                title=self.SHEET_TITLES.get(rtype) or _safe_sheet_title(choice["label"])
            )
            headers = ["Reference", "Submitted at"] + [label for _name, label in columns]
            ws.append(headers)
            for sub in base.filter(respondent_type=rtype):
                # openpyxl cannot write tz-aware datetimes; store as naive.
                submitted = sub.submitted_at.replace(tzinfo=None) if sub.submitted_at else None
                row = [sub.public_reference or "", submitted]
                answers = sub.answers or {}
                for name, _label in columns:
                    if name in CORE_BOOL_FIELDS:
                        row.append("Yes" if getattr(sub, name) else "No")
                    elif name in CORE_TEXT_FIELDS:
                        row.append(_excel_safe(getattr(sub, name) or ""))
                    else:
                        row.append(_excel_safe(answers.get(name, "")))
                ws.append(row)
                total_rows += 1
            self._style_sheet(ws, len(headers))

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        record_audit_event(
            action="export",
            request=request,
            object_type="cso_submission",
            description="Exported CSO mapping submissions to Excel",
            metadata={"search": search or "", "records": total_rows},
        )

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="cso-mapping-submissions.xlsx"'
        return response

    @staticmethod
    def _style_sheet(ws, num_columns):
        """Freeze the header, enable filters, format dates, cap widths, wrap text."""
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        last_col = get_column_letter(num_columns)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
        # Date format on the "Submitted at" column (B), skipping the header.
        for (cell,) in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            cell.number_format = "yyyy-mm-dd hh:mm"
        # Readable-but-capped widths: reference/date narrow, questions wide.
        for idx in range(1, num_columns + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 20 if idx <= 2 else 42
        wrap = Alignment(vertical="top", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap


# ---------------------------------------------------------------------------
# Form editor (admin-only): edit the questionnaire schema without a code deploy.
# ---------------------------------------------------------------------------
def _ensure_schema_seed() -> None:
    """Seed the first DB version from the bundled file if the table is empty, so the
    editor always has a version to load and the DB becomes the source of truth."""
    if not CsoMappingSchemaVersion.objects.exists():
        bundled = _bundled_schema()
        CsoMappingSchemaVersion.objects.create(
            schema=bundled,
            version_label=str(bundled.get("version", ""))[:64],
            note="Initial version (imported from the bundled form).",
            is_active=True,
        )


def _version_summary(version) -> dict:
    return {
        "id": version.pk,
        "version_label": version.version_label,
        "note": version.note,
        "is_active": version.is_active,
        "created_at": version.created_at,
        "created_by": getattr(version.created_by, "username", None),
    }


def _activate_new_version(schema, *, label, note, user):
    """Deactivate the current active version and make `schema` the new active one."""
    with transaction.atomic():
        CsoMappingSchemaVersion.objects.filter(is_active=True).update(is_active=False)
        return CsoMappingSchemaVersion.objects.create(
            schema=schema,
            version_label=str(label)[:64],
            note=str(note)[:255],
            is_active=True,
            created_by=user if getattr(user, "is_authenticated", False) else None,
        )


class CsoMappingSchemaAdminView(APIView):
    """Admin-only: read or replace the active questionnaire schema.

    GET returns the live schema (same shape as the public endpoint). PUT accepts
    ``{"schema": <schema>, "note": <optional str>}`` (or a bare schema), validates
    it, and — only if valid — saves it as the new active version. A rejected edit
    never touches the live form.
    """

    permission_classes = [IsAuthenticated, CanUseCsoMapping]

    def get(self, request):
        _ensure_schema_seed()
        return Response(load_schema())

    def put(self, request):
        payload = request.data if isinstance(request.data, dict) else {}
        schema = payload.get("schema", payload)
        note = payload.get("note", "") if isinstance(payload, dict) else ""
        try:
            validate_schema(schema)
        except SchemaValidationError as exc:
            return Response({"errors": exc.errors}, status=status.HTTP_400_BAD_REQUEST)
        version = _activate_new_version(
            schema,
            label=schema.get("version", "") if isinstance(schema, dict) else "",
            note=note,
            user=request.user,
        )
        record_audit_event(
            action="update",
            request=request,
            object_type="cso_schema",
            object_id=str(version.pk),
            description="Updated the CSO mapping questionnaire form",
            metadata={"version": version.version_label},
        )
        return Response(load_schema())


class CsoMappingSchemaHistoryView(APIView):
    """Admin-only: list saved schema versions (most recent first)."""

    permission_classes = [IsAuthenticated, CanUseCsoMapping]

    def get(self, request):
        _ensure_schema_seed()
        versions = CsoMappingSchemaVersion.objects.select_related("created_by")[:50]
        return Response([_version_summary(v) for v in versions])


class CsoMappingSchemaActivateView(APIView):
    """Admin-only: roll back by re-activating a previous version (copied forward)."""

    permission_classes = [IsAuthenticated, CanUseCsoMapping]

    def post(self, request, pk):
        target = CsoMappingSchemaVersion.objects.filter(pk=pk).first()
        if target is None:
            return Response(
                {"detail": "Version not found."}, status=status.HTTP_404_NOT_FOUND
            )
        try:
            validate_schema(target.schema)
        except SchemaValidationError as exc:
            return Response({"errors": exc.errors}, status=status.HTTP_400_BAD_REQUEST)
        version = _activate_new_version(
            target.schema,
            label=target.version_label,
            note=f"Rolled back to version #{target.pk}.",
            user=request.user,
        )
        record_audit_event(
            action="update",
            request=request,
            object_type="cso_schema",
            object_id=str(version.pk),
            description=f"Rolled back the CSO mapping form to version #{target.pk}",
        )
        return Response(load_schema())
