"""Tests for the CSO Mapping questionnaire API."""
from __future__ import annotations

import hashlib
import json
import uuid
from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase as _DRFAPITestCase

from .models import CsoMappingDraft, CsoMappingSubmission
from .schema import field_is_active, iter_answerable_fields, load_schema

User = get_user_model()


# Isolate every CSO API test to a per-process in-memory cache. The public submit
# throttle stores its counters in the shared file-based cache; without this they
# would accumulate across test runs (and could collide with other state). This is
# also safer than clearing the shared cache. Shadowing the name keeps the test
# classes below unchanged.
@override_settings(
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "cso-mapping-tests",
        }
    },
    CSO_MAPPING_DRAFT_TTL_DAYS=14,
)
class APITestCase(_DRFAPITestCase):
    def _pre_setup(self):
        super()._pre_setup()
        # Reset the (isolated) throttle cache before each test so the public
        # submit/draft rate limits do not accumulate across the suite.
        from django.core.cache import cache

        cache.clear()


# A valid Gaborone office location (inside Botswana's extent — not flagged).
VALID_LOCATION = {
    "latitude": "-24.628200",
    "longitude": "25.923100",
    "location_accuracy": "12.50",
    "location_captured_at": "2026-08-06T10:00:00Z",
    "location_capture_method": "device_gps",
}


def build_valid_payload(respondent_type: str = "cso") -> dict:
    """Construct a complete, valid submission for a given respondent branch."""
    schema = load_schema()
    ctx = {"consent": "yes", "respondent_type": respondent_type, "information_confirmed": "yes"}
    payload = dict(ctx)
    for section, field in iter_answerable_fields(schema):
        # Evaluate relevance against the accumulating payload so conditional
        # fields (e.g. registration_number when "registered") are populated.
        if not field_is_active(section, field, payload):
            continue
        name = field["name"]
        if name in payload:
            continue
        if field["type"] == "select_one":
            payload[name] = field["choices"][0]["name"]
        elif field["type"] == "select_multiple":
            payload[name] = [field["choices"][0]["name"]]
        elif name == "respondent_email":
            payload[name] = "respondent@example.org"
        else:
            payload[name] = f"Answer for {name}"
    # Office GPS location is now required for a submission.
    payload.update(VALID_LOCATION)
    return payload


class SchemaEndpointTests(APITestCase):
    def test_schema_is_public(self):
        resp = self.client.get(reverse("cso-mapping-schema"))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIn("sections", resp.data)
        self.assertTrue(len(resp.data["sections"]) > 0)


class SubmissionCreateTests(APITestCase):
    def setUp(self):
        self.url = reverse("cso-mapping-submit")

    def test_valid_cso_submission_is_stored_locally(self):
        payload = build_valid_payload("cso")
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        self.assertEqual(CsoMappingSubmission.objects.count(), 1)
        sub = CsoMappingSubmission.objects.get()
        self.assertTrue(sub.consent)
        self.assertEqual(sub.respondent_type, "cso")
        self.assertTrue(sub.information_confirmed)
        # Annex 2 answers land in the JSON blob; other annexes are absent.
        self.assertIn("annex2_a2_1a", sub.answers)
        self.assertFalse(any(k.startswith("annex3_") for k in sub.answers))
        self.assertFalse(any(k.startswith("annex4_") for k in sub.answers))

    def test_consent_is_required(self):
        payload = build_valid_payload("cso")
        payload["consent"] = "no"
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("consent", resp.data)
        self.assertEqual(CsoMappingSubmission.objects.count(), 0)

    def test_missing_required_admin_field_rejected(self):
        payload = build_valid_payload("cso")
        payload["responding_entity"] = ""
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("responding_entity", resp.data)

    def test_missing_required_annex_field_rejected(self):
        payload = build_valid_payload("cso")
        payload["annex2_a2_1a"] = ""
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("annex2_a2_1a", resp.data)

    def test_information_confirmed_must_be_yes(self):
        payload = build_valid_payload("cso")
        payload["information_confirmed"] = "no"
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("information_confirmed", resp.data)

    def test_invalid_select_choice_rejected(self):
        payload = build_valid_payload("cso")
        payload["respondent_type"] = "not_a_real_type"
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_inactive_branch_answers_are_ignored(self):
        # A CSO respondent who somehow also sends an Annex 3 answer: it must be
        # neither required nor stored.
        payload = build_valid_payload("cso")
        payload["annex3_a3_1a"] = "stray value"
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        sub = CsoMappingSubmission.objects.get()
        self.assertNotIn("annex3_a3_1a", sub.answers)

    def test_coordinating_body_branch(self):
        payload = build_valid_payload("coordinating_body")
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        sub = CsoMappingSubmission.objects.get()
        self.assertIn("annex3_a3_1a", sub.answers)


class StaffAccessTests(APITestCase):
    """Fix A: raw submissions are admin-only; is_staff alone must not grant access."""

    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin1", email="admin1@example.com", password="x", role="admin"
        )
        self.officer = User.objects.create_user(
            username="officer1", email="officer1@example.com", password="x", role="officer"
        )
        self.manager = User.objects.create_user(
            username="manager1", email="manager1@example.com", password="x", role="manager"
        )
        # A non-administrator who merely has the Django admin-site is_staff flag.
        self.staff_only = User.objects.create_user(
            username="staff1", email="staff1@example.com", password="x",
            role="officer", is_staff=True,
        )
        self.superuser = User.objects.create_user(
            username="root1", email="root1@example.com", password="x",
            role="officer", is_superuser=True,
        )
        self.client.post(
            reverse("cso-mapping-submit"), build_valid_payload("cso"), format="json"
        )

    def _list_status(self, user):
        self.client.force_authenticate(user)
        return self.client.get(reverse("cso-submissions-list")).status_code

    def test_anonymous_denied(self):
        resp = self.client.get(reverse("cso-submissions-list"))
        self.assertIn(resp.status_code, (status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN))

    def test_officer_denied(self):
        self.assertEqual(self._list_status(self.officer), status.HTTP_403_FORBIDDEN)

    def test_manager_denied(self):
        self.assertEqual(self._list_status(self.manager), status.HTTP_403_FORBIDDEN)

    def test_is_staff_alone_denied(self):
        # The core of Fix A: is_staff=True must NOT unlock personal data.
        self.assertEqual(self._list_status(self.staff_only), status.HTTP_403_FORBIDDEN)

    def test_admin_allowed(self):
        self.assertEqual(self._list_status(self.admin), status.HTTP_200_OK)

    def test_superuser_allowed(self):
        self.assertEqual(self._list_status(self.superuser), status.HTTP_200_OK)

    def test_detail_summary_export_are_all_admin_gated(self):
        sub = CsoMappingSubmission.objects.get()
        targets = [
            reverse("cso-submissions-detail", kwargs={"pk": sub.pk}),
            reverse("cso-submissions-summary"),
            reverse("cso-submissions-export"),
        ]
        for url in targets:
            self.client.force_authenticate(self.officer)
            self.assertEqual(self.client.get(url).status_code, status.HTTP_403_FORBIDDEN, url)
            self.client.force_authenticate(self.admin)
            self.assertEqual(self.client.get(url).status_code, status.HTTP_200_OK, url)

    def test_admin_export_is_hardened_three_sheet_workbook(self):
        from openpyxl import load_workbook

        self.client.force_authenticate(self.admin)
        resp = self.client.get(reverse("cso-submissions-export"))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = load_workbook(BytesIO(resp.content))
        self.assertEqual(
            wb.sheetnames,
            ["Health Service CSOs", "Coordinating Bodies", "Strategic Structures"],
        )
        cso = wb["Health Service CSOs"]
        self.assertEqual(cso.max_row, 2)  # header + 1 CSO submission
        self.assertEqual(cso.freeze_panes, "A2")  # header frozen
        self.assertIsNotNone(cso.auto_filter.ref)  # filters enabled
        headers = [cell.value for cell in cso[1]]
        self.assertTrue(any(h and "Nature/type of your organisation" in h for h in headers))
        # A category with no submissions still gets a header-only sheet.
        self.assertEqual(wb["Coordinating Bodies"].max_row, 1)


class IdempotencyTests(APITestCase):
    """Fix B: a repeated questionnaire attempt must not create a second row."""

    def setUp(self):
        self.url = reverse("cso-mapping-submit")

    def test_same_client_id_creates_one_row_and_replays_receipt(self):
        payload = build_valid_payload("cso")
        payload["client_submission_id"] = str(uuid.uuid4())
        first = self.client.post(self.url, payload, format="json")
        self.assertEqual(first.status_code, status.HTTP_201_CREATED, first.data)
        again = self.client.post(self.url, payload, format="json")
        self.assertEqual(again.status_code, status.HTTP_200_OK)  # replay, not created
        self.assertEqual(CsoMappingSubmission.objects.count(), 1)
        self.assertEqual(first.data["reference"], again.data["reference"])

    def test_different_client_id_creates_second_row(self):
        p1 = build_valid_payload("cso"); p1["client_submission_id"] = str(uuid.uuid4())
        p2 = build_valid_payload("cso"); p2["client_submission_id"] = str(uuid.uuid4())
        self.client.post(self.url, p1, format="json")
        self.client.post(self.url, p2, format="json")
        self.assertEqual(CsoMappingSubmission.objects.count(), 2)

    def test_invalid_client_id_rejected(self):
        payload = build_valid_payload("cso")
        payload["client_submission_id"] = "not-a-uuid"
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("client_submission_id", resp.data)

    def test_db_unique_constraint_blocks_concurrent_duplicate(self):
        from django.db import IntegrityError, transaction

        cid = uuid.uuid4()
        CsoMappingSubmission.objects.create(respondent_type="cso", client_submission_id=cid)
        with self.assertRaises(IntegrityError), transaction.atomic():
            CsoMappingSubmission.objects.create(respondent_type="cso", client_submission_id=cid)


class ReceiptTests(APITestCase):
    """Fix G: the response exposes a public reference, not the sequential PK."""

    def test_receipt_shape(self):
        resp = self.client.post(
            reverse("cso-mapping-submit"), build_valid_payload("cso"), format="json"
        )
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        self.assertNotIn("id", resp.data)  # do not leak the sequential PK
        self.assertRegex(resp.data["reference"], r"^CSO-\d{4}-[0-9A-F]{8}$")
        self.assertIn("responding_entity", resp.data)
        self.assertIn("submitted_at", resp.data)


class ExcelInjectionTests(APITestCase):
    """Fix H: respondent-supplied formulas are neutralised in the workbook."""

    def setUp(self):
        self.admin = User.objects.create_user(
            username="ad", email="ad@example.com", password="x", role="admin"
        )

    def test_formula_values_are_neutralised_but_db_unchanged(self):
        from openpyxl import load_workbook

        payload = build_valid_payload("cso")
        # Use free-text fields (the structured ones are now fixed-choice).
        payload["responding_entity"] = '=HYPERLINK("https://example.com","Click")'
        payload["annex2_a2_2a"] = "@SUM(A1:A3)"
        payload["annex2_a2_2b"] = "+1+1"
        payload["annex2_a2_2c"] = "-2+3"
        self.client.post(reverse("cso-mapping-submit"), payload, format="json")

        sub = CsoMappingSubmission.objects.get()
        # The database keeps the exact original value.
        self.assertEqual(sub.responding_entity, '=HYPERLINK("https://example.com","Click")')
        self.assertEqual(sub.answers["annex2_a2_2a"], "@SUM(A1:A3)")

        self.client.force_authenticate(self.admin)
        resp = self.client.get(reverse("cso-submissions-export"))
        wb = load_workbook(BytesIO(resp.content))
        cso = wb["Health Service CSOs"]
        cell_values = [c.value for row in cso.iter_rows(min_row=2) for c in row]
        dangerous = [
            v for v in cell_values if isinstance(v, str) and v[:1] in ("=", "+", "-", "@")
        ]
        self.assertEqual(dangerous, [], f"un-neutralised formula cells: {dangerous}")


class AuditTests(APITestCase):
    """Fix I: admin views and exports are recorded in the audit stream."""

    def setUp(self):
        self.admin = User.objects.create_user(
            username="ad", email="ad@example.com", password="x", role="admin"
        )
        self.client.post(
            reverse("cso-mapping-submit"), build_valid_payload("cso"), format="json"
        )
        self.sub = CsoMappingSubmission.objects.get()

    def test_detail_view_is_audited(self):
        from audit.models import AuditEvent

        self.client.force_authenticate(self.admin)
        self.client.get(reverse("cso-submissions-detail", kwargs={"pk": self.sub.pk}))
        self.assertTrue(
            AuditEvent.objects.filter(
                action="view", object_type="cso_submission", object_id=str(self.sub.pk)
            ).exists()
        )

    def test_export_is_audited_with_count_and_actor(self):
        from audit.models import AuditEvent

        self.client.force_authenticate(self.admin)
        self.client.get(reverse("cso-submissions-export"))
        event = (
            AuditEvent.objects.filter(action="export", object_type="cso_submission")
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(event)
        self.assertEqual(event.metadata.get("records"), 1)
        self.assertEqual(event.actor, self.admin)


class SchemaIntegrityTests(APITestCase):
    """Fix K: branch isolation + final confirmation are enforced by the schema."""

    def test_annex_branches_are_isolated_and_confirmation_enforced(self):
        schema = load_schema()
        gate = {
            "annex2": "cso",
            "annex3": "coordinating_body",
            "annex4": "strategic_structure",
        }
        all_types = ["cso", "coordinating_body", "strategic_structure"]
        for section in schema["sections"]:
            for field in section["fields"]:
                for prefix, owner in gate.items():
                    if not field["name"].startswith(prefix):
                        continue
                    # Active only under its owning respondent type (checks section
                    # AND field relevance — path-intro notes are gated per-field).
                    for rtype in all_types:
                        ctx = {"consent": "yes", "respondent_type": rtype}
                        active = field_is_active(section, field, ctx)
                        if rtype == owner:
                            # "Other — specify" follow-ups are only active once
                            # their trigger option is chosen, so skip them here.
                            if not field["name"].endswith("_other"):
                                self.assertTrue(active, f"{field['name']} inactive for {owner}")
                        else:
                            self.assertFalse(active, f"{field['name']} leaked into {rtype}")
        confirm = next(
            f
            for s in schema["sections"]
            for f in s["fields"]
            if f["name"] == "information_confirmed"
        )
        self.assertTrue(confirm["required"])
        self.assertEqual(confirm["constraint"]["value"], "yes")


def _create_draft(client, answers=None, step=0, client_submission_id=None):
    body = {"answers": answers or {}, "current_step": step, "form_version": "test"}
    if client_submission_id:
        body["client_submission_id"] = client_submission_id
    return client.post(reverse("cso-mapping-draft-create"), body, format="json")


def _detail_url():
    return reverse("cso-mapping-draft-detail")


def _submit_url():
    return reverse("cso-mapping-draft-submit")


def _tok(token):
    # The draft token travels in the X-CSO-Draft-Token header, never the URL.
    return {"HTTP_X_CSO_DRAFT_TOKEN": token}


class DraftLifecycleTests(APITestCase):
    """Section C: create / update / retrieve / delete / expire / complete."""

    def test_create_returns_opaque_token_and_state(self):
        resp = _create_draft(self.client, {"consent": "yes"}, step=1)
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        self.assertIn("resume_token", resp.data)
        self.assertGreater(len(resp.data["resume_token"]), 30)  # high-entropy token
        self.assertEqual(resp.data["current_step"], 1)
        self.assertIn("expires_at", resp.data)

    def test_token_never_appears_in_url(self):
        # Routes are token-free; the token is supplied via header only.
        self.assertEqual(_detail_url(), "/api/cso-mapping/drafts/current/")
        self.assertEqual(_submit_url(), "/api/cso-mapping/drafts/current/submit/")
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        self.assertNotIn(token, _detail_url())
        # Header-based retrieval works.
        self.assertEqual(
            self.client.get(_detail_url(), **_tok(token)).status_code, status.HTTP_200_OK
        )
        # No header -> uniform 404 (does not reveal existence).
        self.assertEqual(self.client.get(_detail_url()).status_code, status.HTTP_404_NOT_FOUND)

    def test_raw_token_is_not_stored_only_its_hash(self):
        resp = _create_draft(self.client, {"consent": "yes"})
        raw = resp.data["resume_token"]
        draft = CsoMappingDraft.objects.get()
        self.assertEqual(draft.token_hash, hashlib.sha256(raw.encode()).hexdigest())
        self.assertNotEqual(draft.token_hash, raw)  # not the raw token
        # The raw token appears in no stored text field.
        for value in (draft.token_hash, draft.form_version, str(draft.answers)):
            self.assertNotIn(raw, value)

    def test_update_and_retrieve(self):
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        put = self.client.put(
            _detail_url(),
            {"answers": {"consent": "yes", "respondent_type": "cso"}, "current_step": 2},
            format="json",
            **_tok(token),
        )
        self.assertEqual(put.status_code, status.HTTP_200_OK)
        got = self.client.get(_detail_url(), **_tok(token))
        self.assertEqual(got.status_code, status.HTTP_200_OK)
        self.assertEqual(got.data["current_step"], 2)
        self.assertEqual(got.data["answers"]["respondent_type"], "cso")

    def test_invalid_token_is_404(self):
        self.assertEqual(
            self.client.get(_detail_url(), **_tok("nope")).status_code, status.HTTP_404_NOT_FOUND
        )
        self.assertEqual(
            self.client.put(_detail_url(), {"answers": {}}, format="json", **_tok("nope")).status_code,
            status.HTTP_404_NOT_FOUND,
        )

    def test_expired_token_is_404(self):
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        CsoMappingDraft.objects.update(expires_at=timezone.now() - timedelta(minutes=1))
        self.assertEqual(
            self.client.get(_detail_url(), **_tok(token)).status_code, status.HTTP_404_NOT_FOUND
        )

    def test_deleted_token_is_404(self):
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        self.assertEqual(
            self.client.delete(_detail_url(), **_tok(token)).status_code, status.HTTP_204_NO_CONTENT
        )
        self.assertEqual(
            self.client.get(_detail_url(), **_tok(token)).status_code, status.HTTP_404_NOT_FOUND
        )

    def test_delete_unknown_token_still_204(self):
        # Does not reveal whether the token existed.
        self.assertEqual(
            self.client.delete(_detail_url(), **_tok("nope")).status_code, status.HTTP_204_NO_CONTENT
        )

    @override_settings(CSO_MAPPING_DRAFT_TTL_DAYS=1)
    def test_expiry_uses_configured_ttl(self):
        resp = _create_draft(self.client, {"consent": "yes"})
        draft = CsoMappingDraft.objects.get()
        delta = draft.expires_at - timezone.now()
        self.assertGreater(delta, timedelta(hours=23))
        self.assertLess(delta, timedelta(hours=25))

    def test_expired_drafts_cleanup_command(self):
        keep = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        gone = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        CsoMappingDraft.objects.filter(
            token_hash=CsoMappingDraft.hash_token(gone)
        ).update(expires_at=timezone.now() - timedelta(days=1))
        call_command("cleanup_cso_drafts")
        self.assertTrue(
            CsoMappingDraft.objects.filter(token_hash=CsoMappingDraft.hash_token(keep)).exists()
        )
        self.assertFalse(
            CsoMappingDraft.objects.filter(token_hash=CsoMappingDraft.hash_token(gone)).exists()
        )


class DraftValidationTests(APITestCase):
    """Section C: drafts are lenient on required, strict on choices/size/branch."""

    def _put(self, answers):
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        return (
            self.client.put(_detail_url(), {"answers": answers}, format="json", **_tok(token)),
            token,
        )

    def test_required_fields_may_stay_incomplete(self):
        resp, _ = self._put({"consent": "yes", "respondent_type": "cso"})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)  # no required-field block

    def test_invalid_choice_rejected(self):
        resp, _ = self._put({"consent": "yes", "respondent_type": "not_valid"})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_oversized_value_rejected(self):
        resp, _ = self._put({"consent": "yes", "respondent_type": "cso", "annex2_a2_1a": "x" * 20001})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_unknown_field_ignored(self):
        resp, token = self._put({"consent": "yes", "totally_unknown": "x"})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        got = self.client.get(_detail_url(), **_tok(token))
        self.assertNotIn("totally_unknown", got.data["answers"])

    def test_inactive_branch_answers_stripped(self):
        resp, token = self._put(
            {"consent": "yes", "respondent_type": "cso", "annex3_a3_1a": ["ngo"]}
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        got = self.client.get(_detail_url(), **_tok(token))
        self.assertNotIn("annex3_a3_1a", got.data["answers"])

    def test_respondent_type_change_removes_old_branch(self):
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        self.client.put(
            _detail_url(),
            {"answers": {"consent": "yes", "respondent_type": "cso", "annex2_a2_1a": "ngo"}},
            format="json",
            **_tok(token),
        )
        self.client.put(
            _detail_url(),
            {"answers": {"consent": "yes", "respondent_type": "coordinating_body"}},
            format="json",
            **_tok(token),
        )
        answers = self.client.get(_detail_url(), **_tok(token)).data["answers"]
        self.assertNotIn("annex2_a2_1a", answers)
        self.assertEqual(answers["respondent_type"], "coordinating_body")


class DraftSubmitTests(APITestCase):
    """Section C: atomic draft -> submission conversion + idempotency."""

    def test_final_submit_validates_all_required(self):
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        # Missing required annex answers -> full validation fails; nothing created.
        resp = self.client.post(
            _submit_url(),
            {"consent": "yes", "respondent_type": "cso", "responding_entity": "X"},
            format="json",
            **_tok(token),
        )
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(CsoMappingSubmission.objects.count(), 0)
        # Draft is untouched (conversion is atomic).
        self.assertEqual(self.client.get(_detail_url(), **_tok(token)).status_code, status.HTTP_200_OK)

    def test_conversion_creates_one_submission_and_completes_draft(self):
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        payload = build_valid_payload("cso")
        payload["client_submission_id"] = str(uuid.uuid4())
        resp = self.client.post(_submit_url(), payload, format="json", **_tok(token))
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        self.assertEqual(CsoMappingSubmission.objects.count(), 1)
        draft = CsoMappingDraft.objects.get()
        self.assertIsNotNone(draft.completed_at)
        self.assertEqual(draft.submission, CsoMappingSubmission.objects.get())
        self.assertEqual(draft.answers, {})  # personal data cleared after conversion
        # A completed draft cannot be resumed.
        self.assertEqual(
            self.client.get(_detail_url(), **_tok(token)).status_code, status.HTTP_404_NOT_FOUND
        )

    def test_repeated_final_submit_returns_same_receipt_no_duplicate(self):
        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        payload = build_valid_payload("cso")
        payload["client_submission_id"] = str(uuid.uuid4())
        first = self.client.post(_submit_url(), payload, format="json", **_tok(token))
        second = self.client.post(_submit_url(), payload, format="json", **_tok(token))
        self.assertEqual(first.status_code, status.HTTP_201_CREATED)
        self.assertEqual(second.status_code, status.HTTP_200_OK)
        self.assertEqual(first.data["reference"], second.data["reference"])
        self.assertEqual(CsoMappingSubmission.objects.count(), 1)

    def test_different_client_id_creates_separate_submission(self):
        for _ in range(2):
            token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
            payload = build_valid_payload("cso")
            payload["client_submission_id"] = str(uuid.uuid4())
            self.client.post(_submit_url(), payload, format="json", **_tok(token))
        self.assertEqual(CsoMappingSubmission.objects.count(), 2)

    def test_public_draft_flow_writes_no_audit_events(self):
        from audit.models import AuditEvent

        token = _create_draft(self.client, {"consent": "yes"}).data["resume_token"]
        self.client.put(
            _detail_url(),
            {"answers": {"consent": "yes", "respondent_type": "cso"}},
            format="json",
            **_tok(token),
        )
        payload = build_valid_payload("cso")
        payload["client_submission_id"] = str(uuid.uuid4())
        self.client.post(_submit_url(), payload, format="json", **_tok(token))
        # No personal questionnaire data is written to the audit stream by the
        # public flow (audit records only admin views/exports).
        self.assertEqual(AuditEvent.objects.count(), 0)

    def test_draft_endpoints_are_throttled(self):
        from .views import DraftCreateView, DraftDetailView, DraftSubmitView

        self.assertEqual(DraftCreateView.throttle_scope, "cso_mapping")
        self.assertEqual(DraftDetailView.throttle_scope, "cso_mapping_draft")
        self.assertEqual(DraftSubmitView.throttle_scope, "cso_mapping")


class DraftConfigCheckTests(APITestCase):
    """Section C: production must configure the draft retention (TTL) explicitly."""

    def test_check_errors_when_prod_ttl_missing(self):
        import os

        from .checks import draft_ttl_configured

        with override_settings(DEBUG=False, TESTING=False):
            saved = os.environ.pop("CSO_MAPPING_DRAFT_TTL_DAYS", None)
            try:
                errors = draft_ttl_configured(None)
                self.assertTrue(any(e.id == "cso_mapping.E001" for e in errors))
            finally:
                if saved is not None:
                    os.environ["CSO_MAPPING_DRAFT_TTL_DAYS"] = saved

    def test_check_passes_when_prod_ttl_set(self):
        import os

        from .checks import draft_ttl_configured

        with override_settings(DEBUG=False, TESTING=False):
            saved = os.environ.get("CSO_MAPPING_DRAFT_TTL_DAYS")
            os.environ["CSO_MAPPING_DRAFT_TTL_DAYS"] = "30"
            try:
                self.assertEqual(draft_ttl_configured(None), [])
            finally:
                if saved is None:
                    os.environ.pop("CSO_MAPPING_DRAFT_TTL_DAYS", None)
                else:
                    os.environ["CSO_MAPPING_DRAFT_TTL_DAYS"] = saved


class CleanupScheduleTests(APITestCase):
    """Section 2: scheduled cleanup — dry-run, failure exit, overlap protection."""

    def test_dry_run_deletes_nothing(self):
        draft = CsoMappingDraft.objects.create(
            token_hash="h1", expires_at=timezone.now() - timedelta(days=2)
        )
        call_command("cleanup_cso_drafts", "--dry-run")
        self.assertTrue(CsoMappingDraft.objects.filter(pk=draft.pk).exists())

    def test_only_expired_deleted_active_and_completed_kept(self):
        active = CsoMappingDraft.objects.create(
            token_hash="a1", expires_at=timezone.now() + timedelta(days=1)
        )
        expired = CsoMappingDraft.objects.create(
            token_hash="e1", expires_at=timezone.now() - timedelta(days=1)
        )
        sub = CsoMappingSubmission.objects.create(respondent_type="cso")
        call_command("cleanup_cso_drafts")
        self.assertTrue(CsoMappingDraft.objects.filter(pk=active.pk).exists())
        self.assertFalse(CsoMappingDraft.objects.filter(pk=expired.pk).exists())
        self.assertTrue(CsoMappingSubmission.objects.filter(pk=sub.pk).exists())

    def test_failure_exits_non_zero_without_leaking_data(self):
        from unittest.mock import patch

        from django.core.management.base import CommandError

        with patch("cso_mapping.models.CsoMappingDraft.objects") as manager:
            manager.filter.side_effect = RuntimeError("db down: secret@example.com")
            with self.assertRaises(CommandError) as ctx:
                call_command("cleanup_cso_drafts")
        # The raised message must not echo the underlying (potentially sensitive) text.
        self.assertNotIn("secret@example.com", str(ctx.exception))

    def test_cron_wrapper_uses_flock_and_logs_no_pii(self):
        import os

        from django.conf import settings

        wrapper = os.path.join(
            os.path.dirname(settings.BASE_DIR), "deploy", "cron", "cso-mapping-cleanup.sh"
        )
        with open(wrapper, encoding="utf-8") as fh:
            content = fh.read()
        self.assertIn("flock", content)  # overlap protection
        self.assertIn("cleanup_cso_drafts", content)
        # No PII field identifiers are referenced (would only appear if values
        # were being logged). Descriptive prose in comments is fine.
        for pii in ("respondent_email", "respondent_name", "responding_entity", "token_hash"):
            self.assertNotIn(pii, content)


class LocationCaptureTests(APITestCase):
    """Server-side validation of the required office GPS location."""

    def setUp(self):
        self.url = reverse("cso-mapping-submit")

    def _post(self, **overrides):
        payload = build_valid_payload("cso")
        for key, value in overrides.items():
            if value is None:
                payload.pop(key, None)
            else:
                payload[key] = value
        return self.client.post(self.url, payload, format="json")

    def test_valid_coordinates_are_accepted_and_stored(self):
        resp = self._post()
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        sub = CsoMappingSubmission.objects.get()
        self.assertEqual(str(sub.latitude), "-24.628200")
        self.assertEqual(str(sub.longitude), "25.923100")
        self.assertEqual(sub.location_capture_method, "device_gps")
        self.assertIsNotNone(sub.location_captured_at)
        self.assertFalse(sub.location_flagged)  # Gaborone is inside Botswana

    def test_missing_coordinates_are_rejected(self):
        resp = self._post(latitude=None, longitude=None)
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("cso_office_location", resp.data)
        self.assertEqual(CsoMappingSubmission.objects.count(), 0)

    def test_missing_only_longitude_is_rejected(self):
        resp = self._post(longitude=None)
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("cso_office_location", resp.data)

    def test_invalid_latitude_out_of_range_is_rejected(self):
        resp = self._post(latitude="120.0")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("cso_office_location", resp.data)

    def test_invalid_longitude_out_of_range_is_rejected(self):
        resp = self._post(longitude="500.0")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("cso_office_location", resp.data)

    def test_malformed_nan_and_infinite_coordinates_are_rejected(self):
        for bad in ("abc", "NaN", "Infinity", "-inf"):
            resp = self._post(latitude=bad)
            self.assertEqual(
                resp.status_code, status.HTTP_400_BAD_REQUEST, f"{bad!r} should be rejected"
            )
        self.assertEqual(CsoMappingSubmission.objects.count(), 0)

    def test_valid_but_outside_botswana_is_accepted_and_flagged(self):
        # Cape Town — a valid coordinate well outside Botswana's extent.
        resp = self._post(latitude="-33.924870", longitude="18.424055")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        sub = CsoMappingSubmission.objects.get()
        self.assertTrue(sub.location_flagged)
        self.assertTrue(sub.location_flag_reason)
        # The coordinate is never altered/replaced.
        self.assertEqual(str(sub.latitude), "-33.924870")


class MapLocationEndpointTests(APITestCase):
    """The read-only map feed + Excel/GeoJSON exports (authorised staff only)."""

    def setUp(self):
        self.admin = User.objects.create_user(
            username="mapadmin", email="mapadmin@example.com", password="x", role="admin"
        )
        self.officer = User.objects.create_user(
            username="mapofficer", email="mapofficer@example.com", password="x", role="officer"
        )
        # One mapped CSO (valid coords) ...
        self.mapped = CsoMappingSubmission.objects.create(
            consent=True,
            respondent_type="cso",
            responding_entity="Example Organisation",
            respondent_name="Jane Doe",
            respondent_phone="+267 555 0000",
            respondent_email="jane@example.org",
            primary_district="Kweneng District",
            latitude="-24.406600",
            longitude="25.495100",
            location_capture_method="device_gps",
            # Organisation type + physical address are read from answers for the
            # map (the direct "nature/type" question + the address field).
            answers={"annex2_a2_1a": "ngo", "physical_address": "Plot 123, Gaborone"},
        )
        # ... and one legacy record without coordinates (must be excluded).
        CsoMappingSubmission.objects.create(
            consent=True,
            respondent_type="cso",
            responding_entity="No Coordinates Org",
            primary_district="Central",
        )
        self.list_url = reverse("cso-mapping-locations")
        self.excel_url = reverse("cso-mapping-locations-export")
        self.geojson_url = reverse("cso-mapping-locations-geojson")

    def test_anonymous_denied(self):
        resp = self.client.get(self.list_url)
        self.assertIn(
            resp.status_code, (status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN)
        )

    def test_officer_without_module_denied(self):
        self.client.force_authenticate(self.officer)
        self.assertEqual(self.client.get(self.list_url).status_code, status.HTTP_403_FORBIDDEN)

    def test_admin_allowed_and_excludes_records_without_coordinates(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(self.list_url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(len(resp.data), 1)
        self.assertEqual(resp.data[0]["cso_name"], "Example Organisation")

    def test_list_returns_only_approved_fields(self):
        self.client.force_authenticate(self.admin)
        row = self.client.get(self.list_url).data[0]
        self.assertEqual(
            set(row.keys()),
            {
                "id",
                "cso_name",
                "organisation_type",
                "district",
                "village_town",
                "physical_address",
                "latitude",
                "longitude",
            },
        )
        # Organisation type is the labelled nature; address is passed through.
        self.assertEqual(
            row["organisation_type"], "Non-Governmental Organisation (NGO)"
        )
        self.assertEqual(row["physical_address"], "Plot 123, Gaborone")
        # No personal / confidential data leaks through the map feed.
        serialized = str(row)
        for leaked in ("jane@example.org", "+267 555 0000", "Jane Doe"):
            self.assertNotIn(leaked, serialized)

    def test_geojson_uses_longitude_before_latitude(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(self.geojson_url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = json.loads(resp.content)
        self.assertEqual(data["type"], "FeatureCollection")
        self.assertEqual(len(data["features"]), 1)
        feature = data["features"][0]
        self.assertEqual(feature["geometry"]["type"], "Point")
        lng, lat = feature["geometry"]["coordinates"]
        self.assertAlmostEqual(lng, 25.495100)  # longitude first
        self.assertAlmostEqual(lat, -24.406600)  # latitude second
        self.assertEqual(feature["properties"]["cso_name"], "Example Organisation")
        # No confidential fields in properties.
        self.assertNotIn("respondent_email", feature["properties"])

    def test_geojson_export_denied_to_officer(self):
        self.client.force_authenticate(self.officer)
        self.assertEqual(self.client.get(self.geojson_url).status_code, status.HTTP_403_FORBIDDEN)

    def test_excel_export_has_expected_headings_and_rows(self):
        from openpyxl import load_workbook

        self.client.force_authenticate(self.admin)
        resp = self.client.get(self.excel_url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["CSO Locations"]
        headers = [c.value for c in ws[1]]
        self.assertEqual(
            headers,
            [
                "CSO Name",
                "Organisation Type",
                "District",
                "Village/Town",
                "Physical Address",
                "Latitude",
                "Longitude",
            ],
        )
        # Header + exactly one mapped row (legacy no-coords excluded).
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws["A2"].value, "Example Organisation")
        self.assertEqual(ws["B2"].value, "Non-Governmental Organisation (NGO)")
        self.assertEqual(ws["C2"].value, "Kweneng District")


class MultiSelectQuestionTests(APITestCase):
    """select_multiple support: storage as lists, validation, and 'selected' relevance."""

    def setUp(self):
        self.url = reverse("cso-mapping-submit")

    def _multi_field(self):
        """A required select_multiple field active for a CSO submission."""
        schema = load_schema()
        ctx = {"consent": "yes", "respondent_type": "cso"}
        for section, field in iter_answerable_fields(schema):
            if (
                field["type"] == "select_multiple"
                and field.get("required")
                and field_is_active(section, field, ctx)
            ):
                return field
        raise AssertionError("no active required select_multiple field found")

    def test_multi_select_stored_as_list(self):
        field = self._multi_field()
        names = [c["name"] for c in field["choices"] if c["name"] != "other"][:2]
        payload = build_valid_payload("cso")
        payload[field["name"]] = names
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        sub = CsoMappingSubmission.objects.get()
        self.assertEqual(sub.answers[field["name"]], names)  # stored as a list

    def test_per_option_comment_is_stored(self):
        field = self._multi_field()
        opt = next(c["name"] for c in field["choices"] if c["name"] != "other")
        key = f"{field['name']}__comment__{opt}"
        payload = build_valid_payload("cso")
        payload[field["name"]] = [opt]
        payload[key] = "A note about this specific option"
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        sub = CsoMappingSubmission.objects.get()
        self.assertEqual(sub.answers[key], "A note about this specific option")

    def test_invalid_multi_select_option_rejected(self):
        field = self._multi_field()
        payload = build_valid_payload("cso")
        payload[field["name"]] = ["definitely_not_a_real_option"]
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn(field["name"], resp.data)

    def test_required_multi_select_empty_rejected(self):
        field = self._multi_field()
        payload = build_valid_payload("cso")
        payload[field["name"]] = []
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn(field["name"], resp.data)

    def test_other_specify_follows_selected_option(self):
        # annex2_a2_1e (services) is multi-select with an "other" option and an
        # annex2_a2_1e_other follow-up gated on selecting "other".
        payload = build_valid_payload("cso")
        payload["annex2_a2_1e"] = ["other"]
        payload["annex2_a2_1e_other"] = "A bespoke service"
        resp = self.client.post(self.url, payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        sub = CsoMappingSubmission.objects.get()
        self.assertEqual(sub.answers["annex2_a2_1e"], ["other"])
        self.assertEqual(sub.answers["annex2_a2_1e_other"], "A bespoke service")

    def test_export_joins_multi_select_labels(self):
        from openpyxl import load_workbook

        field = self._multi_field()
        picks = [c for c in field["choices"] if c["name"] != "other"][:2]
        payload = build_valid_payload("cso")
        payload[field["name"]] = [c["name"] for c in picks]
        self.assertEqual(
            self.client.post(self.url, payload, format="json").status_code,
            status.HTTP_201_CREATED,
        )
        admin = User.objects.create_user(
            username="msadmin", email="ms@example.com", password="x", role="admin"
        )
        self.client.force_authenticate(admin)
        resp = self.client.get(reverse("cso-submissions-export"))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        wb = load_workbook(BytesIO(resp.content))
        joined = ", ".join(c["label"] for c in picks)
        found = any(
            cell.value == joined
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
        )
        self.assertTrue(found, f"expected joined labels {joined!r} in the workbook")
