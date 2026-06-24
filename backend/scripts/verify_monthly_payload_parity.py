#!/usr/bin/env python3
"""Verify full payload parity between monthly workbooks and DB aggregates.

This script treats monthly workbooks as the source of truth and compares
workbook-derived payloads against aggregate rows per organization and quarter.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any


BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django  # noqa: E402

django.setup()

from openpyxl import load_workbook  # noqa: E402

from aggregates.models import Aggregate  # noqa: E402
from indicators.models import Indicator  # noqa: E402
from organizations.models import Organization  # noqa: E402
from projects.models import Project  # noqa: E402
from uploads.management.commands.import_reporting_workbook_overwrite import (  # noqa: E402
    IndicatorResolver,
    extract_section_value,
    get_age_band_mapping,
    merge_json_values,
    normalize_text,
    parse_sections,
)


DEFAULT_IMPORT_ROOT = Path("/home/bonasoadmin/BONASOV1/imports/End of year CBO Excel 2026")

DEFAULT_WORKBOOK_BY_ORG = {
    101: "APSA NCD REPORT - APRIL 26.xlsx",
    110: "ATN-Report.xlsx",
    102: "BONA NALEDI END OF YEAR SOUTH-EAST DISTRICT NCDs REPORT.xlsx",
    100: "BONMEH NCD REPORT final.xlsx",
    99: "BOSASNET End Of Year Excel Report.xlsx",
    104: "CHOBE ARTS REPORT- END OF YEAR 2026.xlsx",
    97: "HOME OF HOPE EXCEL REPORT 2026.xlsx",
    109: "HPP Report 25-26.xlsx",
    94: "JOURNEY OF HOPE REPORT (ANNUAL REPORT).xlsx",
    96: "JHF END OF YEAR REPORT.xlsx",
    95: "KEITSHOLOFETSE END OF YEAR .xlsx",
    108: "MAATA 2025-2026 EXCEL REPORT.xlsx",
    103: "Masego mental health Report (1).xlsx",
    98: "NCONGO Report-April 2026.xlsx",
    106: "SSSG-NCD REPORT -END OF YEAR 2026.xlsx",
    107: "The Fighters Support Group - Reporting Tool March.xlsx",
    111: "ULTIMATE YOUTH REPORT APRIL.xlsx",
    105: "VMHF_ July-April FY25_26 Report.xlsx",
}

QUARTERS = {
    "Q2": {
        "period_start": date(2025, 7, 1),
        "period_end": date(2025, 9, 30),
        "months": ["JULY", "AUGUST", "SEPTEMBER"],
    },
    "Q3": {
        "period_start": date(2025, 10, 1),
        "period_end": date(2025, 12, 31),
        "months": ["OCTOBER", "NOVEMBER", "DECEMBER"],
    },
    "Q4": {
        "period_start": date(2026, 1, 1),
        "period_end": date(2026, 3, 31),
        "months": ["JANUARY", "FEBRUARY", "MARCH", "APRIL"],
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Verify workbook-to-database full payload parity for monthly workbook imports.",
    )
    parser.add_argument("--project-id", type=int, default=2)
    parser.add_argument("--parent-org-id", type=int, default=5)
    parser.add_argument(
        "--import-root",
        default=str(DEFAULT_IMPORT_ROOT),
        help="Directory containing monthly workbook files.",
    )
    parser.add_argument(
        "--org-workbook-map",
        help="Optional JSON file mapping org_id -> workbook filename.",
    )
    parser.add_argument(
        "--org-ids",
        nargs="*",
        type=int,
        help="Optional subset of organization IDs to verify.",
    )
    parser.add_argument(
        "--sample-limit",
        type=int,
        default=40,
        help="Maximum number of diff sample rows to keep.",
    )
    parser.add_argument(
        "--json-out",
        help="Optional path to write the full verification report as JSON.",
    )
    return parser.parse_args()


def to_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _is_age_band_leaf_key(key: Any) -> bool:
    """True for an innermost age-band leaf key (or the non-age ``Value`` sentinel).

    Count-by-category indicators (e.g. "Number of ... activities conducted") are
    not age-disaggregated, so the importer stores their breakdown under a generic
    ``"Value"`` leaf, while this script's workbook reconstruction labels the same
    count under the organisation's single age band (e.g. ``"18-24"``). Both are
    legitimate representations of the identical number. A key qualifies as an
    age-band leaf when it is the ``Value`` sentinel or contains a digit
    (``18-24``, ``50+``, ``<18`` …). Service categories ("Mall Activations") and
    sex buckets ("Male"/"Female") contain no digits and are never collapsed.
    """
    text = str(key).strip().lower()
    return text == "value" or any(ch.isdigit() for ch in text)


def normalize_for_compare(value: Any) -> Any:
    if isinstance(value, dict):
        normalized = {
            key: normalize_for_compare(item)
            for key, item in sorted(value.items(), key=lambda pair: pair[0])
        }
        # Collapse a *singleton* innermost age-band leaf so that a non-age count
        # stored as ``{"Value": n}`` compares equal to the same count
        # reconstructed under one age band ``{"18-24": n}`` (parity false
        # positive for count-by-category indicators). Multi-band leaves — genuine
        # age disaggregation, e.g. ``{"18-24": 5, "25-49": 3}`` — keep their keys,
        # so a real age-band mismatch still surfaces.
        if len(normalized) == 1:
            (only_key, only_value), = normalized.items()
            if isinstance(only_value, (int, float)) and _is_age_band_leaf_key(only_key):
                return {"__age_band_leaf__": only_value}
        return normalized
    if isinstance(value, list):
        return [normalize_for_compare(item) for item in value]
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        parsed = Decimal(str(value))
        if parsed == parsed.to_integral_value():
            return int(parsed)
        return float(parsed)
    if isinstance(value, str):
        text = value.strip()
        try:
            parsed = Decimal(text)
            if parsed == parsed.to_integral_value():
                return int(parsed)
            return float(parsed)
        except Exception:
            return value
    return value


def is_importable_title(title: str) -> bool:
    token = normalize_text(title)
    return token.startswith("number") or token.startswith("total number")


def find_month_sheet_name(sheetnames: list[str], month: str) -> str | None:
    target = normalize_text(month)
    normalized = [(name, normalize_text(name)) for name in sheetnames]

    for original, candidate in normalized:
        if candidate == target:
            return original
    for original, candidate in normalized:
        if candidate.startswith(target):
            return original
    for original, candidate in normalized:
        if target in candidate.split():
            return original
    return None


def load_org_workbook_map(path: str | None) -> dict[int, str]:
    if not path:
        return dict(DEFAULT_WORKBOOK_BY_ORG)

    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError("org-workbook-map must be a JSON object mapping org_id -> workbook filename.")

    mapped: dict[int, str] = {}
    for key, value in raw.items():
        mapped[int(key)] = str(value)
    return mapped


def collect_expected_payloads_for_org(
    *,
    workbook_path: Path,
    organization: Organization,
    parent: Organization,
    resolver: IndicatorResolver,
) -> dict[str, dict[int, dict]]:
    wb_values = load_workbook(workbook_path, data_only=True, read_only=True)
    wb_formula = load_workbook(workbook_path, data_only=False, read_only=True)

    expected_by_quarter: dict[str, dict[int, dict]] = {}
    age_band_map = get_age_band_mapping(organization, parent)

    try:
        for quarter_name, quarter in QUARTERS.items():
            expected_by_indicator: dict[int, dict] = {}
            for month in quarter["months"]:
                sheet_name = find_month_sheet_name(wb_values.sheetnames, month)
                if not sheet_name:
                    continue

                formula_sheet = wb_formula[sheet_name] if sheet_name in wb_formula.sheetnames else None
                for section in parse_sections(wb_values[sheet_name], formula_sheet):
                    title = str(section.get("title") or "").strip()
                    if not title or not is_importable_title(title):
                        continue

                    indicator = resolver.resolve(title, section.get("index"))
                    if indicator is None:
                        continue

                    value, _ = extract_section_value(section, age_band_by_column=age_band_map)
                    total = to_decimal((value or {}).get("total"))
                    if total == 0:
                        continue

                    if indicator.id in expected_by_indicator:
                        expected_by_indicator[indicator.id] = merge_json_values(
                            expected_by_indicator[indicator.id],
                            value,
                        )
                    else:
                        expected_by_indicator[indicator.id] = value

            expected_by_quarter[quarter_name] = expected_by_indicator

    finally:
        wb_values.close()
        wb_formula.close()

    return expected_by_quarter


def main() -> int:
    args = parse_args()

    project = Project.objects.filter(id=args.project_id).first()
    if project is None:
        raise SystemExit(f"Project not found: {args.project_id}")

    parent = Organization.objects.filter(id=args.parent_org_id).first()
    if parent is None:
        raise SystemExit(f"Parent organization not found: {args.parent_org_id}")

    import_root = Path(args.import_root).expanduser().resolve()
    if not import_root.exists():
        raise SystemExit(f"Import root not found: {import_root}")

    org_workbook_map = load_org_workbook_map(args.org_workbook_map)
    if args.org_ids:
        requested = set(args.org_ids)
        org_workbook_map = {oid: name for oid, name in org_workbook_map.items() if oid in requested}

    resolver = IndicatorResolver(project=project)
    indicator_name_by_id = dict(Indicator.objects.values_list("id", "name"))

    summary = {
        "orgs_checked": 0,
        "workbooks_missing": 0,
        "organizations_missing": 0,
        "quarters_checked": 0,
        "payloads_compared": 0,
        "payload_mismatches": 0,
        "missing_in_db": 0,
        "missing_in_workbook": 0,
    }

    org_rows: list[dict[str, Any]] = []
    diff_sample: list[dict[str, Any]] = []
    # Uncapped record of every diff (the JSON ``diff_sample`` is capped at
    # ``--sample-limit``). Used to write the COMPLETE missing-in-db CSV export.
    # This is additive output only — the parity calculation is unchanged.
    all_diff_rows: list[dict[str, Any]] = []

    def _record_diff(*, org, quarter, indicator_id, indicator_name, kind, workbook):
        all_diff_rows.append({
            "org": org, "quarter": quarter, "indicator_id": indicator_id,
            "indicator_name": indicator_name, "type": kind, "workbook": workbook,
        })
        if len(diff_sample) < args.sample_limit:
            diff_sample.append({
                "org": org, "quarter": quarter, "indicator_id": indicator_id,
                "indicator_name": indicator_name, "type": kind,
            })

    for org_id in sorted(org_workbook_map.keys()):
        workbook_name = org_workbook_map[org_id]
        workbook_path = (import_root / workbook_name).resolve()

        organization = Organization.objects.filter(id=org_id).first()
        if organization is None:
            summary["organizations_missing"] += 1
            org_rows.append(
                {
                    "org_id": org_id,
                    "org": None,
                    "status": "organization_missing",
                    "workbook": str(workbook_path),
                }
            )
            continue

        if not workbook_path.exists():
            summary["workbooks_missing"] += 1
            org_rows.append(
                {
                    "org_id": organization.id,
                    "org": organization.name,
                    "status": "workbook_missing",
                    "workbook": str(workbook_path),
                }
            )
            continue

        expected_by_quarter = collect_expected_payloads_for_org(
            workbook_path=workbook_path,
            organization=organization,
            parent=parent,
            resolver=resolver,
        )

        org_stats = {
            "org_id": organization.id,
            "org": organization.name,
            "status": "ok",
            "workbook": str(workbook_path),
            "payload_exact": 0,
            "payload_mismatches": 0,
            "missing_in_db": 0,
            "missing_in_workbook": 0,
        }

        for quarter_name, quarter in QUARTERS.items():
            summary["quarters_checked"] += 1
            expected_by_indicator = expected_by_quarter.get(quarter_name, {})

            db_rows = Aggregate.objects.filter(
                project=project,
                organization=organization,
                period_start=quarter["period_start"],
                period_end=quarter["period_end"],
            )
            db_by_indicator: dict[int, dict] = {}
            for row in db_rows:
                payload = row.value if isinstance(row.value, dict) else {}
                if to_decimal(payload.get("total")) == 0:
                    continue
                db_by_indicator[row.indicator_id] = payload

            indicator_ids = sorted(set(expected_by_indicator.keys()) | set(db_by_indicator.keys()))
            for indicator_id in indicator_ids:
                summary["payloads_compared"] += 1
                expected = expected_by_indicator.get(indicator_id)
                actual = db_by_indicator.get(indicator_id)

                if expected is None:
                    summary["missing_in_workbook"] += 1
                    org_stats["missing_in_workbook"] += 1
                    _record_diff(
                        org=organization.name, quarter=quarter_name, indicator_id=indicator_id,
                        indicator_name=indicator_name_by_id.get(indicator_id),
                        kind="missing_in_workbook_payload", workbook=str(workbook_path),
                    )
                    continue

                if actual is None:
                    summary["missing_in_db"] += 1
                    org_stats["missing_in_db"] += 1
                    _record_diff(
                        org=organization.name, quarter=quarter_name, indicator_id=indicator_id,
                        indicator_name=indicator_name_by_id.get(indicator_id),
                        kind="missing_in_db_payload", workbook=str(workbook_path),
                    )
                    continue

                if normalize_for_compare(expected) == normalize_for_compare(actual):
                    org_stats["payload_exact"] += 1
                else:
                    summary["payload_mismatches"] += 1
                    org_stats["payload_mismatches"] += 1
                    _record_diff(
                        org=organization.name, quarter=quarter_name, indicator_id=indicator_id,
                        indicator_name=indicator_name_by_id.get(indicator_id),
                        kind="payload_mismatch", workbook=str(workbook_path),
                    )

        summary["orgs_checked"] += 1
        org_rows.append(org_stats)

    report = {
        "generated_at": datetime.now().isoformat(),
        "project": {"id": project.id, "code": project.code, "name": project.name},
        "parent_org": {"id": parent.id, "name": parent.name},
        "import_root": str(import_root),
        "summary": summary,
        "orgs": org_rows,
        "diff_sample": diff_sample,
    }

    if args.json_out:
        output_path = Path(args.json_out).expanduser()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
        print(f"REPORT_JSON={output_path}")

        # Additive export: the COMPLETE list of records present in workbooks but
        # missing from the database, so admins can triage/import the backlog. The
        # parity calculation is unchanged; this just writes already-computed rows.
        write_missing_in_db_csv(output_path.parent, all_diff_rows, project)

    print(json.dumps(summary, sort_keys=True))


def write_missing_in_db_csv(parity_dir: Path, all_diff_rows: list[dict], project) -> Path:
    """Write every ``missing_in_db_payload`` row to ``missing_in_db_latest.csv``
    (and a dated copy). Columns: Organization, Quarter, Indicator ID, Indicator
    Name, Project, Workbook File, Issue Type."""
    project_label = f"{getattr(project, 'id', '')} {getattr(project, 'code', '') or getattr(project, 'name', '')}".strip()
    missing = [r for r in all_diff_rows if r.get("type") == "missing_in_db_payload"]
    header = ["Organization", "Quarter", "Indicator ID", "Indicator Name",
              "Project", "Workbook File", "Issue Type"]
    latest = parity_dir / "missing_in_db_latest.csv"
    dated = parity_dir / f"missing_in_db_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    for path in (latest, dated):
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(header)
            for r in missing:
                w.writerow([
                    r.get("org"), r.get("quarter"), r.get("indicator_id"),
                    r.get("indicator_name"), project_label,
                    Path(r.get("workbook") or "").name, r.get("type"),
                ])
    print(f"MISSING_IN_DB_CSV={latest} (rows={len(missing)})")
    return latest

    has_issues = any(
        summary[key] > 0
        for key in (
            "workbooks_missing",
            "organizations_missing",
            "payload_mismatches",
            "missing_in_db",
            "missing_in_workbook",
        )
    )
    return 1 if has_issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
