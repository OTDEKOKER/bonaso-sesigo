"""
Load NAHPA Social Contracting 2026/27 (project id 3) coordinator indicators &
quarterly targets from the authoritative annual workbook
  frontend/docs/04-06-26 ANNUAL_CSO_TARGETS_SOCIAL_CONTRACTING  YEAR 26-27 (1).xlsx

One sheet per coordinator. Column H = indicator name, J = definition,
A-D = Q1-Q4 targets. Indicators are SHARED across coordinators: each workbook
row is matched to an existing indicator (reuse) or a new one is created, then a
coordinator-scoped ProjectIndicatorOrganizationTarget + ProjectIndicatorAssignment
is upserted and the ProjectIndicator quarterly totals are rolled up.

Matching: exact normalized name -> reuse; an explicit OVERRIDE_REUSE / FORCE_CREATE
map handles borderline cases; fuzzy >= 0.95 -> reuse; otherwise create new.

Dry-run by default. Pass --apply to persist. Writes a JSON report.
"""
import argparse
import difflib
import json
import os
import re
import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

import openpyxl

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
import django  # noqa: E402

django.setup()

from django.db import transaction  # noqa: E402
from django.db.models import Sum  # noqa: E402

from indicators.models import Indicator  # noqa: E402
from projects.models import (  # noqa: E402
    Project,
    ProjectIndicator,
    ProjectIndicatorAssignment,
    ProjectIndicatorOrganizationTarget,
    ProjectOrganization,
)

PROJECT_ID = 3
WORKBOOK = str(
    BACKEND_ROOT.parent
    / "frontend/docs/04-06-26 ANNUAL_CSO_TARGETS_SOCIAL_CONTRACTING  YEAR 26-27 (1).xlsx"
)

# sheet -> (coordinator organization id, label)
SHEET_ORG = {
    "MEN AND BOYS": (166, "MBGE"),
    "MAKGABANENG": (5, "MAKGABANENG"),
    "BONEPWA": (112, "BONEPWA+"),
    "HPP": (109, "HPP"),
    "TEBELOPELE": (1, "TEBELOPELE"),
    "MOPIPI": (159, "MOPIPI"),
}

HEADER_TOKENS = {
    "numerator", "denominator", "indicator", "activity", "definition",
    "dissagregation", "m", "f", "age", "location", "gen pop", "",
}

# Curated reuse: normalized workbook name -> existing indicator id.
# Covers method/phrasing/population variants of an existing shared indicator
# and corrects a handful of misleading high-ratio fuzzy hits.
OVERRIDE_REUSE = {
    # syndromic-method wording = same screening indicator everyone else reports
    "number of individuals syndromic screened for gbv": 343,
    "number of people syndromic screened for stis": 351,
    # PLHIV/PLWH wording for the TB indicator BONEPWA already reports (I525)
    "number of people living with hiv who tested positive for tb and are on treatment": 525,
    # MOPIPI "refresher training" CSO row -> shared CSO capacity indicator (I319)
    "number of csos who recieved refresher training to provide comprehensive psychosocial and practical support including hiv related support": 319,
    # MBGE reworded "STI referrals cases completed" -> shared I470
    "number of sti referrals cases completed": 470,
    # service-provider induction/training variants -> shared I370
    "number of service providers community mobilisers m e officers program officers receiving training for the project": 370,
    "number of service providers receiving training community mobilisers m e officers program officers": 370,
    "number of service providers receiving training community mobilisers m e officers program officers stakeholders": 370,
    "number of service providers receiving training counsellors m e officers": 370,
    # MOPIPI "braille condoms distributed" -> brailed condoms (I448), NOT male condoms
    "number of braille condoms distributed": 448,
    # repeat condom collection -> the track project 3 already uses (I356), not
    # the near-identical duplicates I480/I543, to avoid a parallel target track
    "number of people who reported collecting condoms for the repeated time by age and sex": 356,
    "number of people who reported collecting condoms for a repeated time": 356,
    # lubricants -> the track project 3 already uses (I357), not the near-identical I607
    "number of lubricants distributed": 357,
    # mentorship / quality-report "per quarter" wording -> shared monthly indicators
    "number of sub recipients mentored per quarter": 318,
    "number of sub recipients submitting quality reports per quarter it should be timely complete accurate realistic": 320,
    # 0.90-0.95 fuzzy hits that ARE correct (pin them so they never drift to create)
    "number of csos sensitized to provide comprehensive psychosocial and practical support including hiv related support": 319,
    "number of people referred for sti services": 550,
}

# Genuinely new indicators (distinct unit/population/concept) — always create,
# even if a fuzzy match exists.
FORCE_CREATE = {
    "number of counsellors sensitized to provide comprehensive psychosocial and practical support including hiv related support",
    "number of centres submitting quality reports per quarter",
    "number of kps taking up services",
    "number of pwds taking up services",
    "number of radio drama episodes produced and aired",
    "number of condoms distributed to plwh",
}

FUZZY_REUSE_THRESHOLD = 0.95

CATEGORY_KEYWORDS = [
    ("gbv", "gbv"),
    ("sti", "sti"),
    ("ncd", "ncd"),
    ("cancer", "ncd"),
    ("diabetes", "ncd"),
    ("mental health", "mental_health"),
    ("counselling", "mental_health"),
    ("radio", "media"),
    ("media", "media"),
    ("commemorat", "events"),
    ("trained", "trainings"),
    ("training", "trainings"),
    ("sensitiz", "trainings"),
    ("sensitis", "trainings"),
]


def norm(s):
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", str(s or "").lower())).strip()


def to_dec(v):
    if v is None:
        return Decimal(0)
    s = str(v).strip().replace("`", "").replace(",", "").replace("%", "")
    if s in ("", "-"):
        return Decimal(0)
    try:
        return Decimal(s)
    except Exception:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return Decimal(m.group(0)) if m else Decimal(0)


def infer_category(name):
    nn = norm(name)
    for kw, cat in CATEGORY_KEYWORDS:
        if kw in nn:
            return cat
    return "hiv_prevention"


def make_code(name, existing_codes):
    base = "SC2627-" + re.sub(r"[^A-Z0-9]+", "-", name.upper()).strip("-")[:40]
    code = base
    n = 1
    while code in existing_codes or Indicator.objects.filter(code=code).exists():
        n += 1
        code = f"{base[:46]}-{n}"
    existing_codes.add(code)
    return code


def parse_workbook():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    rows = []
    for sheet, (oid, label) in SHEET_ORG.items():
        ws = wb[sheet]
        last = max(
            (r for r in range(1, ws.max_row + 1)
             if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1))),
            default=0,
        )
        for r in range(6, last + 1):
            h = ws.cell(r, 8).value
            name = str(h).strip() if h is not None else ""
            if not name or norm(name) in HEADER_TOKENS:
                continue
            q = [to_dec(ws.cell(r, c).value) for c in range(1, 5)]
            defn = ws.cell(r, 10).value
            rows.append({
                "sheet": sheet, "oid": oid, "label": label, "row": r,
                "name": name, "norm": norm(name),
                "defn": str(defn).strip() if defn else "",
                "q": q,
            })
    return rows


def build_matcher():
    inds = [
        (i, n) for i, n in Indicator.objects
        .exclude(name__istartswith="DEMO").exclude(code__startswith="DEMO-")
        .exclude(is_deprecated=True).values_list("id", "name")
    ]
    bynorm = {}
    for i, n in inds:
        bynorm.setdefault(norm(n), i)

    def resolve(nn):
        # returns (indicator_id or None, method, score)
        if nn in FORCE_CREATE:
            return None, "force_create", 0.0
        if nn in OVERRIDE_REUSE:
            return OVERRIDE_REUSE[nn], "override_reuse", 1.0
        if nn in bynorm:
            return bynorm[nn], "exact", 1.0
        bid, bs = None, -1.0
        for i, n in inds:
            s = difflib.SequenceMatcher(None, nn, norm(n)).ratio()
            if s > bs:
                bid, bs = i, s
        if bs >= FUZZY_REUSE_THRESHOLD:
            return bid, "fuzzy", bs
        return None, "create", bs

    return resolve


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--cleanup", action="store_true",
                    help="also delete coordinator org-targets not in the workbook (zero-data only)")
    ap.add_argument("--report", default=str(BACKEND_ROOT / "reports" / "sc2627_coord_indicators_targets.json"))
    args = ap.parse_args()

    project = Project.objects.get(id=PROJECT_ID)
    rows = parse_workbook()
    resolve = build_matcher()
    coord_po = {
        oid: ProjectOrganization.objects.filter(project=project, organization_id=oid).first()
        for _, (oid, _) in SHEET_ORG.items()
    }
    existing_codes = set(Indicator.objects.values_list("code", flat=True))

    # Decide an indicator per unique normalized name (shared across coordinators).
    name_decision = {}
    for r in rows:
        name_decision.setdefault(r["norm"], {"name": r["name"], "defn": r["defn"]})
    decisions = {}
    for nn, info in name_decision.items():
        iid, method, score = resolve(nn)
        decisions[nn] = {"indicator_id": iid, "method": method, "score": round(score, 3),
                         "name": info["name"], "defn": info["defn"]}

    report = {
        "project": {"id": project.id, "code": project.code, "name": project.name},
        "apply": args.apply,
        "rows_parsed": len(rows),
        "unique_indicator_names": len(decisions),
        "created_indicators": [],
        "reused_indicators": [],
        "summary": defaultdict(int),
        "duplicate_rows_in_sheet": [],
        "per_coordinator": defaultdict(lambda: {"targets_created": 0, "targets_updated": 0, "targets_unchanged": 0, "indicators": 0}),
        "existing_targets_not_in_workbook": [],
        "override_keys_unused": [],
    }

    used_override = set()
    used_force = set()
    touched_pi = set()
    seen_pair = set()  # (oid, indicator_id) duplicate detection within run

    with transaction.atomic():
        # 1. Resolve / create indicators
        for nn, d in decisions.items():
            if d["method"] == "override_reuse":
                used_override.add(nn)
            if d["method"] == "force_create":
                used_force.add(nn)
            if d["indicator_id"] is None:
                ind = Indicator(
                    name=d["name"][:255],
                    code=make_code(d["name"], existing_codes),
                    description=d["defn"],
                    type="number",
                    category=infer_category(d["name"]),
                    aggregation_method="sum",
                    is_active=True,
                )
                ind.save()
                d["indicator_id"] = ind.id
                report["created_indicators"].append({"id": ind.id, "code": ind.code, "name": ind.name, "category": ind.category, "method": d["method"], "best_score": d["score"]})
                report["summary"]["indicators_created"] += 1
            else:
                report["reused_indicators"].append({"id": d["indicator_id"], "method": d["method"], "score": d["score"], "workbook_name": d["name"]})
                report["summary"][f"reuse_{d['method']}"] += 1

        # 2. Upsert ProjectIndicator + org target + assignment per coordinator row
        for r in rows:
            iid = decisions[r["norm"]]["indicator_id"]
            oid = r["oid"]
            pair = (oid, iid)
            if pair in seen_pair:
                report["duplicate_rows_in_sheet"].append({"sheet": r["sheet"], "row": r["row"], "indicator_id": iid, "name": r["name"]})
            seen_pair.add(pair)

            pi, _ = ProjectIndicator.objects.get_or_create(project=project, indicator_id=iid)
            touched_pi.add(pi.id)

            project.organizations.add(oid)
            Indicator.objects.get(id=iid).organizations.add(oid)

            q1, q2, q3, q4 = r["q"]
            tv = q1 + q2 + q3 + q4
            ot = ProjectIndicatorOrganizationTarget.objects.filter(project_indicator=pi, organization_id=oid).first()
            if ot is None:
                ProjectIndicatorOrganizationTarget.objects.create(
                    project_indicator=pi, organization_id=oid,
                    q1_target=q1, q2_target=q2, q3_target=q3, q4_target=q4, target_value=tv)
                report["per_coordinator"][r["label"]]["targets_created"] += 1
            elif (ot.q1_target, ot.q2_target, ot.q3_target, ot.q4_target) == (q1, q2, q3, q4):
                report["per_coordinator"][r["label"]]["targets_unchanged"] += 1
            else:
                ot.q1_target, ot.q2_target, ot.q3_target, ot.q4_target, ot.target_value = q1, q2, q3, q4, tv
                ot.save(update_fields=["q1_target", "q2_target", "q3_target", "q4_target", "target_value"])
                report["per_coordinator"][r["label"]]["targets_updated"] += 1
            report["per_coordinator"][r["label"]]["indicators"] += 1

            ProjectIndicatorAssignment.objects.update_or_create(
                project_indicator=pi, organization_id=oid,
                defaults=dict(assignment_source="organization_target", is_active=True,
                              project_organization=coord_po.get(oid)))

        # 3. Roll up ProjectIndicator quarterly totals from org targets
        for pi_id in touched_pi:
            agg = ProjectIndicatorOrganizationTarget.objects.filter(project_indicator_id=pi_id).aggregate(
                q1=Sum("q1_target"), q2=Sum("q2_target"), q3=Sum("q3_target"), q4=Sum("q4_target"), tv=Sum("target_value"))
            ProjectIndicator.objects.filter(id=pi_id).update(
                q1_target=agg["q1"] or 0, q2_target=agg["q2"] or 0, q3_target=agg["q3"] or 0,
                q4_target=agg["q4"] or 0, target_value=agg["tv"] or 0)

        # 4. Existing coordinator targets NOT covered by the workbook.
        #    Report them; with --cleanup, delete the zero-data ones (superseded
        #    leftovers from the earlier rough load) and their assignments.
        from aggregates.models import Aggregate  # noqa: E402
        wb_pairs = {(r["oid"], decisions[r["norm"]]["indicator_id"]) for r in rows}
        cleanup_pi = set()
        for ot in ProjectIndicatorOrganizationTarget.objects.filter(
            project_indicator__project=project,
            organization_id__in=[oid for _, (oid, _) in SHEET_ORG.items()],
        ).select_related("project_indicator__indicator"):
            if (ot.organization_id, ot.project_indicator.indicator_id) in wb_pairs:
                continue
            iid = ot.project_indicator.indicator_id
            has_data = bool(ot.current_value) or bool(ot.baseline_value) or Aggregate.objects.filter(
                project=project, indicator_id=iid, organization_id=ot.organization_id).exists()
            entry = {
                "organization_id": ot.organization_id,
                "indicator_id": iid,
                "indicator_name": ot.project_indicator.indicator.name,
                "has_data": has_data,
            }
            report["existing_targets_not_in_workbook"].append(entry)
            if args.cleanup and not has_data:
                ProjectIndicatorAssignment.objects.filter(
                    project_indicator=ot.project_indicator, organization_id=ot.organization_id).delete()
                cleanup_pi.add(ot.project_indicator_id)
                ot.delete()
                report["summary"]["leftover_targets_deleted"] += 1
            elif args.cleanup and has_data:
                report["summary"]["leftover_targets_kept_has_data"] += 1

        # Re-roll-up any project indicators whose org targets changed in cleanup
        for pi_id in cleanup_pi:
            agg = ProjectIndicatorOrganizationTarget.objects.filter(project_indicator_id=pi_id).aggregate(
                q1=Sum("q1_target"), q2=Sum("q2_target"), q3=Sum("q3_target"), q4=Sum("q4_target"), tv=Sum("target_value"))
            ProjectIndicator.objects.filter(id=pi_id).update(
                q1_target=agg["q1"] or 0, q2_target=agg["q2"] or 0, q3_target=agg["q3"] or 0,
                q4_target=agg["q4"] or 0, target_value=agg["tv"] or 0)

        report["override_keys_unused"] = sorted(set(OVERRIDE_REUSE) - used_override)
        report["summary"]["force_create_keys_unused"] = len(set(FORCE_CREATE) - used_force)
        report["summary"]["project_indicators_touched"] = len(touched_pi)

        if not args.apply:
            transaction.set_rollback(True)

    report["summary"] = dict(report["summary"])
    report["per_coordinator"] = {k: v for k, v in report["per_coordinator"].items()}
    Path(args.report).parent.mkdir(parents=True, exist_ok=True)
    Path(args.report).write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")

    print(("APPLIED" if args.apply else "DRY-RUN") + " ----------------------------------------")
    print(json.dumps(report["summary"], indent=2))
    print("per coordinator:", json.dumps(report["per_coordinator"], indent=2))
    print(f"created indicators: {len(report['created_indicators'])}")
    for c in report["created_indicators"]:
        print(f"  + I? {c['code']}  [{c['category']}]  {c['name'][:60]!r}  (best fuzzy {c['best_score']})")
    if report["duplicate_rows_in_sheet"]:
        print("DUP rows (same indicator twice for one coord):", report["duplicate_rows_in_sheet"])
    if report["override_keys_unused"]:
        print("WARN unused OVERRIDE_REUSE keys:", report["override_keys_unused"])
    print(f"existing coord targets NOT in workbook: {len(report['existing_targets_not_in_workbook'])}")
    print(f"report -> {args.report}")


if __name__ == "__main__":
    main()
