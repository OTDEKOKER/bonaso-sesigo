"""
Re-store project 3 percentage/rate coordinator targets as percent-numbers and
mark their indicators type='percentage'.

A workbook row is a percentage/rate (not a count) when its ANNUAL target (col E)
parses to <= 1, or any quarter/annual cell literally contains '%'. For those
rows the quarterly fractions are stored x100 (0.014->1.4, 0.8->80, 1.0->100) so
the 2-decimal target field stops rounding sub-1% values, and the rate reads
naturally against the denominator (col I).

The indicator a row maps to is found by the loader's resolver, then disambiguated
to whichever same-named indicator actually holds the project-3 org target for
that coordinator (handles duplicate-named indicators like the two "referred for PEP").

Dry-run by default; --apply to persist. Flags any indicator that is a percentage
in one sheet but a count in another (not auto-flipped).
"""
import argparse
import os
import re
import sys
from decimal import Decimal
from pathlib import Path

import openpyxl

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
import django  # noqa: E402

django.setup()

from django.db import transaction  # noqa: E402
from django.db.models import Sum  # noqa: E402

from indicators.models import Indicator  # noqa: E402
from projects.models import ProjectIndicator, ProjectIndicatorOrganizationTarget as OT  # noqa: E402

import coordinator_indicators_targets_2627 as L  # noqa: E402

PID = 3


def parse_cell(v):
    """raw cell -> (Decimal fraction, had_percent_sign)"""
    if v is None:
        return Decimal(0), False
    s = str(v).strip().replace("`", "").replace(",", "")
    pct = "%" in s
    s = s.replace("%", "")
    if s in ("", "-"):
        return Decimal(0), pct
    try:
        return Decimal(s), pct
    except Exception:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return (Decimal(m.group(0)) if m else Decimal(0)), pct


def parse_rows_with_annual():
    wb = openpyxl.load_workbook(L.WORKBOOK, data_only=True)
    rows = []
    for sheet, (oid, label) in L.SHEET_ORG.items():
        ws = wb[sheet]
        last = max((r for r in range(1, ws.max_row + 1)
                    if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1))), default=0)
        for r in range(6, last + 1):
            h = ws.cell(r, 8).value
            name = str(h).strip() if h is not None else ""
            if not name or L.norm(name) in L.HEADER_TOKENS:
                continue
            q = [parse_cell(ws.cell(r, c).value) for c in range(1, 5)]      # A-D (num, had_%)
            annual, ann_pct = parse_cell(ws.cell(r, 5).value)               # E
            had_pct = ann_pct or any(p for _, p in q)
            # fraction form of annual for classification (100% -> 1.0)
            annual_frac = (annual / 100) if ann_pct else annual
            rows.append({
                "sheet": sheet, "oid": oid, "label": label, "row": r,
                "name": name, "norm": L.norm(name),
                "q": q, "annual_frac": annual_frac, "had_pct": had_pct,
                "denom": str(ws.cell(r, 9).value or "").strip(),
            })
    return rows


def percent_value(num, had_pct):
    """workbook cell -> percent-number. 100% -> 100 ; 0.014 -> 1.4 ; 1.0 -> 100."""
    return num if had_pct else num * 100


def is_percentage(row):
    if row["had_pct"]:
        return True
    a = row["annual_frac"]
    if a is not None and Decimal(0) < a <= Decimal(1):
        return True
    return False


def find_target(nn, oid):
    """resolve indicator + locate the project-3 org target actually holding it."""
    iid, method, score = L.build_matcher()(nn)
    if iid is None:
        # post-load these exist; match by exact normalized name among all indicators
        for i, n in Indicator.objects.values_list("id", "name"):
            if L.norm(n) == nn:
                iid = i
                break
    if iid is None:
        return None, None
    ot = OT.objects.filter(project_indicator__project_id=PID, project_indicator__indicator_id=iid, organization_id=oid).first()
    if ot:
        return iid, ot
    # disambiguate duplicate-named indicators: same name, but the one with the target
    target_name = L.norm(Indicator.objects.get(id=iid).name)
    for i, n in Indicator.objects.values_list("id", "name"):
        if L.norm(n) == target_name and i != iid:
            ot = OT.objects.filter(project_indicator__project_id=PID, project_indicator__indicator_id=i, organization_id=oid).first()
            if ot:
                return i, ot
    return iid, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    rows = parse_rows_with_annual()

    # classify
    pct_ind_ids = set()
    count_ind_ids = set()
    plan = []   # (row, iid, ot)
    by_ind = {}  # iid -> list of (sheet,row,name,is_pct)
    for r in rows:
        iid, ot = find_target(r["norm"], r["oid"])
        if iid is None:
            continue
        # blank/zero-target rows carry no count-vs-percent signal — skip them
        if not r["had_pct"] and (r["annual_frac"] in (None, Decimal(0))) and all(n == 0 for n, _ in r["q"]):
            continue
        p = is_percentage(r)
        by_ind.setdefault(iid, []).append((r["label"], r["row"], r["name"][:35], p))
        if p:
            pct_ind_ids.add(iid)
            plan.append((r, iid, ot))
        else:
            count_ind_ids.add(iid)

    conflicts = pct_ind_ids & count_ind_ids

    updated_targets = 0
    typed = 0
    touched_pi = set()
    examples = []
    missing = []
    with transaction.atomic():
        for r, iid, ot in plan:
            if iid in conflicts:
                continue
            if ot is None:
                missing.append((r["label"], r["row"], r["name"][:40], iid))
                continue
            new = [percent_value(num, pct) for num, pct in r["q"]]
            tv = sum(new)
            old = [ot.q1_target, ot.q2_target, ot.q3_target, ot.q4_target]
            if [Decimal(x) for x in old] != new:
                if args.apply:
                    ot.q1_target, ot.q2_target, ot.q3_target, ot.q4_target, ot.target_value = new[0], new[1], new[2], new[3], tv
                    ot.save(update_fields=["q1_target", "q2_target", "q3_target", "q4_target", "target_value"])
                updated_targets += 1
                touched_pi.add(ot.project_indicator_id)
                if len(examples) < 25:
                    examples.append((r["label"], r["name"][:34], [float(n) for n, _ in r["q"]], [float(x) for x in new]))

        flip_ids = pct_ind_ids - conflicts
        for iid in flip_ids:
            ind = Indicator.objects.get(id=iid)
            if ind.type != "percentage":
                if args.apply:
                    ind.type = "percentage"
                    ind.save(update_fields=["type", "updated_at"])
                typed += 1

        for pi_id in touched_pi:
            agg = OT.objects.filter(project_indicator_id=pi_id).aggregate(
                q1=Sum("q1_target"), q2=Sum("q2_target"), q3=Sum("q3_target"), q4=Sum("q4_target"), tv=Sum("target_value"))
            if args.apply:
                ProjectIndicator.objects.filter(id=pi_id).update(
                    q1_target=agg["q1"] or 0, q2_target=agg["q2"] or 0, q3_target=agg["q3"] or 0,
                    q4_target=agg["q4"] or 0, target_value=agg["tv"] or 0)

        if not args.apply:
            transaction.set_rollback(True)

    print(("APPLIED" if args.apply else "DRY-RUN") + " ----------------------------------")
    print(f"percentage rows: {len(plan)} | indicators flipped to 'percentage': {typed} | targets re-stored x100: {updated_targets}")
    if conflicts:
        print(f"CONFLICT indicators (percentage in one sheet, count in another) — left unchanged: {sorted(conflicts)}")
        for cid in sorted(conflicts):
            print("   I%s %r" % (cid, Indicator.objects.get(id=cid).name[:55]))
            for lbl, rr, nm, p in by_ind.get(cid, []):
                print("       %-10s r%-3s %-37s %s" % (lbl, rr, nm, "PCT" if p else "count"))
    if missing:
        print("rows with no stored target (skipped):")
        for m in missing:
            print("  ", m)
    print("examples (coord | indicator | wb fraction -> stored percent):")
    for e in examples:
        print("  ", e[0], "|", e[1], "|", e[2], "->", e[3])


if __name__ == "__main__":
    main()
