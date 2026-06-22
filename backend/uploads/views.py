import io
import json
import os
import re
import subprocess
import sys
from pathlib import Path

from django.conf import settings
from django.http import FileResponse, HttpResponse
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.response import Response
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Upload, ImportJob, ExportJob
from .serializers import UploadSerializer, ImportJobSerializer, ExportJobSerializer
from .jobs import resolve_report_workbook_import_script, run_aggregate_review_import_job
from organizations.access import is_training_only_request
from users.permissions import HasModulePermission
from aggregates import reporting_workbook as agg_rw


REPORT_WORKBOOK_IMPORT_SCRIPT = "import_selected_q3_workbook.py"
NON_REPORT_SHEET_KEYWORDS = (
    "indicator matrix",
    "summary",
    "instruction",
    "instructions",
    "cover",
    "contents",
    "readme",
    "notes",
)


def _is_truthy(value):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _current_fiscal_quarter():
    """(quarter, fiscal_start_year) for today on the Botswana Apr–Mar fiscal year."""
    from datetime import date
    today = date.today()
    month, year = today.month, today.year
    if 4 <= month <= 6:
        return 1, year
    if 7 <= month <= 9:
        return 2, year
    if 10 <= month <= 12:
        return 3, year
    return 4, year - 1


def _is_sesigo_reporting_workbook(path) -> bool:
    """True when an uploaded file is a SESIGO reporting workbook (carries the
    hidden Metadata + _cellmap sheets). Used to route uploads to the donor-style
    smart importer instead of the legacy fuzzy NAHPA template importer."""
    try:
        workbook = load_workbook(path, read_only=True)
    except Exception:
        return False
    try:
        names = set(workbook.sheetnames)
    finally:
        workbook.close()
    return {agg_rw.SHEET_META, agg_rw.SHEET_CELLMAP}.issubset(names)


def _parse_reporting_period_range(label):
    match = re.match(
        r"^Q([1-4])\s+(\d{4})(?:\s*/\s*(\d{2}|\d{4}))?$",
        str(label or "").strip(),
        flags=re.IGNORECASE,
    )
    if not match:
        return None

    quarter = int(match.group(1))
    year = int(match.group(2))
    next_year = year + 1

    if quarter == 1:
        return f"{year}-04-01", f"{year}-06-30"
    if quarter == 2:
        return f"{year}-07-01", f"{year}-09-30"
    if quarter == 3:
        return f"{year}-10-01", f"{year}-12-31"
    if quarter == 4:
        return f"{next_year}-01-01", f"{next_year}-03-31"
    return None


def _looks_like_non_reporting_sheet(sheet_name):
    normalized = " ".join(str(sheet_name or "").replace("_", " ").split()).lower()
    return any(keyword in normalized for keyword in NON_REPORT_SHEET_KEYWORDS)


def _resolve_report_workbook_import_script():
    return resolve_report_workbook_import_script()


def _get_importable_sheet_names(upload, provided_sheet_names=None):
    sheet_names = []

    if isinstance(provided_sheet_names, (list, tuple)):
        for sheet_name in provided_sheet_names:
            text = str(sheet_name or "").strip()
            if text and not _looks_like_non_reporting_sheet(text) and text not in sheet_names:
                sheet_names.append(text)
        if sheet_names:
            return sheet_names

    workbook = load_workbook(upload.file.path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            if _looks_like_non_reporting_sheet(sheet_name):
                continue
            if sheet_name not in sheet_names:
                sheet_names.append(sheet_name)
    finally:
        workbook.close()

    return sheet_names


def _coerce_mapping_payload(raw_value):
    if isinstance(raw_value, dict):
        return raw_value
    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            return {}
    return {}


def _launch_import_job_worker(job):
    log_dir = Path(settings.BASE_DIR) / "reports" / "import-jobs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"import-job-{job.id}.log"
    env = os.environ.copy()
    env["BONASO_DJANGO_ROOT"] = str(settings.BASE_DIR)
    command = [sys.executable, str(Path(settings.BASE_DIR) / "manage.py"), "process_import_job", str(job.id)]
    with log_path.open("ab") as log_file:
        subprocess.Popen(
            command,
            cwd=str(settings.BASE_DIR),
            env=env,
            stdout=log_file,
            stderr=subprocess.STDOUT,
            start_new_session=True,
        )
    result = job.result if isinstance(job.result, dict) else {}
    result["log_path"] = str(log_path)
    result["worker_started_at"] = timezone.now().isoformat()
    job.result = result
    job.save(update_fields=["result"])
    return log_path


def _launch_export_job_worker(job):
    log_dir = Path(settings.BASE_DIR) / "reports" / "export-jobs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"export-job-{job.id}.log"
    env = os.environ.copy()
    env["BONASO_DJANGO_ROOT"] = str(settings.BASE_DIR)
    command = [sys.executable, str(Path(settings.BASE_DIR) / "manage.py"), "process_export_job", str(job.id)]
    with log_path.open("ab") as log_file:
        subprocess.Popen(
            command,
            cwd=str(settings.BASE_DIR),
            env=env,
            stdout=log_file,
            stderr=subprocess.STDOUT,
            start_new_session=True,
        )
    result = job.result if isinstance(job.result, dict) else {}
    result["log_path"] = str(log_path)
    result["worker_started_at"] = timezone.now().isoformat()
    job.result = result
    job.save(update_fields=["result"])
    return log_path


class UploadViewSet(viewsets.ModelViewSet):
    """ViewSet for managing uploads."""
    
    queryset = Upload.objects.all()
    serializer_class = UploadSerializer
    required_module = 'uploads'
    permission_classes = [IsAuthenticated, HasModulePermission]
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['file_type', 'organization', 'content_type']
    
    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_staff or user.role == 'admin':
            return Upload.objects.all()
        elif user.organization:
            return Upload.objects.filter(organization=user.organization)
        return Upload.objects.filter(created_by=user)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def start_import(self, request, pk=None):
        """Start an import job for the uploaded file."""
        upload = self.get_object()
        queue_aggregate_review = _is_truthy(request.data.get("queue_aggregate_review"))
        dry_run = _is_truthy(request.data.get("dry_run"))

        # Smart routing (Phase 3): a SESIGO reporting workbook carries its own
        # project/organization/quarter metadata, so it bypasses the legacy fuzzy
        # importer and is processed by the donor-style smart importer. Existing
        # (flat / multi-org NAHPA) uploads fall through to the original pipeline.
        try:
            if _is_sesigo_reporting_workbook(upload.file.path):
                return self._import_sesigo_reporting_workbook(request, upload, dry_run=dry_run)
        except Exception:
            pass  # On any detection error, fall back to the legacy pipeline.

        job_status = "pending" if queue_aggregate_review else "ready_for_review"
        job = ImportJob.objects.create(
            upload=upload,
            status=job_status,
            job_type="aggregate_review_import" if queue_aggregate_review else "upload_import",
            created_by=request.user
        )

        if not queue_aggregate_review:
            return Response(ImportJobSerializer(job).data, status=status.HTTP_201_CREATED)

        project_id = request.data.get("project_id") or request.data.get("project")
        try:
            project_id = int(project_id)
        except (TypeError, ValueError):
            project_id = None
        reporting_period = request.data.get("reporting_period") or request.data.get("period_label")
        period_start = request.data.get("period_start")
        period_end = request.data.get("period_end")
        if reporting_period and (not period_start or not period_end):
            parsed_period = _parse_reporting_period_range(reporting_period)
            if parsed_period:
                period_start, period_end = parsed_period

        if not project_id:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": "project_id required for aggregate review queueing"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "project_id required for aggregate review queueing"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Training/live project mismatch guard — enforce at the backend,
        # regardless of frontend routing. A training-mode request must only
        # import into a training project, and a live request must never write
        # into a training project.
        try:
            from projects.models import Project as _Project
            _target_project = _Project.objects.only('id', 'is_training').get(pk=project_id)
            _caller_is_training = is_training_only_request(request)
            if _caller_is_training and not _target_project.is_training:
                return Response(
                    {"error": "Training-mode requests can only import into training projects."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            if not _caller_is_training and _target_project.is_training:
                return Response(
                    {"error": "Live requests cannot import into a training project."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        except Exception:
            pass  # Project lookup failure handled later in the import pipeline

        if not period_start or not period_end:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": "reporting_period or period_start/period_end required"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "reporting_period or period_start/period_end required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            sheet_names = _get_importable_sheet_names(upload, request.data.get("sheet_names"))
        except Exception as exc:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": str(exc)}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": f"Unable to inspect workbook sheets: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not sheet_names:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": "no importable organization sheets found"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "No importable organization sheets found in workbook"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        script_path = _resolve_report_workbook_import_script()
        if not script_path.exists():
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": f"Import script not found: {script_path}"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": f"Import script not found: {script_path}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        report_path = Path(settings.BASE_DIR) / "reports" / "aggregate-review-queue" / f"upload-{upload.id}-job-{job.id}.json"
        indicator_overrides_path = None
        indicator_overrides = _coerce_mapping_payload(request.data.get("indicator_overrides"))
        if indicator_overrides:
            indicator_overrides_path = report_path.with_suffix(".indicator-overrides.json")
            indicator_overrides_path.parent.mkdir(parents=True, exist_ok=True)
            indicator_overrides_path.write_text(
                json.dumps(indicator_overrides, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

        sheet_org_overrides_path = None
        sheet_org_overrides = _coerce_mapping_payload(request.data.get("sheet_org_overrides"))
        if sheet_org_overrides:
            sheet_org_overrides_path = report_path.with_suffix(".sheet-org-overrides.json")
            sheet_org_overrides_path.parent.mkdir(parents=True, exist_ok=True)
            sheet_org_overrides_path.write_text(
                json.dumps(sheet_org_overrides, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

        job.parameters = {
            "project_id": project_id,
            "period_start": str(period_start),
            "period_end": str(period_end),
            "reporting_period": str(reporting_period or ""),
            "sheet_names": sheet_names,
            "dry_run": dry_run,
            "report_path": str(report_path),
            "indicator_overrides_path": str(indicator_overrides_path) if indicator_overrides_path else "",
            "sheet_org_overrides_path": str(sheet_org_overrides_path) if sheet_org_overrides_path else "",
        }
        job.output_file = str(report_path)
        job.save(update_fields=["parameters", "output_file"])

        if dry_run:
            job = run_aggregate_review_import_job(job.id)
            payload = ImportJobSerializer(job).data
            payload.update(job.result or {})
            return Response(payload, status=status.HTTP_201_CREATED)

        try:
            log_path = _launch_import_job_worker(job)
        except Exception as exc:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": f"Unable to start background import worker: {exc}"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "Unable to start background import worker", "details": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        payload = ImportJobSerializer(job).data
        payload["background_job_started"] = True
        payload["dry_run"] = dry_run
        payload["log_path"] = str(log_path)
        return Response(payload, status=status.HTTP_202_ACCEPTED)

    def _import_sesigo_reporting_workbook(self, request, upload, *, dry_run=False):
        """Process a SESIGO reporting workbook upload by reusing the aggregates
        smart-import endpoint (single source of truth for parsing + permissions
        + aggregate writes), recording the outcome on an ImportJob."""
        from rest_framework.test import APIRequestFactory, force_authenticate
        from aggregates.views import AggregateViewSet

        job = ImportJob.objects.create(
            upload=upload,
            status="processing",
            job_type="reporting_workbook_import",
            created_by=request.user,
            started_at=timezone.now(),
        )

        path = "/api/aggregates/import-reporting-workbook/"
        if is_training_only_request(request):
            path += "?training_only=true"
        file_handle = open(upload.file.path, "rb")
        try:
            data = {"file": file_handle}
            if dry_run:
                data["dry_run"] = "true"
            internal = APIRequestFactory().post(path, data=data, format="multipart")
            force_authenticate(internal, user=request.user)
            response = AggregateViewSet.as_view({"post": "import_reporting_workbook"})(internal)
        finally:
            file_handle.close()

        result = response.data if hasattr(response, "data") else {}
        if response.status_code >= 400:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [result]
            job.result = result
            job.save(update_fields=["status", "completed_at", "errors", "result"])
            return Response(result, status=response.status_code)

        summary = result.get("summary", {}) if isinstance(result, dict) else {}
        job.total_rows = int(summary.get("indicators_found", 0) or 0)
        job.processed_rows = job.total_rows
        job.successful_rows = int(summary.get("indicators_valid", 0) or 0)
        job.failed_rows = int(summary.get("indicators_failed", 0) or 0)
        job.status = "validated" if dry_run else "imported"
        job.completed_at = timezone.now()

        # Surface repeated uploads of the same file (idempotency safeguard
        # IMP-1). The import itself is already idempotent (update_or_create on
        # the natural key), so this is informational — it lets the UI warn that
        # the file was processed before and explains an all-"unchanged" result.
        prior = upload.prior_imported_upload()
        if isinstance(result, dict) and prior is not None:
            result = {
                **result,
                "duplicate_file": True,
                "previous_upload_id": prior.id,
                "previous_upload_name": prior.name,
            }

        job.result = result
        job.errors = []
        job.save(update_fields=[
            "total_rows", "processed_rows", "successful_rows", "failed_rows",
            "status", "completed_at", "result", "errors",
        ])
        payload = ImportJobSerializer(job).data
        payload.update(result if isinstance(result, dict) else {})
        return Response(payload, status=response.status_code)

    @action(detail=False, methods=['get'], url_path='download_template')
    def download_template(self, request):
        """Generate and stream a reporting workbook for a project/organization.

        When an organization is selected this returns the donor-style SESIGO
        reporting workbook (delegating to the aggregates generator — the single
        source of truth for assignment/target/disaggregation resolution and
        permission/training enforcement). With no organization it falls back to
        the legacy flat batch-entry template for backward compatibility."""
        from projects.models import Project, ProjectIndicator, ProjectOrganization

        project_id = request.query_params.get('project')
        organization_id = request.query_params.get('organization')

        if not project_id:
            return Response({'detail': 'project query param required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Donor reporting workbook (Phase 1): reuse the aggregates generator.
        if organization_id:
            from rest_framework.test import APIRequestFactory, force_authenticate
            from aggregates.views import AggregateViewSet

            query = {'project': project_id, 'organization': organization_id}
            quarter = request.query_params.get('quarter')
            fiscal_year = request.query_params.get('fiscal_year')
            if quarter:
                query['quarter'] = quarter
                if fiscal_year:
                    query['fiscal_year'] = fiscal_year
            else:
                cq, cfy = _current_fiscal_quarter()
                query['quarter'] = f"Q{cq}"
                query['fiscal_year'] = cfy
            if _is_truthy(request.query_params.get('with_data')):
                query['with_data'] = 'true'
            if is_training_only_request(request):
                query['training_only'] = 'true'
            internal = APIRequestFactory().get("/api/aggregates/reporting-workbook/", query)
            force_authenticate(internal, user=request.user)
            return AggregateViewSet.as_view({"get": "reporting_workbook"})(internal)

        try:
            project = Project.objects.get(pk=project_id)
        except Project.DoesNotExist:
            return Response({'detail': 'Project not found.'}, status=status.HTTP_404_NOT_FOUND)

        pi_qs = ProjectIndicator.objects.filter(project=project).select_related('indicator').order_by('indicator__name')
        if organization_id:
            assigned_indicator_ids = set(
                project.projectindicatorassignment_set.filter(
                    organization_id=organization_id, is_active=True
                ).values_list('project_indicator__indicator_id', flat=True)
            ) if hasattr(project, 'projectindicatorassignment_set') else set()
            if assigned_indicator_ids:
                pi_qs = pi_qs.filter(indicator_id__in=assigned_indicator_ids)

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Entry"

        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", wrap_text=True)

        fixed_headers = ["Organization Code", "Organization Name", "Reporting Period", "Quarter"]
        for col, header in enumerate(fixed_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = 20

        indicator_headers = []
        for pi in pi_qs:
            ind = pi.indicator
            label = f"{ind.code} – {ind.name}" if ind.code else ind.name
            indicator_headers.append(label)

        for col_offset, label in enumerate(indicator_headers):
            col = len(fixed_headers) + 1 + col_offset
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = 30

        ws.row_dimensions[1].height = 40

        # Instruction sheet
        ws_info = wb.create_sheet(title="Instructions")
        ws_info["A1"] = "Batch Data Entry Template"
        ws_info["A1"].font = Font(bold=True, size=14)
        ws_info["A3"] = f"Project: {project.name} ({project.code})"
        ws_info["A4"] = "Fill in the 'Data Entry' sheet. Do not rename or delete columns."
        ws_info["A5"] = "Reporting Period format: YYYY-QN (e.g. 2025-Q4)"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"batch_template_{project.code or project_id}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ImportJobViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing import jobs."""
    
    queryset = ImportJob.objects.all()
    serializer_class = ImportJobSerializer
    # Import jobs are part of the uploads module; honour an explicit deny.
    # (get_queryset already restricts to the user's own jobs.)
    required_module = 'uploads'
    permission_classes = [IsAuthenticated, HasModulePermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'upload']
    
    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_staff or user.role == 'admin':
            return ImportJob.objects.all()
        return ImportJob.objects.filter(created_by=user)

    @action(detail=True, methods=['post'])
    def retry(self, request, pk=None):
        job = self.get_object()
        if job.job_type != "aggregate_review_import":
            return Response({"error": "Only aggregate review import jobs can be retried."}, status=status.HTTP_400_BAD_REQUEST)
        if job.status not in {"failed", "pending", "processing"}:
            return Response({"error": "Only failed or stuck import jobs can be retried."}, status=status.HTTP_400_BAD_REQUEST)
        if not job.parameters:
            return Response({"error": "This import job has no saved parameters to retry."}, status=status.HTTP_400_BAD_REQUEST)

        job.status = "pending"
        job.total_rows = 0
        job.processed_rows = 0
        job.successful_rows = 0
        job.failed_rows = 0
        job.started_at = None
        job.completed_at = None
        job.errors = []
        job.result = {
            **(job.result if isinstance(job.result, dict) else {}),
            "retried_at": timezone.now().isoformat(),
            "retry_requested_by": getattr(request.user, "id", None),
        }
        job.save(
            update_fields=[
                "status",
                "total_rows",
                "processed_rows",
                "successful_rows",
                "failed_rows",
                "started_at",
                "completed_at",
                "errors",
                "result",
            ]
        )

        try:
            _launch_import_job_worker(job)
        except Exception as exc:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": f"Unable to restart background import worker: {exc}"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "Unable to restart background import worker", "details": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        payload = ImportJobSerializer(job, context={"request": request}).data
        payload["background_job_started"] = True
        return Response(payload, status=status.HTTP_202_ACCEPTED)


class ExportJobViewSet(viewsets.ModelViewSet):
    """Create and view export jobs."""

    queryset = ExportJob.objects.all()
    serializer_class = ExportJobSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'job_type']

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_staff or user.role == 'admin':
            return ExportJob.objects.all()
        return ExportJob.objects.filter(created_by=user)

    def create(self, request, *args, **kwargs):
        job_type = str(request.data.get("job_type") or "aggregate_export")
        if job_type != "aggregate_export":
            return Response({"error": "Unsupported export job type."}, status=status.HTTP_400_BAD_REQUEST)

        parameters = request.data.get("parameters") or {}
        if not isinstance(parameters, dict):
            return Response({"error": "parameters must be an object."}, status=status.HTTP_400_BAD_REQUEST)

        from organizations.access import request_mode_value

        job = ExportJob.objects.create(
            job_type=job_type,
            status="pending",
            parameters=parameters,
            # Bind the export to the caller's environment so the tokenless
            # background worker re-enters the same training/live mode (audit S-1).
            mode=request_mode_value(request),
            created_by=request.user,
        )

        try:
            log_path = _launch_export_job_worker(job)
        except Exception as exc:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": f"Unable to start background export worker: {exc}"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "Unable to start background export worker", "details": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        payload = ExportJobSerializer(job, context={"request": request}).data
        payload["background_job_started"] = True
        payload["log_path"] = str(log_path)
        return Response(payload, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def retry(self, request, pk=None):
        job = self.get_object()
        if job.job_type != "aggregate_export":
            return Response({"error": "Only aggregate export jobs can be retried."}, status=status.HTTP_400_BAD_REQUEST)
        if job.status not in {"failed", "pending", "processing"}:
            return Response({"error": "Only failed or stuck export jobs can be retried."}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(job.parameters, dict):
            return Response({"error": "This export job has no saved parameters to retry."}, status=status.HTTP_400_BAD_REQUEST)

        job.status = "pending"
        job.started_at = None
        job.completed_at = None
        job.output_file = ""
        job.file_name = ""
        job.content_type = ""
        job.errors = []
        job.result = {
            **(job.result if isinstance(job.result, dict) else {}),
            "retried_at": timezone.now().isoformat(),
            "retry_requested_by": getattr(request.user, "id", None),
        }
        job.save(
            update_fields=[
                "status",
                "started_at",
                "completed_at",
                "output_file",
                "file_name",
                "content_type",
                "errors",
                "result",
            ]
        )

        try:
            _launch_export_job_worker(job)
        except Exception as exc:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": f"Unable to restart background export worker: {exc}"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "Unable to restart background export worker", "details": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        payload = ExportJobSerializer(job, context={"request": request}).data
        payload["background_job_started"] = True
        return Response(payload, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        job = self.get_object()
        if job.status != "completed" or not job.output_file:
            return Response({"error": "Export is not ready for download."}, status=status.HTTP_400_BAD_REQUEST)
        path = Path(job.output_file)
        if not path.exists():
            return Response({"error": "Export file is missing."}, status=status.HTTP_404_NOT_FOUND)
        response = FileResponse(path.open("rb"), content_type=job.content_type or "application/octet-stream")
        response["Content-Disposition"] = f'attachment; filename="{job.file_name or path.name}"'
        return response

