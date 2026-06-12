"""Tests for the native Excel chart export (analysis/chart_export.py)."""
from io import BytesIO

from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from openpyxl import load_workbook

from analysis.chart_export import build_chart_workbook
from analysis.chart_theme import SESIGO_CHART_PALETTE, normalize_hex, series_color

User = get_user_model()

SPEC = {
    "title": "People reached",
    "subtitle": "NAHPA Social Contracting | 3 periods",
    "chart_type": "column",
    "categories": ["10-14", "15-19", "20-24"],
    "series": [
        {"name": "Male", "color": "#20A3D3", "values": [3327, 3359, 937]},
        {"name": "Female", "color": "#0FA546", "values": [3968, 3481, 2063]},
    ],
    "source_filters": {"Project": "NAHPA Social Contracting", "Coordinator": "MAKGABANENG"},
}


class ChartThemeTests(TestCase):
    def test_palette_matches_frontend_order(self):
        # Guard: the first colours must stay aligned with frontend/lib/chart-theme.ts.
        self.assertEqual(SESIGO_CHART_PALETTE[0], "#20A3D3")
        self.assertEqual(SESIGO_CHART_PALETTE[1], "#0FA546")

    def test_normalize_hex_strips_and_falls_back(self):
        self.assertEqual(normalize_hex("#20A3D3"), "20A3D3")
        self.assertEqual(normalize_hex("20a3d3"), "20A3D3")
        # Unresolved CSS var falls back to the first palette colour, never random.
        self.assertEqual(normalize_hex("hsl(var(--chart-2))"), "20A3D3")
        self.assertEqual(normalize_hex(None), "20A3D3")

    def test_series_color_cycles(self):
        self.assertEqual(series_color(0), "20A3D3")
        self.assertEqual(series_color(len(SESIGO_CHART_PALETTE)), "20A3D3")  # wraps


class ChartWorkbookTests(TestCase):
    def _wb(self, spec=None):
        return load_workbook(BytesIO(build_chart_workbook(spec or SPEC)))

    def test_sheets_present(self):
        wb = self._wb()
        self.assertEqual(wb.sheetnames, ["Cover", "Metadata", "Summary", "Pivot", "Charts", "Raw Data", "Data Dictionary"])

    def test_chart_sheet_has_native_chart(self):
        wb = self._wb()
        self.assertEqual(len(wb["Charts"]._charts), 1)

    def test_series_colours_match_spec(self):
        wb = self._wb()
        chart = wb["Charts"]._charts[0]

        def _hex(fill):
            # openpyxl stores solidFill either as a plain 'RRGGBB' str or a
            # ColorChoice whose .srgbClr is itself a plain str.
            if isinstance(fill, str):
                return fill
            srgb = fill.srgbClr
            return srgb if isinstance(srgb, str) else srgb.value

        fills = [_hex(s.graphicalProperties.solidFill) for s in chart.series]
        self.assertEqual(fills, ["20A3D3", "0FA546"])

    def test_pivot_totals_correct(self):
        ws = self._wb()["Pivot"]
        # Total row is the last row; grand total in the final column.
        last = [c.value for c in ws[ws.max_row]]
        self.assertEqual(last[0], "Total")
        self.assertEqual(last[-1], 3327 + 3359 + 937 + 3968 + 3481 + 2063)

    def test_raw_data_is_long_format(self):
        ws = self._wb()["Raw Data"]
        # header + (3 categories x 2 series) = 7 rows
        self.assertEqual(ws.max_row, 7)
        self.assertEqual([c.value for c in ws[2]], ["10-14", "Male", 3327])

    def test_chart_has_data_table_with_keys(self):
        # Donor house style: a native data table with legend colour keys below
        # the plot. openpyxl can't read dTable back, so assert on the chart XML.
        import zipfile

        raw = build_chart_workbook(SPEC)
        z = zipfile.ZipFile(BytesIO(raw))
        chart_xml = next(
            z.read(n).decode() for n in z.namelist()
            if "chart" in n and n.endswith(".xml")
        )
        self.assertIn("dTable", chart_xml)
        self.assertIn("showKeys", chart_xml)
        self.assertIn('barDir val="col"', chart_xml)
        self.assertIn('grouping val="clustered"', chart_xml)

    def test_pie_colours_each_slice(self):
        spec = dict(SPEC, chart_type="pie", series=[SPEC["series"][0]])
        chart = self._wb(spec)["Charts"]._charts[0]
        self.assertEqual(len(chart.series[0].data_points), 3)

    def test_cover_carries_environment_stamp(self):
        ws = self._wb(dict(SPEC, environment="TRAINING", exported_by="alice"))["Cover"]
        self.assertEqual(ws["A1"].value, "TRAINING ENVIRONMENT")
        # exporter shown somewhere on the cover
        values = [c.value for row in ws.iter_rows(max_col=2) for c in row]
        self.assertIn("alice", values)

    def test_data_dictionary_present(self):
        ws = self._wb()["Data Dictionary"]
        labels = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
        self.assertTrue({"Cover", "Charts", "Raw Data", "Environment"} <= labels)


class ChartExportEndpointTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(username="ce_admin", password="x", is_superuser=True, is_staff=True)
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def test_endpoint_returns_xlsx(self):
        resp = self.client.post("/api/analysis/chart-export/", SPEC, format="json")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn(".xlsx", resp["Content-Disposition"])
        wb = load_workbook(BytesIO(b"".join(resp.streaming_content) if resp.streaming else resp.content))
        self.assertIn("Charts", wb.sheetnames)
        self.assertIn("Cover", wb.sheetnames)

    def test_endpoint_rejects_empty(self):
        resp = self.client.post("/api/analysis/chart-export/", {"title": "x", "categories": [], "series": []}, format="json")
        self.assertEqual(resp.status_code, 400)

    def test_endpoint_requires_auth(self):
        client = APIClient()
        resp = client.post("/api/analysis/chart-export/", SPEC, format="json")
        self.assertIn(resp.status_code, (401, 403))

    def test_endpoint_stamps_environment_from_token(self):
        from rest_framework_simplejwt.tokens import AccessToken
        # Training-bound token -> the cover must be stamped TRAINING, server-side.
        token = AccessToken.for_user(self.admin)
        token["mode"] = "training"
        self.client.force_authenticate(self.admin, token=token)
        resp = self.client.post("/api/analysis/chart-export/", SPEC, format="json")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("training", resp["Content-Disposition"])
        wb = load_workbook(BytesIO(resp.content))
        self.assertEqual(wb["Cover"]["A1"].value, "TRAINING ENVIRONMENT")
