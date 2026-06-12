"""
Native Excel chart export for Sesigo analysis charts.

Builds an .xlsx that mirrors a portal chart as closely as Excel allows: a title,
the source-filter context (project/coordinator/organization/period), the
chart-ready data table, and an actual editable Excel chart whose series colours
come from the shared palette (analysis/chart_theme.py), so the download matches
the on-screen chart instead of re-colouring randomly.

The chart spec is produced by the frontend (frontend/lib/chart-export.ts) using
the same palette, then POSTed here. Keeping colour resolution on both ends from
one ordered list guarantees a series keeps its colour in the UI and the file.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .chart_theme import normalize_hex, series_color

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
_TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="0A2B73")
_SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
_META_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="595959")
_META_VALUE_FONT = Font(name="Calibri", size=10, color="262626")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="0A2B73")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_VALID_TYPES = {"bar", "column", "line", "stacked-bar", "stacked-column", "pie"}


def _coerce_number(value: Any):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return value


def _apply_series_color(series: Series, hex_color: str, *, line: bool) -> None:
    """Colour a chart series from a hex value (fill for bars, stroke for lines)."""
    color = normalize_hex(hex_color)
    gp = GraphicalProperties()
    if line:
        gp.line = LineProperties(solidFill=color, w=28575)  # ~2.25pt
        series.graphicalProperties = gp
        series.marker = Marker(symbol="circle", size=5)
        if series.marker.graphicalProperties is None:
            series.marker.graphicalProperties = GraphicalProperties()
        series.marker.graphicalProperties.solidFill = color
        series.smooth = False
    else:
        gp.solidFill = color
        gp.line = LineProperties(solidFill=color)
        series.graphicalProperties = gp


def build_chart_workbook(spec: dict[str, Any]) -> bytes:
    """Render a chart spec to a donor/M&E-grade .xlsx workbook.

    Sheets (in order): Cover, Metadata, Summary, Pivot, Charts, Raw Data, Data
    Dictionary. Every sheet carries the environment stamp (LIVE/TRAINING) so a
    training-mode export can never be mistaken for live reporting, plus who
    exported it and when.

    spec = {
        "title": str,
        "subtitle": str | None,
        "chart_type": "bar"|"column"|"line"|"stacked-bar"|"stacked-column"|"pie",
        "categories": [str, ...],
        "series": [{"name": str, "color": "#RRGGBB"|None, "values": [num, ...]}],
        "source_filters": {"Project": "...", "Period": "...", ...},
        "value_axis_title": str | None,
        "category_axis_title": str | None,
        # injected server-side by the view (not trusted from the client):
        "environment": "LIVE"|"TRAINING",
        "exported_by": str,
        "exported_at": str (ISO),
    }
    """
    title = str(spec.get("title") or "Chart").strip() or "Chart"
    subtitle = (spec.get("subtitle") or "").strip()
    chart_type = str(spec.get("chart_type") or "bar").strip().lower()
    if chart_type not in _VALID_TYPES:
        chart_type = "bar"
    categories = [str(c) for c in (spec.get("categories") or [])]
    raw_series = spec.get("series") or []
    source_filters = spec.get("source_filters") or {}
    environment = str(spec.get("environment") or "LIVE").strip().upper()
    if environment not in {"LIVE", "TRAINING"}:
        environment = "LIVE"
    exported_by = str(spec.get("exported_by") or "—")
    exported_at = str(spec.get("exported_at") or "")

    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    _write_cover_sheet(cover, title, subtitle, environment, exported_by, exported_at, source_filters)
    _write_metadata_sheet(
        wb.create_sheet("Metadata"), title, subtitle, chart_type, environment,
        exported_by, exported_at, source_filters, categories, raw_series,
    )
    _write_summary_sheet(wb.create_sheet("Summary"), title, subtitle, source_filters, categories, raw_series)
    _write_pivot_sheet(wb.create_sheet("Pivot"), title, categories, raw_series)
    _write_chart_sheet(wb.create_sheet("Charts"), title, subtitle, chart_type, categories, raw_series, source_filters, spec)
    _write_raw_sheet(wb.create_sheet("Raw Data"), categories, raw_series)
    _write_data_dictionary_sheet(wb.create_sheet("Data Dictionary"))

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# Environment stamp colours.
_ENV_FILL = {"LIVE": "0FA546", "TRAINING": "F1A100"}


def _write_cover_sheet(ws, title, subtitle, environment, exported_by, exported_at, source_filters):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    badge = ws.cell(row=1, column=1, value=f"{environment} ENVIRONMENT")
    badge.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    badge.fill = PatternFill("solid", fgColor=_ENV_FILL.get(environment, "0FA546"))
    badge.alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")
    if environment == "TRAINING":
        warn = ws.cell(row=2, column=1, value="Training data — NOT for official reporting.")
        warn.font = Font(italic=True, bold=True, color="B45309")
        ws.merge_cells("A2:B2")

    ws["A4"] = title
    ws["A4"].font = _TITLE_FONT
    ws.merge_cells("A4:B4")
    if subtitle:
        ws["A5"] = subtitle
        ws["A5"].font = _SUBTITLE_FONT
        ws.merge_cells("A5:B5")

    rows = [("Sesigo Data Portal", ""), ("Exported by", exported_by), ("Exported at", exported_at)]
    r = 7
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = _META_LABEL_FONT
        ws.cell(row=r, column=2, value=str(value)).font = _META_VALUE_FONT
        r += 1
    if source_filters:
        r += 1
        ws.cell(row=r, column=1, value="Filters applied").font = _META_LABEL_FONT
        r += 1
        for label, value in source_filters.items():
            if value in (None, "", []):
                continue
            ws.cell(row=r, column=1, value=str(label)).font = _META_LABEL_FONT
            ws.cell(row=r, column=2, value=str(value)).font = _META_VALUE_FONT
            r += 1


def _write_metadata_sheet(ws, title, subtitle, chart_type, environment, exported_by,
                          exported_at, source_filters, categories, raw_series):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60
    _style_header(ws.cell(row=1, column=1, value="Field"))
    _style_header(ws.cell(row=1, column=2, value="Value"))
    meta = [
        ("Report title", title),
        ("Subtitle", subtitle),
        ("Environment", environment),
        ("Exported by", exported_by),
        ("Exported at", exported_at),
        ("Chart type", chart_type),
        ("Categories", len(categories)),
        ("Series", len(raw_series)),
        ("Series names", ", ".join(str(s.get("name") or "") for s in raw_series)),
        ("Source system", "Sesigo Data Portal (powered by BONASO)"),
    ]
    for label, value in source_filters.items():
        meta.append((f"Filter — {label}", value))
    r = 2
    for label, value in meta:
        ws.cell(row=r, column=1, value=str(label)).border = _BORDER
        ws.cell(row=r, column=2, value="" if value in (None,) else str(value)).border = _BORDER
        r += 1


def _write_data_dictionary_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80
    _style_header(ws.cell(row=1, column=1, value="Sheet / Column"))
    _style_header(ws.cell(row=1, column=2, value="Description"))
    entries = [
        ("Cover", "Title, environment stamp (LIVE/TRAINING), exporter and timestamp, and the filters applied."),
        ("Metadata", "Machine-readable export context: title, environment, user, time, chart type, series."),
        ("Summary", "Grand total plus per-series and per-category totals."),
        ("Pivot", "Category × series matrix with a Total column and a Total row."),
        ("Charts", "The editable Excel chart with a colour-keyed data table beneath it."),
        ("Raw Data", "Tidy long-format rows — one row per (Category, Series, Value) — for re-analysis."),
        ("— Category", "The grouping shown on the chart's category axis (e.g. organization, age band)."),
        ("— Series", "A measured dimension/series name (e.g. a message type or sex)."),
        ("— Value", "The numeric value for that category and series."),
        ("Environment", "LIVE = production reporting data. TRAINING = practice data, not for official use."),
    ]
    r = 2
    for label, value in entries:
        c = ws.cell(row=r, column=1, value=label)
        c.font = _META_LABEL_FONT
        c.border = _BORDER
        v = ws.cell(row=r, column=2, value=value)
        v.alignment = Alignment(wrap_text=True, vertical="top")
        v.border = _BORDER
        r += 1


def _write_chart_sheet(ws, title, subtitle, chart_type, categories, raw_series, source_filters, spec):
    ws.sheet_view.showGridLines = False

    # --- Title + subtitle ---------------------------------------------------
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    row = 2
    if subtitle:
        ws[f"A{row}"] = subtitle
        ws[f"A{row}"].font = _SUBTITLE_FONT
        row += 1

    # --- Source filters block ----------------------------------------------
    row += 1
    if source_filters:
        ws[f"A{row}"] = "Filters applied"
        ws[f"A{row}"].font = _META_LABEL_FONT
        row += 1
        for label, value in source_filters.items():
            if value in (None, "", []):
                continue
            ws[f"A{row}"] = str(label)
            ws[f"A{row}"].font = _META_LABEL_FONT
            ws[f"B{row}"] = str(value)
            ws[f"B{row}"].font = _META_VALUE_FONT
            row += 1
        row += 1

    # --- Data table ---------------------------------------------------------
    header_row = row
    ws.cell(row=header_row, column=1, value="Category")
    for col_idx, s in enumerate(raw_series, start=2):
        ws.cell(row=header_row, column=col_idx, value=str(s.get("name") or f"Series {col_idx - 1}"))
    for cell in ws[header_row]:
        if cell.column <= 1 + len(raw_series):
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER

    first_data_row = header_row + 1
    for r, category in enumerate(categories):
        excel_row = first_data_row + r
        c = ws.cell(row=excel_row, column=1, value=category)
        c.border = _BORDER
        for col_idx, s in enumerate(raw_series, start=2):
            values = s.get("values") or []
            value = _coerce_number(values[r]) if r < len(values) else None
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="right")
    last_data_row = first_data_row + max(len(categories) - 1, 0)

    # --- Column widths ------------------------------------------------------
    ws.column_dimensions["A"].width = max(18, min(48, max((len(c) for c in categories), default=10) + 2))
    for col_idx, s in enumerate(raw_series, start=2):
        letter = ws.cell(row=header_row, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(14, len(str(s.get("name") or "")) + 2)

    # --- Native chart -------------------------------------------------------
    if categories and raw_series:
        chart = _build_chart(chart_type, title, spec)
        data = Reference(
            ws, min_col=2, max_col=1 + len(raw_series),
            min_row=header_row, max_row=last_data_row,
        )
        cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Colour each series from the spec (which the UI filled from the palette).
        is_line = chart_type == "line"
        is_pie = chart_type == "pie"
        if is_pie:
            # Pie: colour each data point (one series, many slices).
            from openpyxl.chart.series import DataPoint
            ser = chart.series[0]
            for i, _cat in enumerate(categories):
                color = series_color(i)
                dp = DataPoint(idx=i)
                dp.graphicalProperties = GraphicalProperties(solidFill=color)
                ser.data_points.append(dp)
        else:
            for i, s in enumerate(chart.series):
                spec_color = raw_series[i].get("color") if i < len(raw_series) else None
                _apply_series_color(s, spec_color or series_color(i), line=is_line)

        anchor_row = header_row
        anchor_col_letter = ws.cell(row=1, column=3 + len(raw_series)).column_letter
        ws.add_chart(chart, f"{anchor_col_letter}{anchor_row}")


def _series_values(raw_series: list[dict], count: int) -> list[list]:
    """Per-series numeric value lists, padded to `count` entries."""
    out = []
    for s in raw_series:
        vals = [_coerce_number(v) for v in (s.get("values") or [])]
        vals = vals[:count] + [None] * (count - len(vals))
        out.append(vals)
    return out


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else 0


def _style_header(cell):
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = _BORDER


def _write_pivot_sheet(ws, title, categories, raw_series):
    """Category × series matrix with a Total column and a Total row."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    names = [str(s.get("name") or f"Series {i + 1}") for i, s in enumerate(raw_series)]
    values = _series_values(raw_series, len(categories))

    header_row = 3
    ws.cell(row=header_row, column=1, value="Category")
    for c, name in enumerate(names, start=2):
        ws.cell(row=header_row, column=c, value=name)
    ws.cell(row=header_row, column=2 + len(names), value="Total")
    for cell in ws[header_row]:
        if cell.column <= 2 + len(names):
            _style_header(cell)

    for r, category in enumerate(categories):
        excel_row = header_row + 1 + r
        ws.cell(row=excel_row, column=1, value=category).border = _BORDER
        row_total = 0
        for c in range(len(names)):
            v = values[c][r]
            row_total += _num(v)
            cell = ws.cell(row=excel_row, column=2 + c, value=v)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="right")
        tcell = ws.cell(row=excel_row, column=2 + len(names), value=row_total)
        tcell.font = Font(bold=True)
        tcell.border = _BORDER
        tcell.alignment = Alignment(horizontal="right")

    total_row = header_row + 1 + len(categories)
    tot = ws.cell(row=total_row, column=1, value="Total")
    tot.font = Font(bold=True)
    grand = 0
    for c in range(len(names)):
        col_total = sum(_num(values[c][r]) for r in range(len(categories)))
        grand += col_total
        cell = ws.cell(row=total_row, column=2 + c, value=col_total)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
    gcell = ws.cell(row=total_row, column=2 + len(names), value=grand)
    gcell.font = Font(bold=True)
    gcell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = max(18, min(48, max((len(c) for c in categories), default=10) + 2))
    for c, name in enumerate(names, start=2):
        ws.column_dimensions[ws.cell(row=header_row, column=c).column_letter].width = max(14, len(name) + 2)


def _write_summary_sheet(ws, title, subtitle, source_filters, categories, raw_series):
    """Headline totals: grand total, per-series totals, per-category totals."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    row = 2
    if subtitle:
        ws[f"A{row}"] = subtitle
        ws[f"A{row}"].font = _SUBTITLE_FONT
        row += 1
    row += 1

    if source_filters:
        ws.cell(row=row, column=1, value="Filters applied").font = _META_LABEL_FONT
        row += 1
        for label, value in source_filters.items():
            if value in (None, "", []):
                continue
            ws.cell(row=row, column=1, value=str(label)).font = _META_LABEL_FONT
            ws.cell(row=row, column=2, value=str(value)).font = _META_VALUE_FONT
            row += 1
        row += 1

    names = [str(s.get("name") or f"Series {i + 1}") for i, s in enumerate(raw_series)]
    values = _series_values(raw_series, len(categories))
    grand = sum(_num(values[c][r]) for c in range(len(names)) for r in range(len(categories)))

    ws.cell(row=row, column=1, value="Grand total").font = _META_LABEL_FONT
    gc = ws.cell(row=row, column=2, value=grand)
    gc.font = Font(bold=True, size=12)
    row += 2

    _style_header(ws.cell(row=row, column=1, value="Series"))
    _style_header(ws.cell(row=row, column=2, value="Total"))
    row += 1
    for c, name in enumerate(names):
        ws.cell(row=row, column=1, value=name).border = _BORDER
        cell = ws.cell(row=row, column=2, value=sum(_num(values[c][r]) for r in range(len(categories))))
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="right")
        row += 1
    row += 1

    _style_header(ws.cell(row=row, column=1, value="Category"))
    _style_header(ws.cell(row=row, column=2, value="Total"))
    row += 1
    for r, category in enumerate(categories):
        ws.cell(row=row, column=1, value=category).border = _BORDER
        cell = ws.cell(row=row, column=2, value=sum(_num(values[c][r]) for c in range(len(names))))
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="right")
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16


def _write_raw_sheet(ws, categories, raw_series):
    """Tidy long-format: one row per (category, series, value) for re-analysis."""
    ws.sheet_view.showGridLines = False
    for col, head in enumerate(["Category", "Series", "Value"], start=1):
        _style_header(ws.cell(row=1, column=col, value=head))
    names = [str(s.get("name") or f"Series {i + 1}") for i, s in enumerate(raw_series)]
    values = _series_values(raw_series, len(categories))
    r = 2
    for ci, category in enumerate(categories):
        for si, name in enumerate(names):
            ws.cell(row=r, column=1, value=category)
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=values[si][ci])
            r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14


def _build_chart(chart_type: str, title: str, spec: dict[str, Any]):
    value_axis_title = spec.get("value_axis_title") or "Value"
    category_axis_title = spec.get("category_axis_title") or "Category"

    if chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
        chart.title = title
        chart.height = 10
        chart.width = 18
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        return chart
    else:
        chart = BarChart()
        chart.type = "col" if chart_type in {"column", "stacked-column"} else "bar"
        if chart_type in {"stacked-bar", "stacked-column"}:
            chart.grouping = "stacked"
            chart.overlap = 100

    chart.title = title
    chart.height = 11
    chart.width = 24
    chart.style = 10
    if hasattr(chart, "x_axis"):
        chart.x_axis.title = category_axis_title
        chart.x_axis.delete = False
    if hasattr(chart, "y_axis"):
        chart.y_axis.title = value_axis_title
        chart.y_axis.delete = False

    # Donor/M&E house style (see the example reports): attach a native data
    # table with legend colour keys directly beneath the plot, and drop the
    # separate legend (the keys live in the table). This is the built-in Excel
    # chart "Data Table" element, so it stays in sync if the user edits values.
    from openpyxl.chart.plotarea import DataTable

    chart.plot_area.dTable = DataTable(
        showHorzBorder=True,
        showVertBorder=True,
        showOutline=True,
        showKeys=True,
    )
    chart.legend = None
    return chart
