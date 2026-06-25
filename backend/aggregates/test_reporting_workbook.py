"""Tests for the SESIGO reporting workbook generator + importer.

Covers the four success-criteria areas:

  * Generation — correct project/org/quarter metadata, indicators, and that the
    disaggregation matrix is driven by the indicator's configuration.
  * Import validation — non-SESIGO files, missing sheets, bad version, and a
    workbook whose embedded metadata points at a missing project are rejected
    with friendly messages (never "missing indicator_id column").
  * Permissions — project assignment / organization scope / training isolation.
  * Round trip — generate → enter values → upload → download-with-data yields
    identical numbers.
"""
import tempfile
from datetime import date
from io import BytesIO

from django.test import override_settings
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from aggregates import reporting_workbook as rw
from aggregates.models import Aggregate
from indicators.models import Indicator
from organizations.models import Organization
from projects.models import Project, ProjectIndicator, ProjectIndicatorOrganizationTarget
from users.models import User

WORKBOOK_URL = "/api/aggregates/reporting-workbook/"
IMPORT_URL = "/api/aggregates/import-reporting-workbook/"

# Botswana FY 2025/26, Q3 = Oct–Dec 2025.
Q3_START = date(2025, 10, 1)
Q3_END = date(2025, 12, 31)

SEX_AGE_CONFIG = {
    "enabled": True,
    "layout": "matrix",
    "dimensions": [
        {"key": "sex", "label": "Sex", "values": ["Male", "Female"]},
        {"key": "age_band", "label": "Age Range", "values": ["10-14", "15-19", "20-24"]},
    ],
}


class _BaseSetup(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.org = Organization.objects.create(name="Home of Hope", code="WB_HOH", type="cso")
        cls.other_org = Organization.objects.create(name="Other CBO", code="WB_OTHER", type="cso")
        cls.admin = User.objects.create_user(
            username="wb_admin", email="wb_admin@example.com", password="TestPass123!",
            role="admin", organization=cls.org,
        )
        cls.officer = User.objects.create_user(
            username="wb_off", email="wb_off@example.com", password="TestPass123!",
            role="officer", organization=cls.org,
        )

        cls.project = Project.objects.create(
            name="Sesigo", code="WB-1", start_date=date(2025, 4, 1), end_date=date(2026, 3, 31),
            created_by=cls.admin,
        )
        cls.project.organizations.add(cls.org)

        cls.plain = Indicator.objects.create(
            name="Number of CSOs trained", code="WB_PLAIN", type="number",
            category="hiv_prevention", created_by=cls.admin,
        )
        cls.matrix = Indicator.objects.create(
            name="Number of people reached", code="WB_MATRIX", type="number",
            category="hiv_prevention", aggregate_disaggregation_config=SEX_AGE_CONFIG,
            created_by=cls.admin,
        )
        pi_plain = ProjectIndicator.objects.create(
            project=cls.project, indicator=cls.plain, q3_target=50,
        )
        pi_matrix = ProjectIndicator.objects.create(
            project=cls.project, indicator=cls.matrix, q3_target=300,
        )
        # Organization-scoped targets define the org's assignment for this project.
        ProjectIndicatorOrganizationTarget.objects.create(
            project_indicator=pi_plain, organization=cls.org, q3_target=25,
        )
        ProjectIndicatorOrganizationTarget.objects.create(
            project_indicator=pi_matrix, organization=cls.org, q3_target=300,
        )

    def _download_blank(self, user=None, **params):
        self.client.force_authenticate(user or self.officer)
        query = {"project": self.project.id, "organization": self.org.id,
                 "quarter": "Q3", "fiscal_year": 2025}
        query.update(params)
        return self.client.get(WORKBOOK_URL, query)


class WorkbookGenerationTests(_BaseSetup):
    def test_blank_workbook_has_expected_sheets_and_metadata(self):
        response = self._download_blank()
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", response["Content-Type"])

        wb = load_workbook(BytesIO(response.content))
        self.assertIn(rw.SHEET_FORM, wb.sheetnames)
        self.assertIn(rw.SHEET_META, wb.sheetnames)
        self.assertIn(rw.SHEET_CELLMAP, wb.sheetnames)
        self.assertIn(rw.SHEET_INSTRUCTIONS, wb.sheetnames)
        # Hidden machine sheets must not be visible to users.
        self.assertEqual(wb[rw.SHEET_META].sheet_state, "hidden")
        self.assertEqual(wb[rw.SHEET_CELLMAP].sheet_state, "veryHidden")

        meta = {row[0]: row[1] for row in wb[rw.SHEET_META].iter_rows(min_row=2, max_col=2, values_only=True)}
        self.assertEqual(int(meta["project_id"]), self.project.id)
        self.assertEqual(int(meta["organization_id"]), self.org.id)
        self.assertEqual(int(meta["quarter"]), 3)
        self.assertEqual(meta["period_start"], Q3_START.isoformat())
        self.assertEqual(meta["period_end"], Q3_END.isoformat())
        self.assertEqual(meta["workbook_version"], rw.WORKBOOK_VERSION)

    def test_form_columns_driven_by_indicator_matrix(self):
        """The matrix indicator must render its sex/age bands; the plain one must not."""
        response = self._download_blank()
        wb = load_workbook(BytesIO(response.content))
        form_text = {
            cell.value
            for row in wb[rw.SHEET_FORM].iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        }
        self.assertIn("Number of people reached", form_text)
        self.assertIn("Number of CSOs trained", form_text)
        # Age bands from the indicator config appear as column headers.
        self.assertIn("10-14", form_text)
        self.assertIn("15-19", form_text)
        # The secondary-axis column header is the indicator config's actual
        # dimension label (here "Sex"), never a hardcoded "SEX"/"AGE/SEX".
        self.assertIn("Sex", form_text)
        self.assertNotIn("AGE/SEX", form_text)

        # The cell map carries no bands for the plain indicator and the full
        # sex×age grid (2×3 = 6 cells) for the matrix indicator.
        cm_rows = [r for r in wb[rw.SHEET_CELLMAP].iter_rows(min_row=2, values_only=True) if r[0]]
        matrix_cells = [r for r in cm_rows if int(r[0]) == self.matrix.id]
        plain_cells = [r for r in cm_rows if int(r[0]) == self.plain.id]
        self.assertEqual(len(matrix_cells), 6)
        self.assertEqual(len(plain_cells), 1)
        self.assertEqual(plain_cells[0][2], "total")

    def test_non_sex_secondary_uses_real_dimension_label(self):
        """A non-sex SECONDARY dimension must render its own label, not 'SEX'.

        The secondary (column C) axis is occupied by a non-sex dimension when an
        indicator has two non-special dimensions (here Message Type x Key
        Population). The header for that column must be the config's actual label.
        """
        ind = Indicator.objects.create(
            name="People engaged with NCD messages", code="WB_KP2", type="number",
            category="media", created_by=self.admin,
            aggregate_disaggregation_config={
                "enabled": True, "layout": "matrix", "dimensions": [
                    {"key": "ncd_prevention_messages", "label": "NCD Prevention Messages",
                     "values": ["Alcohol", "Tobacco"]},
                    {"key": "key_population", "label": "Key Population",
                     "values": ["FSW", "MSM", "PWID"]},
                ]},
        )
        pi = ProjectIndicator.objects.create(project=self.project, indicator=ind, q3_target=10)
        ProjectIndicatorOrganizationTarget.objects.create(
            project_indicator=pi, organization=self.org, q3_target=10)

        response = self._download_blank()
        wb = load_workbook(BytesIO(response.content))
        form_text = {
            cell.value for row in wb[rw.SHEET_FORM].iter_rows() for cell in row
            if isinstance(cell.value, str)
        }
        # Real labels/values appear; the secondary column is NOT mislabelled "SEX".
        self.assertIn("Key Population", form_text)
        self.assertIn("FSW", form_text)
        self.assertNotIn("SEX", form_text)
        self.assertNotIn("AGE/SEX", form_text)

    def test_total_column_aligned_across_indicators(self):
        """Every indicator's TOTAL header must land in the same column, regardless
        of how many band columns it has (plain, sex+age, and a wider age set)."""
        wide = Indicator.objects.create(
            name="Wider age indicator", code="WB_WIDE", type="number", category="hiv_prevention",
            created_by=self.admin,
            aggregate_disaggregation_config={"enabled": True, "layout": "matrix", "dimensions": [
                {"key": "sex", "label": "Sex", "values": ["Male", "Female"]},
                {"key": "age_band", "label": "Age Range",
                 "values": ["10-14", "15-19", "20-24", "25-29", "30-34", "35-39"]},
            ]})
        pi = ProjectIndicator.objects.create(project=self.project, indicator=wide, q3_target=10)
        ProjectIndicatorOrganizationTarget.objects.create(
            project_indicator=pi, organization=self.org, q3_target=10)

        response = self._download_blank()
        wb = load_workbook(BytesIO(response.content))
        ws = wb[rw.SHEET_FORM]
        total_cols = {
            cell.column for row in ws.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.strip() == "TOTAL"
        }
        self.assertEqual(len(total_cols), 1, f"TOTAL headers not aligned: columns {sorted(total_cols)}")

    def test_only_assigned_indicators_are_included(self):
        unassigned = Indicator.objects.create(
            name="Unassigned indicator", code="WB_UNASSIGNED", type="number", created_by=self.admin,
        )
        response = self._download_blank()
        wb = load_workbook(BytesIO(response.content))
        form_text = {
            cell.value for row in wb[rw.SHEET_FORM].iter_rows() for cell in row
            if isinstance(cell.value, str)
        }
        self.assertNotIn(unassigned.name, form_text)


class WorkbookImportValidationTests(_BaseSetup):
    def _post(self, content, name="wb.xlsx", **extra):
        self.client.force_authenticate(self.officer)
        upload = BytesIO(content)
        upload.name = name
        data = {"file": upload}
        data.update(extra)
        return self.client.post(IMPORT_URL, data, format="multipart")

    def test_rejects_non_excel_file(self):
        response = self._post(b"this is not excel", name="notes.txt")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("messages", response.data)

    def test_rejects_arbitrary_workbook_without_metadata(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(["indicator_id", "value"])
        buf = BytesIO()
        wb.save(buf)
        response = self._post(buf.getvalue())
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        joined = " ".join(response.data.get("messages", []))
        self.assertIn("not a SESIGO reporting workbook", joined)
        # Never surfaces a technical column error.
        self.assertNotIn("indicator_id", joined)

    def test_friendly_error_lists_missing_pieces(self):
        # Generate a valid workbook then wipe the version to force a friendly fail.
        buf = self._download_blank().content
        wb = load_workbook(BytesIO(buf))
        for row in wb[rw.SHEET_META].iter_rows(min_row=2, max_col=2):
            if row[0].value == "workbook_version":
                row[1].value = "something-else"
        out = BytesIO()
        wb.save(out)
        response = self._post(out.getvalue())
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["error"], "Workbook Validation Failed")


class WorkbookPermissionTests(_BaseSetup):
    def test_requires_writable_organization(self):
        # Officer scoped to cls.org cannot generate a workbook for other_org.
        self.client.force_authenticate(self.officer)
        response = self.client.get(WORKBOOK_URL, {
            "project": self.project.id, "organization": self.other_org.id,
            "quarter": "Q3", "fiscal_year": 2025,
        })
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_client_role_cannot_generate(self):
        client_user = User.objects.create_user(
            username="wb_client", email="wb_client@example.com", password="TestPass123!",
            role="client", organization=self.org,
        )
        response = self._download_blank(user=client_user)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_live_request_cannot_import_training_workbook(self):
        training_project = Project.objects.create(
            name="Training", code="WB-TRAIN", start_date=date(2025, 4, 1),
            end_date=date(2026, 3, 31), is_training=True, created_by=self.admin,
        )
        training_project.organizations.add(self.org)
        ProjectIndicator.objects.create(project=training_project, indicator=self.plain)

        # Build a workbook for the training project, then upload it as a live request.
        plans = [rw.IndicatorPlan(indicator=self.plain, config=rw.resolve_matrix_config(self.plain), target=None)]
        buf = rw.generate_workbook(
            project=training_project, organization=self.org, quarter=3,
            fiscal_start_year=2025, indicator_plans=plans,
        )
        self.client.force_authenticate(self.officer)
        upload = BytesIO(buf.getvalue())
        upload.name = "training.xlsx"
        response = self.client.post(IMPORT_URL, {"file": upload}, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class LegacyEntrypointConvergenceTests(_BaseSetup):
    """The legacy uploads entrypoints reuse the donor workbook engine
    (Phase 1/3) without breaking the existing flat/NAHPA pipeline."""

    def test_download_template_with_org_returns_donor_workbook(self):
        self.client.force_authenticate(self.officer)
        response = self.client.get(
            "/api/uploads/download_template/",
            {"project": self.project.id, "organization": self.org.id,
             "quarter": "Q3", "fiscal_year": 2025},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        # It is the donor workbook (hidden machine sheets), not the flat template.
        self.assertIn(rw.SHEET_META, wb.sheetnames)
        self.assertIn(rw.SHEET_CELLMAP, wb.sheetnames)

    def test_start_import_routes_sesigo_workbook_to_smart_importer(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from uploads.models import Upload

        # Build a filled donor workbook for the matrix indicator.
        plans = [rw.IndicatorPlan(indicator=self.matrix, config=rw.resolve_matrix_config(self.matrix), target=None)]
        buf = rw.generate_workbook(project=self.project, organization=self.org,
                                   quarter=3, fiscal_start_year=2025, indicator_plans=plans)
        wb = load_workbook(buf)
        for r in wb[rw.SHEET_CELLMAP].iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            wb[rw.SHEET_FORM][r[6].split("!", 1)[1]] = 2
        out = BytesIO(); wb.save(out)

        upload = Upload.objects.create(
            name="filled.xlsx",
            file=SimpleUploadedFile("filled.xlsx", out.getvalue()),
            organization=self.org, created_by=self.officer,
        )
        self.client.force_authenticate(self.officer)
        response = self.client.post(f"/api/uploads/{upload.id}/start_import/", {}, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.assertEqual(response.data.get("job_type"), "reporting_workbook_import")
        self.assertEqual(response.data.get("status"), "imported")
        self.assertTrue(
            Aggregate.objects.filter(indicator=self.matrix, project=self.project, organization=self.org).exists()
        )

    def test_start_import_leaves_non_sesigo_uploads_on_legacy_path(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from openpyxl import Workbook
        from uploads.models import Upload

        wb = Workbook(); wb.active.append(["indicator_id", "value"]); buf = BytesIO(); wb.save(buf)
        upload = Upload.objects.create(
            name="flat.xlsx", file=SimpleUploadedFile("flat.xlsx", buf.getvalue()),
            organization=self.org, created_by=self.officer,
        )
        self.client.force_authenticate(self.officer)
        response = self.client.post(f"/api/uploads/{upload.id}/start_import/", {}, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        # Legacy pipeline: a non-queued upload becomes a review job, untouched.
        self.assertEqual(response.data.get("job_type"), "upload_import")
        self.assertEqual(response.data.get("status"), "ready_for_review")


class WorkbookRoundTripTests(_BaseSetup):
    def test_generate_fill_upload_export_is_identical(self):
        # 1. Download blank.
        blank = self._download_blank().content
        wb = load_workbook(BytesIO(blank))

        # 2. Enter values into every input cell via the cell map.
        cellmap = wb[rw.SHEET_CELLMAP]
        form = wb[rw.SHEET_FORM]
        counter = 1
        entered = {}
        for r in cellmap.iter_rows(min_row=2, values_only=True):
            iid, code, kind, primary, secondary, band, coord = (list(r) + [None] * 7)[:7]
            if iid is None:
                continue
            cell = coord.split("!", 1)[1]
            form[cell] = counter
            entered[(int(iid), kind, primary, secondary, band)] = counter
            counter += 1
        out = BytesIO()
        wb.save(out)

        # 3. Upload.
        self.client.force_authenticate(self.officer)
        upload = BytesIO(out.getvalue())
        upload.name = "filled.xlsx"
        response = self.client.post(IMPORT_URL, {"file": upload}, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.assertEqual(response.data["summary"]["indicators_failed"], 0)

        # The aggregates landed pending review with the canonical value shape.
        matrix_agg = Aggregate.objects.get(indicator=self.matrix, project=self.project, organization=self.org)
        self.assertEqual(matrix_agg.status, "pending")
        self.assertIn("disaggregates", matrix_agg.value)
        plain_agg = Aggregate.objects.get(indicator=self.plain, project=self.project, organization=self.org)
        self.assertIn("total", plain_agg.value)

        # 4. Download with data and confirm the numbers match what was entered.
        with_data = self.client.get(WORKBOOK_URL, {
            "project": self.project.id, "organization": self.org.id,
            "quarter": "Q3", "fiscal_year": 2025, "with_data": "true",
        })
        self.assertEqual(with_data.status_code, status.HTTP_200_OK)
        wb2 = load_workbook(BytesIO(with_data.content), data_only=False)
        cellmap2 = wb2[rw.SHEET_CELLMAP]
        form2 = wb2[rw.SHEET_FORM]
        for r in cellmap2.iter_rows(min_row=2, values_only=True):
            iid, code, kind, primary, secondary, band, coord = (list(r) + [None] * 7)[:7]
            if iid is None:
                continue
            cell = coord.split("!", 1)[1]
            expected = entered[(int(iid), kind, primary, secondary, band)]
            self.assertEqual(form2[cell].value, expected,
                             f"cell {cell} for indicator {iid} did not round-trip")


class WorkbookDuplicatePreventionTests(_BaseSetup):
    """Idempotency + duplicate safeguards for the reporting-workbook importer
    (IMP-1): re-uploading the same file must not create duplicate rows, inflate
    analytics, or silently discard a reviewer's sign-off, and the preview must
    tell the user exactly what an import will create/update/skip."""

    def _filled_workbook_bytes(self, value=7):
        """A SESIGO workbook with ``value`` typed into every input cell."""
        blank = self._download_blank().content
        wb = load_workbook(BytesIO(blank))
        cellmap, form = wb[rw.SHEET_CELLMAP], wb[rw.SHEET_FORM]
        for r in cellmap.iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            form[r[6].split("!", 1)[1]] = value
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def _import(self, content, name="filled.xlsx", **extra):
        self.client.force_authenticate(self.officer)
        upload = BytesIO(content)
        upload.name = name
        data = {"file": upload}
        data.update(extra)
        return self.client.post(IMPORT_URL, data, format="multipart")

    def test_reupload_is_idempotent_no_duplicate_rows(self):
        content = self._filled_workbook_bytes()

        first = self._import(content)
        self.assertEqual(first.status_code, status.HTTP_201_CREATED, first.data)
        self.assertEqual(first.data["summary"]["created"], 2)
        rows_after_first = Aggregate.objects.count()

        # Re-upload the identical file: no new rows, everything "unchanged".
        second = self._import(content)
        self.assertEqual(second.status_code, status.HTTP_201_CREATED, second.data)
        self.assertEqual(Aggregate.objects.count(), rows_after_first)
        self.assertEqual(second.data["summary"]["created"], 0)
        self.assertEqual(second.data["summary"]["updated"], 0)
        self.assertEqual(second.data["summary"]["unchanged"], 2)

    def test_reupload_preserves_approved_status(self):
        content = self._filled_workbook_bytes()
        self._import(content)

        # A manager approves one of the imported records.
        agg = Aggregate.objects.get(indicator=self.matrix, project=self.project, organization=self.org)
        agg.status = "approved"
        agg.save(update_fields=["status"])

        # Re-uploading the *same* numbers must not knock the approval back.
        second = self._import(content)
        self.assertEqual(second.data["summary"]["unchanged"], 2)
        agg.refresh_from_db()
        self.assertEqual(agg.status, "approved")

    def test_changed_reupload_resets_reviewed_record_and_reports_it(self):
        content = self._filled_workbook_bytes(value=7)
        self._import(content)
        agg = Aggregate.objects.get(indicator=self.matrix, project=self.project, organization=self.org)
        agg.status = "approved"
        agg.save(update_fields=["status"])

        # Different numbers => the approved record is legitimately re-opened for
        # review, but the import reports it so reviewers are not surprised.
        changed = self._import(self._filled_workbook_bytes(value=9))
        summary = changed.data["summary"]
        self.assertEqual(summary["reset_from_review"], 1)
        self.assertIn(self.matrix.code, summary["reset_from_review_indicators"])
        agg.refresh_from_db()
        self.assertEqual(agg.status, "pending")

    def test_dry_run_preview_classifies_each_indicator(self):
        content = self._filled_workbook_bytes(value=7)
        # First import the plain indicator's record so the dry-run sees a mix of
        # create (matrix) and unchanged (plain)… by importing everything first.
        self._import(content)

        preview = self._import(content, dry_run="true")
        self.assertEqual(preview.status_code, status.HTTP_200_OK)
        self.assertTrue(preview.data["dry_run"])
        summary = preview.data["summary"]
        # Nothing changed since the prior import: all unchanged, no writes.
        self.assertEqual(summary["unchanged"], 2)
        self.assertEqual(summary["to_create"], 0)
        self.assertEqual(summary["to_update"], 0)
        outcomes = {p["outcome"] for p in preview.data["preview"]}
        self.assertEqual(outcomes, {"unchanged"})
        # Dry-run must not have written anything.
        self.assertEqual(Aggregate.objects.count(), 2)


class BulkCreateDuplicateTests(_BaseSetup):
    """The bulk_create endpoint must upsert (never duplicate) and report when
    the same indicator is sent twice in one payload."""

    BULK_URL = "/api/aggregates/bulk_create/"

    def _payload(self, rows):
        return {
            "project": self.project.id, "organization": self.org.id,
            "period_start": Q3_START.isoformat(), "period_end": Q3_END.isoformat(),
            "data": rows,
        }

    def test_duplicate_indicator_in_payload_is_reported(self):
        self.client.force_authenticate(self.officer)
        resp = self.client.post(self.BULK_URL, self._payload([
            {"indicator": self.plain.id, "value": 5},
            {"indicator": self.plain.id, "value": 8},  # duplicate row, last wins
        ]), format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        # Only one row exists for the indicator despite two payload entries.
        self.assertEqual(
            Aggregate.objects.filter(indicator=self.plain, project=self.project, organization=self.org).count(),
            1,
        )
        self.assertIn(self.plain.id, resp.data["duplicate_indicators_in_payload"])

    def test_identical_resubmit_counts_as_unchanged(self):
        self.client.force_authenticate(self.officer)
        rows = [{"indicator": self.plain.id, "value": 5}]
        first = self.client.post(self.BULK_URL, self._payload(rows), format="json")
        self.assertEqual(first.data["created"], 1)
        second = self.client.post(self.BULK_URL, self._payload(rows), format="json")
        self.assertEqual(second.data["created"], 0)
        self.assertEqual(second.data["unchanged"], 1)
        self.assertEqual(Aggregate.objects.count(), 1)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class UploadFingerprintTests(_BaseSetup):
    """Uploads carry a SHA-256 fingerprint so repeated uploads of the same file
    are detectable (and the smart-import response flags them)."""

    def _filled_workbook_bytes(self, value=4):
        plans = [rw.IndicatorPlan(indicator=self.matrix, config=rw.resolve_matrix_config(self.matrix), target=None)]
        buf = rw.generate_workbook(project=self.project, organization=self.org,
                                   quarter=3, fiscal_start_year=2025, indicator_plans=plans)
        wb = load_workbook(buf)
        for r in wb[rw.SHEET_CELLMAP].iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            wb[rw.SHEET_FORM][r[6].split("!", 1)[1]] = value
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def test_save_populates_file_hash(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from uploads.models import Upload

        content = b"hello fingerprint"
        upload = Upload.objects.create(
            name="a.txt", file=SimpleUploadedFile("a.txt", content),
            organization=self.org, created_by=self.officer,
        )
        import hashlib
        self.assertEqual(upload.file_hash, hashlib.sha256(content).hexdigest())

    def test_duplicate_file_flagged_on_reimport(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from uploads.models import Upload

        content = self._filled_workbook_bytes()
        self.client.force_authenticate(self.officer)

        first_upload = Upload.objects.create(
            name="filled.xlsx", file=SimpleUploadedFile("filled.xlsx", content),
            organization=self.org, created_by=self.officer,
        )
        first = self.client.post(f"/api/uploads/{first_upload.id}/start_import/", {}, format="multipart")
        self.assertEqual(first.status_code, status.HTTP_201_CREATED, first.data)
        self.assertFalse(first.data.get("duplicate_file", False))

        # Same bytes uploaded again -> same fingerprint -> flagged as duplicate.
        second_upload = Upload.objects.create(
            name="filled-again.xlsx", file=SimpleUploadedFile("filled-again.xlsx", content),
            organization=self.org, created_by=self.officer,
        )
        self.assertEqual(second_upload.file_hash, first_upload.file_hash)
        second = self.client.post(f"/api/uploads/{second_upload.id}/start_import/", {}, format="multipart")
        self.assertEqual(second.status_code, status.HTTP_201_CREATED, second.data)
        self.assertTrue(second.data.get("duplicate_file"))
        self.assertEqual(second.data.get("previous_upload_id"), first_upload.id)


class ApprovedDataLifecycleTests(_BaseSetup):
    """Permanent best-practice rule (IMP-1): any *real* change to approved data
    returns it to Pending — even for an admin — and every write keeps a full
    before/after audit trail (old/new value, prev/new status, user, source)."""

    DETAIL = "/api/aggregates/{}/"

    def _approved(self, value=10):
        return Aggregate.objects.create(
            indicator=self.plain, project=self.project, organization=self.org,
            period_start=Q3_START, period_end=Q3_END, value={"total": value},
            status="approved", created_by=self.officer,
        )

    def _last_event(self, agg):
        from audit.models import AuditEvent
        return (
            AuditEvent.objects.filter(object_type="aggregate", object_id=str(agg.id))
            .order_by("-created_at").first()
        )

    def test_admin_value_change_resets_to_pending_with_audit(self):
        agg = self._approved(10)
        self.client.force_authenticate(self.admin)
        resp = self.client.patch(self.DETAIL.format(agg.id), {"value": {"total": 20}}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_200_OK, resp.data)
        agg.refresh_from_db()
        self.assertEqual(agg.status, "pending")
        self.assertIsNone(agg.reviewed_by)

        event = self._last_event(agg)
        self.assertIsNotNone(event)
        self.assertEqual(event.actor_id, self.admin.id)
        self.assertEqual(event.metadata["source"], "direct_edit")
        self.assertEqual(event.metadata["previous_status"], "approved")
        self.assertEqual(event.metadata["new_status"], "pending")
        self.assertEqual(event.metadata["old_value"], {"total": 10})
        self.assertEqual(event.metadata["new_value"], {"total": 20})
        self.assertEqual(event.metadata["outcome"], "reset_from_review")

    def test_no_value_change_preserves_approval(self):
        agg = self._approved(10)
        self.client.force_authenticate(self.admin)
        resp = self.client.patch(self.DETAIL.format(agg.id), {"notes": "context only"}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_200_OK, resp.data)
        agg.refresh_from_db()
        self.assertEqual(agg.status, "approved")

    def test_admin_can_reapprove_after_correction(self):
        agg = self._approved(10)
        self.client.force_authenticate(self.admin)
        self.client.patch(self.DETAIL.format(agg.id), {"value": {"total": 20}}, format="json")
        # Pending → re-approve closes the lifecycle.
        resp = self.client.post(f"/api/aggregates/{agg.id}/approve/", {}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_200_OK, resp.data)
        agg.refresh_from_db()
        self.assertEqual(agg.status, "approved")
        self.assertEqual(agg.reviewed_by_id, self.admin.id)

    def test_workbook_import_writes_audit_with_source(self):
        from audit.models import AuditEvent

        blank = self._download_blank().content
        wb = load_workbook(BytesIO(blank))
        for r in wb[rw.SHEET_CELLMAP].iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            wb[rw.SHEET_FORM][r[6].split("!", 1)[1]] = 3
        out = BytesIO(); wb.save(out)

        self.client.force_authenticate(self.officer)
        upload = BytesIO(out.getvalue()); upload.name = "filled.xlsx"
        resp = self.client.post(IMPORT_URL, {"file": upload}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED, resp.data)
        created = resp.data["summary"]["created"]
        self.assertEqual(
            AuditEvent.objects.filter(
                object_type="aggregate", action="create", metadata__source="workbook_import",
            ).count(),
            created,
        )
