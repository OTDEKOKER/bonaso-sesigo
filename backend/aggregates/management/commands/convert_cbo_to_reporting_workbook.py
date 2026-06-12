"""Convert a CBO partner Excel report (file = organisation, sheets = period)
into a SESIGO Reporting Workbook that the existing smart importer ingests.

Design (no change to import code): generate the real reporting workbook for the
target project + organisation via aggregates.reporting_workbook.generate_workbook
(so Reporting Form / Metadata / _cellmap are valid by construction), then read
the _cellmap and write values pulled from the CBO report's chosen period sheet.

Example:
    python manage.py convert_cbo_to_reporting_workbook \
        --file "/path/APSA NCD REPORT - APRIL 26.xlsx" \
        --project-id 4 --org-id 101 --quarter 3 --fiscal-start-year 2025 \
        --period-sheet APRIL --out /tmp/APSA_reporting_workbook.xlsx
"""
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from aggregates import reporting_workbook as rw
from aggregates.reporting_workbook import generate_workbook
from indicators.models import Indicator
from organizations.models import Organization
from projects.models import Project
from projects.assignment_rules import get_assigned_indicator_ids_for_organization


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def _numbers_on_sheet(ws):
    """Collect (normalized_label -> number) pairs: for each numeric cell, take the
    nearest non-numeric label to its left on the same row as its key."""
    out = {}
    for row in ws.iter_rows(values_only=True):
        label = None
        for cell in row:
            if isinstance(cell, str) and cell.strip() and not _is_num(cell):
                label = _norm(cell)
            elif _is_num(cell) and label:
                val = float(cell)
                # keep the largest number seen for a label (usually the total)
                if label not in out or val > out[label]:
                    out[label] = val
    return out


def _is_num(v):
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v.replace(",", "").strip())
            return True
        except ValueError:
            return False
    return False


class Command(BaseCommand):
    help = "Convert a CBO Excel report into a SESIGO reporting workbook."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--project-id", type=int, required=True)
        parser.add_argument("--org-id", type=int, required=True)
        parser.add_argument("--quarter", type=int, required=True)
        parser.add_argument("--fiscal-start-year", type=int, required=True)
        parser.add_argument("--period-sheet", required=True,
                            help="CBO period sheet to read (e.g. APRIL). NOT the TOTAL sheet.")
        parser.add_argument("--out", required=True)

    def handle(self, *args, **o):
        src = Path(o["file"])
        if not src.exists():
            raise CommandError(f"File not found: {src}")
        if _norm(o["period_sheet"]) == "total":
            raise CommandError("Refusing to use the TOTAL sheet; pass a period (month) sheet.")

        project = Project.objects.get(id=o["project_id"])
        org = Organization.objects.get(id=o["org_id"])

        indicator_ids = list(get_assigned_indicator_ids_for_organization(
            project=project, organization_id=org.id))
        if not indicator_ids:
            raise CommandError(
                f"{org.name} has no assigned indicators in project {project.code}. "
                f"Run mirror_org_assignments_to_training first."
            )
        indicators = list(
            Indicator.objects.filter(id__in=indicator_ids, is_active=True)
            .exclude(canonical_indicator__isnull=False).order_by("name")
        )
        plans = [rw.IndicatorPlan(indicator=ind, config=rw.resolve_matrix_config(ind),
                                  target=None, existing_cells={}) for ind in indicators]

        # 1. Generate the structurally-valid workbook.
        buf = generate_workbook(
            project=project, organization=org, quarter=o["quarter"],
            fiscal_start_year=o["fiscal_start_year"], indicator_plans=plans,
            generated_by="cbo-converter",
        )
        wb = load_workbook(buf)
        cellmap = wb[rw.SHEET_CELLMAP]
        form = wb[rw.SHEET_FORM]

        # 2. Read the CBO period sheet's numbers.
        cbo = load_workbook(src, read_only=True, data_only=True)
        if o["period_sheet"] not in cbo.sheetnames:
            raise CommandError(f"Sheet '{o['period_sheet']}' not in workbook: {cbo.sheetnames}")
        cbo_numbers = _numbers_on_sheet(cbo[o["period_sheet"]])
        cbo.close()

        # 3. Fill the single-value ("Value" band) input cell of each indicator whose
        #    name matches a labelled number in the CBO sheet. (Conservative: matrix
        #    disaggregate cells are left for review-stage entry.)
        ind_by_code = {ind.code: ind for ind in indicators}
        filled = 0
        matched_indicators = set()
        header = [c.value for c in cellmap[1]]
        for r in range(2, cellmap.max_row + 1):
            rowvals = {header[i]: cellmap.cell(row=r, column=i + 1).value for i in range(len(header))}
            band = str(rowvals.get("band") or "")
            code = rowvals.get("indicator_code")
            coord = rowvals.get("coordinate")
            ind = ind_by_code.get(code)
            if not ind or band != rw.NO_BAND or not coord:
                continue
            key = _norm(ind.name)
            value = cbo_numbers.get(key)
            if value is None:
                # fuzzy contains-match against CBO labels
                for lbl, num in cbo_numbers.items():
                    if key and (key in lbl or lbl in key) and len(lbl) > 6:
                        value = num
                        break
            if value is not None:
                cell_ref = coord.split("!", 1)[1]
                form[cell_ref] = int(value)
                filled += 1
                matched_indicators.add(code)

        out = Path(o["out"])
        wb.save(out)
        self.stdout.write(self.style.SUCCESS(
            f"Wrote {out} | indicators={len(indicators)} "
            f"single-value cells filled={filled} matched_indicators={len(matched_indicators)}"
        ))
