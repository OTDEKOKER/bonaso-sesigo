"""Tests for the coordinator rollup workbook (one sheet per sub + TOTAL sheet).

The TOTAL sheet's cells must SUM the *matching* (indicator, primary, secondary,
band) cell on each sub that reports that indicator — not a blind same-coordinate
reference — even though subs have different indicator subsets.
"""
import re
from io import BytesIO
from types import SimpleNamespace

from django.test import SimpleTestCase
from openpyxl import load_workbook

from aggregates import reporting_workbook as rw

STD = ['10-14','15-19','20-24','25-29','30-34','35-39','40-44','45-49','50-54','55-59','60-64','65+']


def _indicator(iid, code, name):
    return SimpleNamespace(
        id=iid, code=code, name=name, type="number",
        aggregate_disaggregation_config={
            "enabled": True,
            "dimensions": [
                {"label": "Sex", "key": "sex", "values": ["Male", "Female"]},
                {"label": "Age Range", "key": "age_band", "values": list(STD)},
            ],
        },
    )


def _plan(ind):
    return rw.IndicatorPlan(indicator=ind, config=rw.resolve_matrix_config(ind), target=None, existing_cells={})


class CoordinatorWorkbookTests(SimpleTestCase):
    def _build(self):
        i1 = _indicator(1, "IND1", "Indicator One")
        i2 = _indicator(2, "IND2", "Indicator Two")
        project = SimpleNamespace(id=99, name="Test Project", code="TP", is_training=False)
        coord = SimpleNamespace(id=5, name="Coordinator Org", code="COORD")
        subA = SimpleNamespace(id=11, name="Sub A", code="A")
        subB = SimpleNamespace(id=12, name="Sub B", code="B")
        # Sub A reports only indicator 1; Sub B reports both (different layouts).
        sub_specs = [(subA, [_plan(i1)]), (subB, [_plan(i2), _plan(i1)])]
        coord_plans = [_plan(i1), _plan(i2)]
        buf = rw.generate_coordinator_workbook(
            project=project, coordinator=coord, sub_specs=sub_specs,
            coordinator_plans=coord_plans, quarter=1, fiscal_start_year=2026,
        )
        return load_workbook(BytesIO(buf.getvalue()))

    def test_has_total_sheet_last_and_one_sheet_per_sub(self):
        wb = self._build()
        visible = [s.title for s in wb.worksheets if s.sheet_state == "visible"]
        self.assertEqual(visible[-1], rw.SHEET_TOTAL)  # TOTAL is the last visible sheet
        self.assertIn("Sub A", visible)
        self.assertIn("Sub B", visible)

    def test_total_cells_reference_matching_indicator_cells(self):
        wb = self._build()
        cm = wb["_cellmap"]
        by_sheet = {}
        for r in cm.iter_rows(values_only=True):
            ind_id, code, kind, primary, secondary, band, coord = r
            if kind not in ("cell", "total"):
                continue
            sheet, cell = str(coord).split("!", 1)
            by_sheet.setdefault(sheet, {})[cell] = (ind_id, primary, secondary, band)

        total = wb[rw.SHEET_TOTAL]
        checked = ref_to_a = 0
        for row in total.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=SUM(") and "'!" in v):
                    continue
                tot_key = by_sheet[total.title][c.coordinate]
                for sheet, cell in re.findall(r"'([^']+)'!([A-Z]+\d+)", v):
                    checked += 1
                    # every ref must point at the SAME indicator/band on the sub
                    self.assertEqual(by_sheet[sheet][cell], tot_key)
                    if sheet == "Sub A":
                        ref_to_a += 1
        self.assertGreater(checked, 0)
        # Indicator 1 (in both subs) → Sub A is referenced; indicator 2 (only Sub
        # B) → its total cells reference only Sub B. So Sub A appears for IND1 only.
        self.assertGreater(ref_to_a, 0)

    def test_no_injected_cached_values_like_single_org(self):
        """The coordinator workbook must finalize like the single-org workbook: a
        plain openpyxl save with NO raw-XML cached-value injection. Stamping
        ``<v>real</v>`` into the cross-sheet formula cells is what made Excel
        "repair" the file on open — stripping merged-cell/border records (the table
        grid) and zeroing the TOTAL sheet. openpyxl's native ``<f>…</f><v></v>`` is
        exactly what the never-repairing single-org sheet ships, and Excel
        recomputes the rollups on open. Regression guard for the 'coordinator
        workbook repairs / tables & borders broken' report."""
        import zipfile
        i1 = _indicator(1, "IND1", "Indicator One")
        i2 = _indicator(2, "IND2", "Indicator Two")
        project = SimpleNamespace(id=99, name="Test Project", code="TP", is_training=False)
        coord = SimpleNamespace(id=5, name="Coordinator Org", code="COORD")
        subA = SimpleNamespace(id=11, name="Sub A", code="A")
        subB = SimpleNamespace(id=12, name="Sub B", code="B")
        buf = rw.generate_coordinator_workbook(
            project=project, coordinator=coord,
            sub_specs=[(subA, [_plan(i1)]), (subB, [_plan(i2), _plan(i1)])],
            coordinator_plans=[_plan(i1), _plan(i2)], quarter=1, fiscal_start_year=2026,
        )
        injected = formulas = 0
        with zipfile.ZipFile(BytesIO(buf.getvalue())) as z:
            for name in z.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    xml = z.read(name).decode("utf-8")
                    formulas += xml.count("</f>")
                    # a NON-empty cached value stamped right after a formula
                    injected += len(re.findall(r"</f><v>[^<]+</v>", xml))
        self.assertGreater(formulas, 0)
        self.assertEqual(injected, 0,
                         "formula cells must NOT carry injected cached values — that "
                         "raw-XML injection triggers Excel's repair (tables/borders stripped)")

    def test_total_sheet_rolls_up_by_cross_sheet_formula(self):
        """The TOTAL sheet rolls up via live cross-sheet ``=SUM`` formulas that Excel
        recomputes on open — not injected static values. Guards that each rollup
        cell references the matching (primary, secondary, band) cell on every
        reporting sub."""
        i1 = _indicator(1, "IND1", "Indicator One")
        project = SimpleNamespace(id=99, name="Test Project", code="TP", is_training=False)
        coord = SimpleNamespace(id=5, name="Coordinator Org", code="COORD")
        subA = SimpleNamespace(id=11, name="Sub A", code="A")
        subB = SimpleNamespace(id=12, name="Sub B", code="B")
        key = (rw.ALL_PRIMARY, "Male", "10-14")  # one filled (primary, secondary, band) cell
        planA = rw.IndicatorPlan(indicator=i1, config=rw.resolve_matrix_config(i1),
                                 target=None, existing_cells={key: 5})
        planB = rw.IndicatorPlan(indicator=i1, config=rw.resolve_matrix_config(i1),
                                 target=None, existing_cells={key: 7})
        buf = rw.generate_coordinator_workbook(
            project=project, coordinator=coord,
            sub_specs=[(subA, [planA]), (subB, [planB])],
            coordinator_plans=[_plan(i1)], quarter=1, fiscal_start_year=2026,
            with_data=True,
        )
        wb = load_workbook(BytesIO(buf.getvalue()))  # formulas, not cached values
        cellmap = wb["_cellmap"]
        prefix = rw.SHEET_TOTAL + "!"
        total_cells = [
            coord_ref.split("!", 1)[1]
            for (ind_id, code, kind, primary, secondary, band, coord_ref)
            in cellmap.iter_rows(min_row=2, values_only=True)
            if str(ind_id) == "1" and kind == "cell" and coord_ref and coord_ref.startswith(prefix)
        ]
        total_sheet = wb[rw.SHEET_TOTAL]
        sum_formulas = [total_sheet[c].value for c in total_cells
                        if isinstance(total_sheet[c].value, str) and total_sheet[c].value.startswith("=SUM(")]
        self.assertTrue(sum_formulas, "TOTAL cells must be cross-sheet =SUM formulas")
        joined = " ".join(sum_formulas)
        self.assertIn("'Sub A'!", joined)
        self.assertIn("'Sub B'!", joined)

    def test_no_disaggregate_total_cell_is_cross_sheet_formula(self):
        """A no-disaggregate (count) indicator's TOTAL cell must be a live
        cross-sheet ``=SUM`` — not a dead literal 0. The plain cell is recorded
        with EMPTY primary/secondary/band, so the TOTAL provider must look it up
        under the same empties (regression guard: the block used to pass
        ALL_PRIMARY/NO_BAND, which never matched ref_index → no auto-calc)."""
        plain = SimpleNamespace(id=1, code="CNT", name="Number of support groups",
                                type="number", aggregate_disaggregation_config={"enabled": False})
        self.assertFalse(rw.resolve_matrix_config(plain).has_disaggregates)
        project = SimpleNamespace(id=99, name="P", code="P", is_training=False)
        coord = SimpleNamespace(id=5, name="Coord", code="C")
        subA = SimpleNamespace(id=11, name="Sub A", code="A")
        subB = SimpleNamespace(id=12, name="Sub B", code="B")
        kp = (rw.ALL_PRIMARY, rw.ALL_PRIMARY, rw.NO_BAND)
        buf = rw.generate_coordinator_workbook(
            project=project, coordinator=coord,
            sub_specs=[(subA, [rw.IndicatorPlan(indicator=plain, config=rw.resolve_matrix_config(plain), target=None, existing_cells={kp: 3})]),
                       (subB, [rw.IndicatorPlan(indicator=plain, config=rw.resolve_matrix_config(plain), target=None, existing_cells={kp: 4})])],
            coordinator_plans=[_plan(plain)], quarter=1, fiscal_start_year=2026, with_data=True,
        )
        wb = load_workbook(BytesIO(buf.getvalue()))
        cm = wb["_cellmap"]
        prefix = rw.SHEET_TOTAL + "!"
        total_cells = [ref.split("!", 1)[1] for (i, c, k, pr, se, ba, ref) in cm.iter_rows(min_row=2, values_only=True)
                       if str(i) == "1" and ref and ref.startswith(prefix)]
        total = wb[rw.SHEET_TOTAL]
        self.assertTrue(total_cells)
        for c in total_cells:
            v = total[c].value
            self.assertTrue(isinstance(v, str) and v.startswith("=SUM(") and "!" in v,
                            f"no-disaggregate TOTAL cell {c} must be a cross-sheet =SUM, got {v!r}")
        # and the plain cell's merge must match the sub sheet's (consistent tables)
        r = int(re.sub(r"[A-Z]", "", total_cells[0]))
        tm = {str(m) for m in total.merged_cells.ranges if m.min_row <= r <= m.max_row}
        sm = {str(m) for m in wb["Sub A"].merged_cells.ranges if m.min_row <= r <= m.max_row}
        self.assertEqual(tm, sm, "no-disaggregate row merges must match the sub sheet")

    def test_no_single_cell_merges_anywhere(self):
        """A 1x1 merged range (openpyxl still writes <mergeCell ref="B6"/>) is
        invalid OOXML — Excel "repairs" the file on open (Removed Records: Merged
        Cells), stripping the table grid and showing the "Repaired" prompt. Guard
        that NO sheet emits a single-cell merge. Uses an indicator with single-row
        key-population groups (primary + age, no sex), which produced degenerate
        B/TOTAL-column merges before the _merge guard. Root-cause regression."""
        kp = SimpleNamespace(id=3, code="KP", name="KP indicator", type="number",
            aggregate_disaggregation_config={"enabled": True, "dimensions": [
                {"label": "Key Population", "key": "kp", "values": ["FSW", "MSM"]},
                {"label": "Age Range", "key": "age_band", "values": ["10-14", "15-19"]}]})
        project = SimpleNamespace(id=99, name="P", code="P", is_training=False)
        coord = SimpleNamespace(id=5, name="Coord", code="C")
        subA = SimpleNamespace(id=11, name="Sub A", code="A")
        subB = SimpleNamespace(id=12, name="Sub B", code="B")
        buf = rw.generate_coordinator_workbook(
            project=project, coordinator=coord,
            sub_specs=[(subA, [_plan(kp)]), (subB, [_plan(kp)])],
            coordinator_plans=[_plan(kp)], quarter=1, fiscal_start_year=2026, with_data=True)
        wb = load_workbook(BytesIO(buf.getvalue()))
        for ws in wb.worksheets:
            for m in ws.merged_cells.ranges:
                self.assertFalse(m.min_row == m.max_row and m.min_col == m.max_col,
                                 f"single-cell merge {m} on sheet {ws.title!r}")

    def test_indicator_only_in_one_sub_sums_only_that_sub(self):
        wb = self._build()
        cm = wb["_cellmap"]
        # find an IND2 cell on the total sheet
        total_title = rw.SHEET_TOTAL
        ind2_coord = None
        for r in cm.iter_rows(values_only=True):
            ind_id, code, kind, primary, secondary, band, coord = r
            if ind_id == 2 and kind == "cell" and str(coord).startswith(total_title + "!"):
                ind2_coord = str(coord).split("!", 1)[1]
                break
        self.assertIsNotNone(ind2_coord)
        formula = wb[total_title][ind2_coord].value
        self.assertIn("'Sub B'!", formula)
        self.assertNotIn("'Sub A'!", formula)  # Sub A doesn't report IND2
