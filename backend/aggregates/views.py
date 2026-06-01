from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from django_filters.rest_framework import DjangoFilterBackend
import django_filters
from rest_framework.filters import OrderingFilter
from django.conf import settings
from django.db import transaction
from django.db.models import Sum
from django.utils import timezone
from django.http import HttpResponse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from decimal import Decimal, InvalidOperation
from io import BytesIO

from .models import Aggregate
from .pagination import AggregatePagination
from .serializers import AggregateSerializer
from indicators.models import Indicator
from flags.models import Flag
from projects.models import Project
from projects.assignment_rules import (
    disaggregation_fields_by_indicator,
    is_indicator_assigned_to_organization,
    is_organization_in_project_scope,
)
from projects.hierarchy import resolve_organization_scope_with_project_hierarchy
from respondents.models import Response as InteractionResponse
from organizations.access import get_user_organization_ids, is_organization_admin, filter_queryset_by_org_ids, apply_training_filter, assert_project_write_allowed
from organizations.models import Organization
from respondents.rollups import sync_project_indicator_total
from users.models import User
from messaging.notifications import (
    AggregateNotificationContext,
    notify_aggregate_status_to_submitter,
    notify_aggregate_submitted_for_review,
)
from uploads.management.commands.import_reporting_workbook_overwrite import (
    IndicatorResolver as WorkbookIndicatorResolver,
    OrganizationResolver as WorkbookOrganizationResolver,
    get_age_band_mapping as workbook_get_age_band_mapping,
    normalize_text as workbook_normalize_text,
    snapshot_row_values as workbook_snapshot_row_values,
    row_contains_age_header as workbook_row_contains_age_header,
    SECTION_INDEX_PATTERN as WORKBOOK_SECTION_INDEX_PATTERN,
    TOTAL_COLUMNS as WORKBOOK_TOTAL_COLUMNS,
    resolve_section_age_band_mapping as workbook_resolve_section_age_band_mapping,
    is_skipped_sheet as workbook_is_skipped_sheet,
)


class AggregateFilterSet(django_filters.FilterSet):
    """Allow comma-separated status filters from the review queue UI."""

    status = django_filters.CharFilter(method="filter_status")
    organization = django_filters.CharFilter(method="filter_organization")
    coordinator = django_filters.CharFilter(method="filter_coordinator")
    include_org_descendants = django_filters.CharFilter(method="filter_include_org_descendants")
    date_from = django_filters.DateFilter(field_name="period_start", lookup_expr="gte")
    date_to = django_filters.DateFilter(field_name="period_end", lookup_expr="lte")

    class Meta:
        model = Aggregate
        fields = {
            "indicator": ["exact"],
            "project": ["exact"],
            "period_start": ["exact"],
            "period_end": ["exact"],
        }

    def _to_int(self, value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _org_scope_with_descendants(self, org_id: int, project_id: int | None = None):
        return list(resolve_organization_scope_with_project_hierarchy(org_id, project_id=project_id))

    def _is_truthy(self, value):
        return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on"}

    def filter_status(self, queryset, name, value):
        statuses = [item.strip() for item in str(value).split(",") if item.strip()]
        if not statuses:
            return queryset
        return queryset.filter(status__in=statuses)

    def filter_organization(self, queryset, name, value):
        raw_values = [item.strip() for item in str(value or "").split(",") if item.strip()]
        org_ids = []
        for raw in raw_values or [value]:
            org_id = self._to_int(raw)
            if org_id is not None:
                org_ids.append(org_id)
        if not org_ids:
            return queryset.none()
        include_descendants = self._is_truthy(self.data.get("include_org_descendants"))
        project_id_value = self._to_int(self.data.get("project"))
        if include_descendants:
            scoped_org_ids = []
            seen = set()
            for org_id in org_ids:
                for scoped_org_id in self._org_scope_with_descendants(org_id, project_id=project_id_value):
                    if scoped_org_id in seen:
                        continue
                    seen.add(scoped_org_id)
                    scoped_org_ids.append(scoped_org_id)
            if not scoped_org_ids:
                return queryset.none()
            return queryset.filter(organization_id__in=scoped_org_ids)
        return queryset.filter(organization_id__in=org_ids)

    def filter_coordinator(self, queryset, name, value):
        coordinator_id = self._to_int(value)
        if coordinator_id is None:
            return queryset.none()
        project_id_value = self._to_int(self.data.get("project"))
        scoped_org_ids = self._org_scope_with_descendants(coordinator_id, project_id=project_id_value)
        if not scoped_org_ids:
            return queryset.none()
        return queryset.filter(organization_id__in=scoped_org_ids)

    def filter_include_org_descendants(self, queryset, name, value):
        return queryset


class AggregateViewSet(viewsets.ModelViewSet):
    """ViewSet for managing aggregate data."""
    
    queryset = Aggregate.objects.all()
    serializer_class = AggregateSerializer
    pagination_class = AggregatePagination
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = AggregateFilterSet
    ordering_fields = ['period_start', 'period_end', 'created_at']
    ordering = ['-period_start']

    def _notification_context(self, aggregate: Aggregate) -> AggregateNotificationContext:
        indicator_label = (
            (getattr(aggregate.indicator, 'code', None) or '').strip()
            or (getattr(aggregate.indicator, 'name', None) or '').strip()
            or f"Indicator {aggregate.indicator_id}"
        )
        organization_name = (
            (getattr(aggregate.organization, 'name', None) or '').strip()
            or f"Organization {aggregate.organization_id}"
        )
        return AggregateNotificationContext(
            aggregate_id=int(aggregate.id),
            organization_id=int(aggregate.organization_id),
            indicator_label=indicator_label,
            organization_name=organization_name,
            period_start=aggregate.period_start.isoformat() if aggregate.period_start else '',
            period_end=aggregate.period_end.isoformat() if aggregate.period_end else '',
        )

    def _notify_pending_submission(self, aggregate: Aggregate) -> None:
        notify_aggregate_submitted_for_review(
            context=self._notification_context(aggregate),
            actor=self.request.user,
        )

    def _notify_status_change(self, aggregate: Aggregate, status_value: str) -> None:
        notify_aggregate_status_to_submitter(
            context=self._notification_context(aggregate),
            actor=self.request.user,
            submitter=aggregate.created_by,
            status_value=status_value,
        )
    
    def get_queryset(self):
        queryset = Aggregate.objects.select_related(
            'indicator', 'project', 'organization', 'created_by', 'reviewed_by'
        )
        user = self.request.user

        # Exclude training-project aggregates from all live views by default.
        # Admin users may opt-in with ?include_training=true.
        queryset = apply_training_filter(queryset, self.request, project_lookup="project")

        if is_organization_admin(user):
            return queryset
        org_ids = get_user_organization_ids(user)
        if org_ids:
            return filter_queryset_by_org_ids(queryset, 'organization_id', org_ids)
        return Aggregate.objects.none()

    def _allowed_org_ids_for_user(self):
        user = self.request.user
        if is_organization_admin(user):
            return None
        return set(get_user_organization_ids(user) or [])

    def _assert_write_scope(self, *, project, organization, indicator=None):
        user = self.request.user
        # Enforce the Sesigo Training/Live boundary before any other scope check.
        assert_project_write_allowed(self.request, project)
        allowed_org_ids = self._allowed_org_ids_for_user()
        if allowed_org_ids is not None:
            if not organization or organization.id not in allowed_org_ids:
                raise PermissionDenied('You do not have permission to submit aggregates for this organization.')
            project_visible = (
                project.created_by_id == user.id
                or project.organizations.filter(id__in=allowed_org_ids).exists()
            )
            if not project_visible:
                raise PermissionDenied('You do not have permission to submit aggregates for this project.')

        if not is_organization_in_project_scope(project, organization.id):
            raise PermissionDenied('Selected organization is not assigned to the selected project.')
        if indicator is not None:
            if not is_indicator_assigned_to_organization(
                project=project,
                indicator_id=indicator.id,
                organization_id=organization.id,
            ):
                raise PermissionDenied('Selected indicator is not assigned to this organization for the project.')
    
    def _upsert_pending_aggregate(
        self,
        *,
        indicator,
        project,
        organization,
        period_start,
        period_end,
        value,
        notes,
    ):
        existing = Aggregate.objects.filter(
            indicator=indicator,
            project=project,
            organization=organization,
            period_start=period_start,
            period_end=period_end,
        ).first()
        previous_status = existing.status if existing else None
        aggregate, created = Aggregate.objects.update_or_create(
            indicator=indicator,
            project=project,
            organization=organization,
            period_start=period_start,
            period_end=period_end,
            defaults={
                'value': value,
                'notes': notes or "",
                'status': 'pending',
                'reviewed_at': None,
                'reviewed_by': None,
                'created_by': self.request.user,
            },
        )
        if previous_status == 'approved':
            sync_project_indicator_total(aggregate.project_id, aggregate.indicator_id)
        return aggregate, created

    def perform_create(self, serializer):
        validated = serializer.validated_data
        self._assert_write_scope(
            project=validated['project'],
            organization=validated['organization'],
            indicator=validated['indicator'],
        )
        aggregate, _created = self._upsert_pending_aggregate(
            indicator=validated['indicator'],
            project=validated['project'],
            organization=validated['organization'],
            period_start=validated['period_start'],
            period_end=validated['period_end'],
            value=validated.get('value'),
            notes=validated.get('notes'),
        )
        serializer.instance = aggregate
        self._notify_pending_submission(aggregate)

    def perform_update(self, serializer):
        next_project = serializer.validated_data.get('project', serializer.instance.project)
        next_organization = serializer.validated_data.get('organization', serializer.instance.organization)
        next_indicator = serializer.validated_data.get('indicator', serializer.instance.indicator)
        self._assert_write_scope(
            project=next_project,
            organization=next_organization,
            indicator=next_indicator,
        )
        previous_status = serializer.instance.status
        aggregate = serializer.save(
            status='pending',
            reviewed_at=None,
            reviewed_by=None,
        )
        self._notify_pending_submission(aggregate)
        if previous_status == 'approved' or aggregate.status == 'approved':
            sync_project_indicator_total(aggregate.project_id, aggregate.indicator_id)

    def perform_destroy(self, instance):
        project_id = instance.project_id
        indicator_id = instance.indicator_id
        should_sync = instance.status == 'approved'
        super().perform_destroy(instance)
        if should_sync:
            sync_project_indicator_total(project_id, indicator_id)

    def _mark_review_state(self, aggregate, status_value):
        if status_value in {'reviewed', 'approved', 'rejected', 'flagged'} and not is_organization_admin(self.request.user):
            raise PermissionDenied('Only admins can review or approve aggregate records.')
        previous_status = aggregate.status
        aggregate.status = status_value
        if status_value == 'pending':
            aggregate.reviewed_at = None
            aggregate.reviewed_by = None
        else:
            aggregate.reviewed_at = timezone.now()
            aggregate.reviewed_by = self.request.user
        aggregate.save(update_fields=['status', 'reviewed_at', 'reviewed_by', 'updated_at'])
        self._notify_status_change(aggregate, status_value)
        if previous_status == 'approved' or status_value == 'approved':
            sync_project_indicator_total(aggregate.project_id, aggregate.indicator_id)
        return aggregate

    def _reporting_queryset(self):
        queryset = self.filter_queryset(self.get_queryset())
        status_filter = self.request.query_params.get('status')
        if status_filter:
            statuses = [value.strip() for value in status_filter.split(',') if value.strip()]
            if statuses:
                return queryset.filter(status__in=statuses)
        return queryset.filter(status='approved')

    def _extract_total(self, value):
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, dict):
            if value.get('total') is not None:
                return float(value.get('total') or 0)
            male = float(value.get('male') or 0)
            female = float(value.get('female') or 0)
            return male + female
        return 0.0

    def _normalize_match_tokens(self, raw_value):
        values = raw_value if isinstance(raw_value, list) else [raw_value]
        tokens = set()
        for value in values:
            if value is None:
                continue
            normalized = str(value).strip().lower()
            if normalized:
                tokens.add(normalized)
        return tokens

    def _response_matches(self, response_value, operator, match_tokens):
        if not match_tokens:
            return response_value not in (None, "", [], {})

        if isinstance(response_value, list):
            response_tokens = {
                str(item).strip().lower()
                for item in response_value
                if str(item).strip()
            }
            response_text = " ".join(sorted(response_tokens))
        elif isinstance(response_value, dict):
            response_tokens = {
                str(item).strip().lower()
                for item in response_value.values()
                if str(item).strip()
            }
            response_text = json.dumps(response_value, sort_keys=True).lower()
        else:
            normalized = str(response_value).strip().lower() if response_value is not None else ""
            response_tokens = {normalized} if normalized else set()
            response_text = normalized

        if operator == 'contains':
            return any(token in response_text for token in match_tokens)
        if operator == 'not_equals':
            return not any(token in response_tokens for token in match_tokens)
        return any(token in response_tokens for token in match_tokens)

    @action(detail=False, methods=['post'])
    def generate_from_interactions(self, request):
        """Compute an aggregate from respondent interactions and optionally save it."""
        output_indicator_id = request.data.get('output_indicator')
        source_indicator_id = request.data.get('source_indicator')
        project_id = request.data.get('project')
        organization_id = request.data.get('organization')
        period_start = request.data.get('period_start')
        period_end = request.data.get('period_end')
        operator = str(request.data.get('operator') or 'equals')
        match_value = request.data.get('match_value')
        count_distinct = str(request.data.get('count_distinct') or 'respondent')
        save_rule = bool(request.data.get('save_rule', False))
        save_aggregate = bool(request.data.get('save_aggregate', False))

        if not all([output_indicator_id, source_indicator_id, project_id, organization_id, period_start, period_end]):
            return Response(
                {'detail': 'output_indicator, source_indicator, project, organization, period_start, and period_end are required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if operator not in {'equals', 'not_equals', 'contains'}:
            return Response(
                {'detail': 'operator must be equals, not_equals, or contains.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if count_distinct not in {'respondent', 'interaction'}:
            return Response(
                {'detail': 'count_distinct must be respondent or interaction.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            organization_id_int = int(organization_id)
        except (TypeError, ValueError):
            return Response({'detail': 'organization must be a valid id.'}, status=status.HTTP_400_BAD_REQUEST)

        allowed_org_ids = set(get_user_organization_ids(request.user) or [])
        if not is_organization_admin(request.user) and organization_id_int not in allowed_org_ids:
            return Response(
                {'detail': 'You do not have permission to generate aggregates for this organization.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            output_indicator = Indicator.objects.get(id=output_indicator_id)
        except Indicator.DoesNotExist:
            return Response({'detail': 'Output indicator not found.'}, status=status.HTTP_404_NOT_FOUND)

        if not Indicator.objects.filter(id=source_indicator_id).exists():
            return Response({'detail': 'Source indicator not found.'}, status=status.HTTP_404_NOT_FOUND)

        if not Project.objects.filter(id=project_id).exists():
            return Response({'detail': 'Project not found.'}, status=status.HTTP_404_NOT_FOUND)

        match_tokens = self._normalize_match_tokens(match_value)
        responses = InteractionResponse.objects.select_related(
            'interaction',
            'interaction__respondent',
        ).filter(
            indicator_id=source_indicator_id,
            interaction__project_id=project_id,
            interaction__respondent__organization_id=organization_id_int,
            interaction__date__gte=period_start,
            interaction__date__lte=period_end,
        )

        matched_responses = [
            response
            for response in responses
            if self._response_matches(response.value, operator, match_tokens)
        ]

        if count_distinct == 'interaction':
            computed = len({response.interaction_id for response in matched_responses})
        else:
            computed = len({response.interaction.respondent_id for response in matched_responses})

        aggregate_payload = None
        project = Project.objects.get(id=project_id)
        organization = Organization.objects.get(id=organization_id_int)

        self._assert_write_scope(
            project=project,
            organization=organization,
            indicator=output_indicator,
        )
        if not is_indicator_assigned_to_organization(
            project=project,
            indicator_id=int(source_indicator_id),
            organization_id=organization.id,
        ):
            return Response(
                {'detail': 'Source indicator is not assigned to this organization for the project.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if save_aggregate:
            aggregate, _created = self._upsert_pending_aggregate(
                indicator=output_indicator,
                project=project,
                organization=organization,
                period_start=period_start,
                period_end=period_end,
                value={'total': computed},
                notes=(
                    f"Auto-calculated from source indicator {source_indicator_id} "
                    f"using operator '{operator}' and distinct-by '{count_distinct}'."
                ),
            )
            self._notify_pending_submission(aggregate)
            aggregate_payload = AggregateSerializer(aggregate).data

        return Response(
            {
                'computed': computed,
                'rule': {
                    'output_indicator': int(output_indicator_id),
                    'output_indicator_code': output_indicator.code,
                    'source_indicator': int(source_indicator_id),
                    'operator': operator,
                    'match_value': match_value,
                    'count_distinct': count_distinct,
                    'project': int(project_id),
                    'organization': organization_id_int,
                    'period_start': period_start,
                    'period_end': period_end,
                    'save_rule_requested': save_rule,
                },
                'aggregate': aggregate_payload,
            }
        )
    
    @action(detail=False, methods=['get'])
    def by_indicator(self, request):
        """Get aggregates grouped by indicator."""
        indicator_id = request.query_params.get('indicator_id')
        if not indicator_id:
            return Response({'error': 'indicator_id required'}, status=400)
        
        aggregates = self._reporting_queryset().filter(indicator_id=indicator_id)
        return Response(AggregateSerializer(aggregates, many=True).data)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Bulk create aggregates."""
        project_id = request.data.get('project')
        organization_id = request.data.get('organization')
        period_start = request.data.get('period_start')
        period_end = request.data.get('period_end')
        data = request.data.get('data', [])

        if not project_id or not organization_id or not period_start or not period_end:
            return Response(
                {'error': 'project, organization, period_start, period_end required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not isinstance(data, list) or not data:
            return Response({'error': 'data list required'}, status=status.HTTP_400_BAD_REQUEST)

        project = Project.objects.filter(id=project_id).first()
        organization = Organization.objects.filter(id=organization_id).first()
        if not project or not organization:
            return Response({'error': 'project and organization must reference existing records'}, status=status.HTTP_400_BAD_REQUEST)
        self._assert_write_scope(project=project, organization=organization)

        results = []
        created_count = 0
        updated_count = 0
        try:
            with transaction.atomic():
                for item in data:
                    serializer = AggregateSerializer(data={
                        'indicator': item.get('indicator'),
                        'project': project_id,
                        'organization': organization_id,
                        'period_start': period_start,
                        'period_end': period_end,
                        'value': item.get('value'),
                        'notes': item.get('notes'),
                    })
                    serializer.is_valid(raise_exception=True)
                    validated = serializer.validated_data
                    self._assert_write_scope(
                        project=validated['project'],
                        organization=validated['organization'],
                        indicator=validated['indicator'],
                    )
                    aggregate, created = self._upsert_pending_aggregate(
                        indicator=validated['indicator'],
                        project=validated['project'],
                        organization=validated['organization'],
                        period_start=validated['period_start'],
                        period_end=validated['period_end'],
                        value=validated.get('value'),
                        notes=validated.get('notes'),
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                    self._notify_pending_submission(aggregate)
                    results.append(AggregateSerializer(aggregate).data)
        except Exception as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                'created': created_count,
                'updated': updated_count,
                'results': results,
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['get'])
    def templates(self, request):
        """Return aggregate templates (indicator sets)."""
        project_id = request.query_params.get('project')
        organization_id = request.query_params.get('organization')
        indicators = Indicator.objects.filter(is_active=True)
        organization_id_value = None
        if organization_id not in (None, ""):
            try:
                organization_id_value = int(organization_id)
            except (TypeError, ValueError):
                organization_id_value = None

        if project_id and Project.objects.filter(id=project_id).exists():
            project = Project.objects.filter(id=project_id).first()
            indicators = indicators.filter(projects__id=project_id).distinct()
            if project is not None and organization_id_value is not None:
                assigned_ids = {
                    int(indicator_id)
                    for indicator_id in indicators.values_list('id', flat=True)
                    if is_indicator_assigned_to_organization(
                        project=project,
                        indicator_id=indicator_id,
                        organization_id=organization_id_value,
                    )
                }
                if assigned_ids:
                    indicators = indicators.filter(id__in=assigned_ids)
                else:
                    indicators = indicators.none()
            template_name = f"Project {project_id} Indicators"
        elif organization_id:
            indicators = indicators.filter(organizations__id=organization_id).distinct()
            template_name = "Organization Indicators"
        else:
            template_name = "All Indicators"

        disaggregation_fields_map = {}
        if project_id and Project.objects.filter(id=project_id).exists():
            project = Project.objects.filter(id=project_id).first()
            if project is not None:
                disaggregation_fields_map = disaggregation_fields_by_indicator(
                    project=project,
                    organization_id=organization_id_value,
                    indicator_ids=list(indicators.values_list('id', flat=True)),
                )

        payload = [{
            'id': 1,
            'name': template_name,
            'indicators': [
                {
                    'id': indicator.id,
                    'name': indicator.name,
                    'code': indicator.code,
                    'type': indicator.type,
                    'disaggregation_fields': disaggregation_fields_map.get(indicator.id, []),
                }
                for indicator in indicators
            ],
        }]
        return Response(payload)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get aggregate summary by indicator."""
        queryset = self._reporting_queryset()
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(period_start__gte=date_from)
        if date_to:
            queryset = queryset.filter(period_end__lte=date_to)
        totals = {}
        counts = {}
        for agg in queryset:
            totals[agg.indicator_id] = totals.get(agg.indicator_id, 0.0) + self._extract_total(agg.value)
            counts[agg.indicator_id] = counts.get(agg.indicator_id, 0) + 1

        indicators = Indicator.objects.filter(id__in=totals.keys())
        results = []
        for indicator in indicators:
            results.append({
                'indicator_id': indicator.id,
                'indicator_name': indicator.name,
                'total_value': totals.get(indicator.id, 0.0),
                'period_count': counts.get(indicator.id, 0),
                'trend': 'stable',
            })
        return Response(results)

    def _note_template_metadata(self, notes: str):
        text = str(notes or "").strip()
        if not text:
            return None
        match = re.search(
            r"Imported from (?P<workbook>.+?)\s*\|\s*sheet=(?P<sheet>[^|]+?)\s*\|",
            text,
            flags=re.IGNORECASE,
        )
        if not match:
            return None
        workbook_name = str(match.group("workbook") or "").strip()
        sheet_name = str(match.group("sheet") or "").strip()
        if not workbook_name:
            return None
        return {
            "workbook": workbook_name,
            "sheet": sheet_name,
        }

    def _normalized_filename_token(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())

    def _template_search_roots(self):
        project_root = Path(settings.BASE_DIR).parent
        return [
            project_root / "imports",
            Path(settings.MEDIA_ROOT) / "uploads",
        ]

    def _collect_template_files(self):
        files = []
        for root in self._template_search_roots():
            if not root.exists():
                continue
            for pattern in ("**/*.xlsx", "**/*.xlsm", "**/*.xls"):
                files.extend(root.glob(pattern))
        return files

    def _resolve_template_workbook_path(self, workbook_name: str, indexed_files):
        workbook_name = str(workbook_name or "").strip()
        if not workbook_name:
            return None
        for root in self._template_search_roots():
            direct = root / workbook_name
            if direct.exists():
                return direct

        target = self._normalized_filename_token(workbook_name)
        if not target:
            return None
        exact_matches = []
        partial_matches = []
        for file_path in indexed_files:
            file_token = self._normalized_filename_token(file_path.name)
            if not file_token:
                continue
            if file_token == target:
                exact_matches.append(file_path)
            elif target in file_token or file_token in target:
                partial_matches.append(file_path)
        if exact_matches:
            return sorted(exact_matches, key=lambda path: len(path.name))[0]
        if partial_matches:
            return sorted(partial_matches, key=lambda path: len(path.name))[0]
        return None

    def _select_fallback_template_workbook(self, indexed_files):
        if not indexed_files:
            return None
        cluster_candidates = [
            file_path
            for file_path in indexed_files
            if "cluster" in file_path.name.lower() and "statistics" in file_path.name.lower()
        ]
        if cluster_candidates:
            return sorted(cluster_candidates, key=lambda path: path.stat().st_size, reverse=True)[0]
        return sorted(indexed_files, key=lambda path: path.stat().st_size, reverse=True)[0]

    def _safe_sheet_title(self, workbook, raw_title: str):
        cleaned = re.sub(r'[:/\\?*\[\]]+', ' ', str(raw_title or '').strip())
        cleaned = re.sub(r'\s+', ' ', cleaned).strip() or "Organization"
        base = cleaned[:31]
        candidate = base
        suffix = 2
        existing_titles = set(workbook.sheetnames)
        while candidate in existing_titles:
            suffix_text = f" ({suffix})"
            candidate = f"{base[:max(1, 31 - len(suffix_text))]}{suffix_text}"
            suffix += 1
        return candidate

    def _parse_template_sections_with_rows(self, ws_values, ws_formula):
        sections = []
        current = None
        current_age_header = None
        value_rows = ws_values.iter_rows(min_row=1, min_col=2, max_col=27, values_only=True)
        formula_rows = ws_formula.iter_rows(min_row=1, min_col=2, max_col=27, values_only=True)

        for row_number, (row_values, formula_row_values) in enumerate(zip(value_rows, formula_rows), start=1):
            row_snapshot = workbook_snapshot_row_values(row_values, formula_row_values)
            if workbook_row_contains_age_header(row_snapshot):
                current_age_header = row_snapshot

            index_value = str(row_snapshot.get("B") or "").strip()
            title = str(row_snapshot.get("C") or "").strip()
            is_section_start = bool(WORKBOOK_SECTION_INDEX_PATTERN.fullmatch(index_value))

            if is_section_start:
                if current:
                    sections.append(current)
                current = {
                    "index": index_value,
                    "title": title,
                    "header_row": current_age_header,
                    "rows": [{"row_number": row_number, "row": row_snapshot}],
                }
                continue

            if current:
                if not current["title"] and title:
                    current["title"] = title
                current["rows"].append({"row_number": row_number, "row": row_snapshot})

        if current:
            sections.append(current)
        return sections

    def _decimal_number(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    def _json_number(self, value: Decimal | None):
        if value is None:
            return None
        return int(value) if value == value.to_integral_value() else float(value)

    def _normalize_row_label(self, value: str):
        return workbook_normalize_text(value).replace("-", " ").strip()

    def _is_formula_cell(self, ws, column: str, row_number: int):
        cell = ws[f"{column}{row_number}"]
        return getattr(cell, "data_type", None) == "f" or (
            isinstance(cell.value, str) and str(cell.value).startswith("=")
        )

    def _pick_writable_total_column(self, ws, row_snapshot: dict, row_number: int):
        for column in WORKBOOK_TOTAL_COLUMNS:
            if self._is_formula_cell(ws, column, row_number):
                continue
            value = row_snapshot.get(column)
            if value not in (None, ""):
                return column
        for column in WORKBOOK_TOTAL_COLUMNS:
            if not self._is_formula_cell(ws, column, row_number):
                return column
        return None

    def _write_number(self, ws, column: str, row_number: int, value: Decimal | None):
        if column is None:
            return
        if self._is_formula_cell(ws, column, row_number):
            return
        ws[f"{column}{row_number}"] = self._json_number(value if value is not None else Decimal("0"))

    def _apply_payload_to_template_section(self, ws, section: dict, payload: dict, age_band_by_column: dict[str, str]):
        if not isinstance(payload, dict):
            return

        disaggregates = payload.get("disaggregates") if isinstance(payload.get("disaggregates"), dict) else {}
        normalized_disaggregate_keys = {
            self._normalize_row_label(key): str(key)
            for key in disaggregates.keys()
            if str(key).strip()
        }
        normalized_disaggregate_keys.setdefault("all", normalized_disaggregate_keys.get("all", "All"))
        section_age_mapping = workbook_resolve_section_age_band_mapping(section, age_band_by_column)

        male_total_from_rows = Decimal("0")
        female_total_from_rows = Decimal("0")

        current_primary_label = ""
        for row_item in section.get("rows", []):
            row_number = row_item.get("row_number")
            row_snapshot = row_item.get("row") or {}
            if not row_number:
                continue

            raw_primary = str(row_snapshot.get("E") or "").strip()
            raw_sex = str(row_snapshot.get("F") or "").strip()
            normalized_primary = self._normalize_row_label(raw_primary)
            normalized_sex = self._normalize_row_label(raw_sex)

            if raw_primary and normalized_primary not in {"total", "sub total", "subtotal", "total male", "total female"}:
                current_primary_label = raw_primary
            primary_label = raw_primary or current_primary_label or "All"

            disaggregate_key = normalized_disaggregate_keys.get(
                self._normalize_row_label(primary_label),
                normalized_disaggregate_keys.get("all"),
            )

            if disaggregates and disaggregate_key and normalized_sex in {"male", "female"}:
                sex_label = "Male" if normalized_sex == "male" else "Female"
                band_values = (
                    disaggregates.get(disaggregate_key, {}).get(sex_label, {})
                    if isinstance(disaggregates.get(disaggregate_key), dict)
                    else {}
                )
                if isinstance(band_values, dict) and band_values:
                    row_sum = Decimal("0")
                    for column, age_band in section_age_mapping.items():
                        age_key = str(age_band)
                        cell_value = band_values.get(age_key)
                        number = self._decimal_number(cell_value)
                        if number is None:
                            continue
                        self._write_number(ws, column, row_number, number)
                        row_sum += number
                    total_column = self._pick_writable_total_column(ws, row_snapshot, row_number)
                    self._write_number(ws, total_column, row_number, row_sum)
                    if sex_label == "Male":
                        male_total_from_rows += row_sum
                    else:
                        female_total_from_rows += row_sum
                    continue

            if normalized_primary == "total male":
                total_column = self._pick_writable_total_column(ws, row_snapshot, row_number)
                self._write_number(
                    ws,
                    total_column,
                    row_number,
                    self._decimal_number(payload.get("male")) if payload.get("male") is not None else male_total_from_rows,
                )
                continue

            if normalized_primary == "total female":
                total_column = self._pick_writable_total_column(ws, row_snapshot, row_number)
                self._write_number(
                    ws,
                    total_column,
                    row_number,
                    self._decimal_number(payload.get("female")) if payload.get("female") is not None else female_total_from_rows,
                )
                continue

            if normalized_primary in {"total", "sub total", "subtotal"}:
                total_column = self._pick_writable_total_column(ws, row_snapshot, row_number)
                self._write_number(ws, total_column, row_number, self._decimal_number(payload.get("total")))
                continue

        if not section.get("rows"):
            return

        first_row = section["rows"][0]
        first_row_number = first_row.get("row_number")
        first_row_snapshot = first_row.get("row") or {}
        if not first_row_number:
            return
        fallback_total_column = self._pick_writable_total_column(ws, first_row_snapshot, first_row_number)
        self._write_number(ws, fallback_total_column, first_row_number, self._decimal_number(payload.get("total")))

    def _try_export_template_workbook(self, queryset):
        rows = list(queryset)
        if not rows:
            return None

        project_ids = {row.project_id for row in rows if row.project_id}
        if len(project_ids) != 1:
            return None
        selected_org_ids = {row.organization_id for row in rows if row.organization_id}
        if not selected_org_ids:
            return None
        project_id = next(iter(project_ids))
        project = Project.objects.filter(id=project_id).first()
        if not project:
            return None

        workbook_counter = Counter()
        for row in rows:
            metadata = self._note_template_metadata(row.notes or "")
            if not metadata:
                continue
            workbook_counter[metadata["workbook"]] += 1

        indexed_files = self._collect_template_files()
        if not indexed_files:
            return None

        template_path = None
        if workbook_counter:
            for workbook_name, _count in workbook_counter.most_common():
                template_path = self._resolve_template_workbook_path(workbook_name, indexed_files)
                if template_path:
                    break
        if not template_path:
            if len(selected_org_ids) == 1:
                template_path = self._select_fallback_template_workbook(indexed_files)
            else:
                return None
        if not template_path:
            return None

        try:
            import openpyxl
        except Exception:
            return None

        keep_vba = template_path.suffix.lower() == ".xlsm"
        wb_values = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
        wb_formula = openpyxl.load_workbook(template_path, data_only=False, keep_vba=keep_vba)

        aggregate_value_by_org_indicator = {}
        for row in rows:
            value = row.value
            if isinstance(value, dict):
                payload = value
            elif isinstance(value, (int, float)):
                payload = {"total": value}
            else:
                payload = {}
            aggregate_value_by_org_indicator[(row.organization_id, row.indicator_id)] = payload
        total_payload_count = len(aggregate_value_by_org_indicator)
        applied_payload_keys = set()

        indicator_resolver = WorkbookIndicatorResolver(project=project)
        organization_resolver = WorkbookOrganizationResolver()
        organization_by_id = {
            organization.id: organization
            for organization in Organization.objects.filter(id__in={row.organization_id for row in rows})
        }
        selected_org_ids = set(organization_by_id.keys())
        filled_org_ids = set()
        filled_sheet_titles = set()

        try:
            template_sheet_for_clone = None
            for sheet_name in list(wb_formula.sheetnames):
                if workbook_is_skipped_sheet(sheet_name):
                    continue
                if sheet_name not in wb_values.sheetnames:
                    continue
                if template_sheet_for_clone is None:
                    template_sheet_for_clone = sheet_name
                organization = organization_resolver.resolve(sheet_name)
                if not organization:
                    continue
                if organization.id not in organization_by_id:
                    continue

                ws_values = wb_values[sheet_name]
                ws_formula = wb_formula[sheet_name]
                sections = self._parse_template_sections_with_rows(ws_values, ws_formula)
                age_band_by_column = workbook_get_age_band_mapping(
                    organization_by_id.get(organization.id),
                    None,
                )

                for section in sections:
                    title = str(section.get("title") or "").strip()
                    if not title:
                        continue
                    indicator = indicator_resolver.resolve(title, section.get("index"))
                    if not indicator:
                        continue
                    payload = aggregate_value_by_org_indicator.get((organization.id, indicator.id))
                    if not payload:
                        continue
                    self._apply_payload_to_template_section(
                        ws_formula,
                        section,
                        payload,
                        age_band_by_column=age_band_by_column,
                    )
                    applied_payload_keys.add((organization.id, indicator.id))
                filled_org_ids.add(organization.id)
                filled_sheet_titles.add(sheet_name)

            if template_sheet_for_clone and template_sheet_for_clone in wb_formula.sheetnames:
                source_formula_sheet = wb_formula[template_sheet_for_clone]
                source_values_sheet = (
                    wb_values[template_sheet_for_clone]
                    if template_sheet_for_clone in wb_values.sheetnames
                    else None
                )
                source_sections = (
                    self._parse_template_sections_with_rows(source_values_sheet, source_formula_sheet)
                    if source_values_sheet is not None
                    else self._parse_template_sections_with_rows(source_formula_sheet, source_formula_sheet)
                )

                for organization_id in sorted(selected_org_ids - filled_org_ids):
                    organization = organization_by_id.get(organization_id)
                    if not organization:
                        continue
                    cloned_sheet = wb_formula.copy_worksheet(source_formula_sheet)
                    cloned_sheet.title = self._safe_sheet_title(wb_formula, organization.name or f"Org {organization.id}")
                    age_band_by_column = workbook_get_age_band_mapping(organization, None)

                    for section in source_sections:
                        title = str(section.get("title") or "").strip()
                        if not title:
                            continue
                        indicator = indicator_resolver.resolve(title, section.get("index"))
                        if not indicator:
                            continue
                        payload = aggregate_value_by_org_indicator.get((organization.id, indicator.id))
                        if not payload:
                            continue
                        self._apply_payload_to_template_section(
                            cloned_sheet,
                            section,
                            payload,
                            age_band_by_column=age_band_by_column,
                        )
                        applied_payload_keys.add((organization.id, indicator.id))
                    filled_org_ids.add(organization.id)
                    filled_sheet_titles.add(cloned_sheet.title)

            for sheet_name in list(wb_formula.sheetnames):
                if sheet_name in filled_sheet_titles:
                    continue
                if workbook_is_skipped_sheet(sheet_name):
                    continue
                organization = organization_resolver.resolve(sheet_name)
                if organization and organization.id not in selected_org_ids:
                    del wb_formula[sheet_name]

            coverage_ratio = (
                (len(applied_payload_keys) / total_payload_count)
                if total_payload_count
                else 0.0
            )
            if coverage_ratio < 0.5:
                return None

            output = BytesIO()
            wb_formula.save(output)
            output.seek(0)
            return output.read(), template_path.name
        finally:
            wb_values.close()
            if hasattr(wb_formula, "close"):
                wb_formula.close()

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export aggregates to CSV or Excel."""
        fmt = request.query_params.get('file_format', 'csv').lower()
        if fmt not in {'csv', 'excel'}:
            fmt = 'csv'
        sheet_layout = str(request.query_params.get('sheet_layout', '') or '').strip().lower()
        organization_sheet_layout = sheet_layout in {'organization', 'per_organization', 'by_organization'}
        template_sheet_layout = sheet_layout in {'template', 'organization_template', 'workbook_template'}
        queryset = self._reporting_queryset().select_related('indicator', 'project', 'organization')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(period_start__gte=date_from)
        if date_to:
            queryset = queryset.filter(period_end__lte=date_to)

        rows = []
        for agg in queryset:
            value = agg.value
            male = None
            female = None
            total = None
            if isinstance(value, dict):
                male = value.get('male')
                female = value.get('female')
                total = value.get('total')
            elif isinstance(value, (int, float)):
                total = value

            rows.append({
                'indicator_id': agg.indicator_id or '',
                'indicator_name': agg.indicator.name if agg.indicator_id else '',
                'indicator_code': agg.indicator.code if agg.indicator_id else '',
                'project_id': agg.project_id or '',
                'project_name': agg.project.name if agg.project_id else '',
                'organization_id': agg.organization_id or '',
                'organization_name': agg.organization.name if agg.organization_id else '',
                'period_start': agg.period_start.isoformat(),
                'period_end': agg.period_end.isoformat(),
                'male': male if male is not None else '',
                'female': female if female is not None else '',
                'total': total if total is not None else '',
                'value_json': json.dumps(value, ensure_ascii=False) if value is not None else '',
                'notes': agg.notes or '',
            })

        if fmt == 'excel':
            try:
                import openpyxl
            except Exception as exc:
                return Response({'error': f'Excel export not available: {exc}'}, status=500)

            if template_sheet_layout:
                template_export = self._try_export_template_workbook(queryset)
                if template_export:
                    payload_bytes, source_filename = template_export
                    source_suffix = Path(source_filename).suffix.lower() or ".xlsx"
                    content_type = (
                        'application/vnd.ms-excel.sheet.macroEnabled.12'
                        if source_suffix == '.xlsm'
                        else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response = HttpResponse(
                        payload_bytes,
                        content_type=content_type,
                    )
                    output_name = f"aggregates_template_export_{Path(source_filename).stem}{source_suffix}"
                    response['Content-Disposition'] = f'attachment; filename="{output_name}"'
                    return response

            def _sheet_title(value: str, seen: set[str]) -> str:
                cleaned = re.sub(r'[:/\\?*\[\]]+', ' ', str(value or '').strip())
                cleaned = re.sub(r'\s+', ' ', cleaned).strip()
                base = cleaned or 'Organization'
                title = base[:31]
                suffix = 2
                while title in seen:
                    suffix_text = f' ({suffix})'
                    title = f"{base[:max(1, 31 - len(suffix_text))]}{suffix_text}"
                    suffix += 1
                seen.add(title)
                return title

            wb = openpyxl.Workbook()
            columns = list(rows[0].keys()) if rows else [
                'indicator_id',
                'indicator_name',
                'indicator_code',
                'project_id',
                'project_name',
                'organization_id',
                'organization_name',
                'period_start',
                'period_end',
                'male',
                'female',
                'total',
                'value_json',
                'notes',
            ]

            if organization_sheet_layout and rows:
                grouped_rows = {}
                for row in rows:
                    org_name = str(row.get('organization_name') or f"Organization {row.get('organization_id') or ''}").strip()
                    grouped_rows.setdefault(org_name, []).append(row)

                sheet_names_seen = set()
                active_sheet = wb.active
                first_sheet = True
                for org_name in sorted(grouped_rows.keys(), key=lambda value: value.lower()):
                    if first_sheet:
                        ws = active_sheet
                        first_sheet = False
                    else:
                        ws = wb.create_sheet()
                    ws.title = _sheet_title(org_name, sheet_names_seen)
                    ws.append(columns)
                    for row in grouped_rows[org_name]:
                        ws.append([row.get(column, '') for column in columns])
            else:
                ws = wb.active
                ws.title = 'aggregates'
                ws.append(columns)
                for row in rows:
                    ws.append([row.get(column, '') for column in columns])

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="aggregates.xlsx"'
            return response

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="aggregates.csv"'
        writer = csv.writer(response)
        if rows:
            writer.writerow(rows[0].keys())
            for row in rows:
                writer.writerow(row.values())
        return response

    @action(detail=True, methods=['post'])
    def review(self, request, pk=None):
        """Mark an aggregate as reviewed and send it back through the approval queue."""
        aggregate = self.get_object()
        return Response(AggregateSerializer(self._mark_review_state(aggregate, 'reviewed')).data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve an aggregate so it appears in the main aggregate view."""
        aggregate = self.get_object()
        return Response(AggregateSerializer(self._mark_review_state(aggregate, 'approved')).data)

    @action(detail=False, methods=['post'])
    def bulk_approve(self, request):
        """Approve multiple queued aggregates at once."""
        if not is_organization_admin(request.user):
            return Response({'error': 'admin access required'}, status=status.HTTP_403_FORBIDDEN)

        raw_ids = request.data.get('ids', [])
        if not isinstance(raw_ids, list) or not raw_ids:
            return Response({'error': 'ids list required'}, status=status.HTTP_400_BAD_REQUEST)

        aggregate_ids = []
        for value in raw_ids:
            try:
                aggregate_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        if not aggregate_ids:
            return Response({'error': 'ids list required'}, status=status.HTTP_400_BAD_REQUEST)

        approved_count = 0
        skipped = 0
        affected_pairs = set()
        reviewed_at = timezone.now()
        with transaction.atomic():
            aggregate_rows = list(
                self.get_queryset().filter(
                    id__in=aggregate_ids,
                    status__in=['pending', 'reviewed'],
                ).values(
                    'id',
                    'project_id',
                    'indicator_id',
                    'created_by_id',
                    'organization_id',
                    'organization__name',
                    'indicator__code',
                    'indicator__name',
                    'period_start',
                    'period_end',
                )
            )

            found_ids = {row['id'] for row in aggregate_rows}
            affected_pairs = {
                (row['project_id'], row['indicator_id'])
                for row in aggregate_rows
            }

            if found_ids:
                approved_count = Aggregate.objects.filter(id__in=found_ids).update(
                    status='approved',
                    reviewed_at=reviewed_at,
                    reviewed_by_id=request.user.id,
                    updated_at=reviewed_at,
                )

            skipped = len(set(aggregate_ids) - found_ids)

        for project_id, indicator_id in affected_pairs:
            sync_project_indicator_total(project_id, indicator_id)

        for row in aggregate_rows:
            submitter_id = row.get('created_by_id')
            if not submitter_id or submitter_id == request.user.id:
                continue
            submitter = User.objects.filter(id=submitter_id, is_active=True).first()
            if not submitter:
                continue
            indicator_label = (row.get('indicator__code') or row.get('indicator__name') or '').strip() or f"Indicator {row.get('indicator_id')}"
            organization_name = (row.get('organization__name') or '').strip() or f"Organization {row.get('organization_id')}"
            context = AggregateNotificationContext(
                aggregate_id=int(row['id']),
                organization_id=int(row['organization_id']),
                indicator_label=indicator_label,
                organization_name=organization_name,
                period_start=row['period_start'].isoformat() if row.get('period_start') else '',
                period_end=row['period_end'].isoformat() if row.get('period_end') else '',
            )
            notify_aggregate_status_to_submitter(
                context=context,
                actor=request.user,
                submitter=submitter,
                status_value='approved',
            )

        return Response(
            {
                'approved': approved_count,
                'skipped': skipped,
                'results': [],
            }
        )

    @action(detail=True, methods=['post'])
    def flag(self, request, pk=None):
        """Flag an aggregate for data quality review."""
        aggregate = self.get_object()
        reason = str(request.data.get('reason') or 'other')
        description = str(request.data.get('description') or '').strip()
        severity = str(request.data.get('severity') or 'medium')
        priority_map = {
            'low': 'low',
            'medium': 'medium',
            'high': 'high',
            'critical': 'critical',
        }
        flag = Flag.objects.create(
            flag_type='data_quality',
            status='open',
            priority=priority_map.get(severity, 'medium'),
            title=f"Aggregate review flag: {aggregate.indicator.code or aggregate.indicator.name}",
            description=description or f"Aggregate flagged during review ({reason}).",
            content_type='aggregate',
            object_id=aggregate.id,
            organization=aggregate.organization,
            created_by=request.user,
        )
        flagged_aggregate = self._mark_review_state(aggregate, 'flagged')
        response_data = AggregateSerializer(flagged_aggregate).data
        response_data['flag_id'] = flag.id
        return Response(response_data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject an aggregate so it does not appear in aggregate reporting."""
        aggregate = self.get_object()
        return Response(AggregateSerializer(self._mark_review_state(aggregate, 'rejected')).data)

    @action(detail=True, methods=['post'])
    def unflag(self, request, pk=None):
        """Dismiss open data-quality flags on an aggregate and restore it to approved."""
        aggregate = self.get_object()
        notes = str(request.data.get('notes') or '').strip()
        with transaction.atomic():
            open_flags = Flag.objects.filter(
                content_type='aggregate',
                object_id=aggregate.id,
                status='open',
            )
            for flag in open_flags:
                flag.status = 'dismissed'
                flag.resolution_notes = notes or 'Verified correct by admin.'
                flag.resolved_at = timezone.now()
                flag.resolved_by = request.user
                flag.save(update_fields=['status', 'resolution_notes', 'resolved_at', 'resolved_by'])
            aggregate.status = 'approved'
            aggregate.copy_paste_verified = True
            aggregate.save(update_fields=['status', 'copy_paste_verified', 'updated_at'])
        return Response(AggregateSerializer(aggregate).data)

    @action(detail=False, methods=['post'])
    def detect_copy_paste(self, request):
        """Detect and flag org-period groups where 5+ indicators share the same value.
        Skips any aggregate that has been marked copy_paste_verified=True by an admin."""
        min_count = max(2, int(request.data.get('min_count') or 5))
        tolerance = min(0.5, float(request.data.get('tolerance') or 0.05))
        project_id = request.data.get('project_id')
        dry_run = bool(request.data.get('dry_run'))

        qs = Aggregate.objects.filter(
            status__in=['approved', 'pending', 'reviewed'],
            copy_paste_verified=False,
        ).select_related('indicator', 'organization', 'project')
        if project_id:
            qs = qs.filter(project_id=project_id)

        from collections import defaultdict
        groups = defaultdict(list)
        for agg in qs:
            total = self._extract_total(agg.value)
            if total > 0:
                key = (agg.organization_id, agg.project_id, agg.period_start, agg.period_end)
                groups[key].append((agg, total))

        flagged_groups = []
        for (org_id, proj_id, period_start, period_end), agg_totals in groups.items():
            if len(agg_totals) < min_count:
                continue
            sorted_totals = sorted(t for _, t in agg_totals)
            median = sorted_totals[len(sorted_totals) // 2]
            if median < 10:
                continue
            within = [(agg, t) for agg, t in agg_totals if abs(t - median) / median <= tolerance]
            if len(within) >= min_count:
                flagged_groups.append({
                    'org': within[0][0].organization,
                    'period_start': period_start,
                    'period_end': period_end,
                    'median': median,
                    'aggregates': [agg for agg, _ in within],
                })

        if dry_run:
            return Response({
                'total_groups': len(flagged_groups),
                'total_aggregates': sum(len(g['aggregates']) for g in flagged_groups),
                'groups': [
                    {
                        'org': g['org'].name,
                        'period': f"{g['period_start']} – {g['period_end']}",
                        'median': g['median'],
                        'count': len(g['aggregates']),
                        'indicators': [a.indicator.code or a.indicator.name for a in g['aggregates']],
                    }
                    for g in flagged_groups
                ],
            })

        created_flags = 0
        flagged_aggs = 0
        with transaction.atomic():
            for group in flagged_groups:
                org = group['org']
                indicator_codes = ', '.join(
                    a.indicator.code or a.indicator.name for a in group['aggregates']
                )
                cluster_desc = (
                    f"Copy-paste pattern: {len(group['aggregates'])} indicators all report "
                    f"≈{group['median']:.0f} for {org.name} "
                    f"({group['period_start']} – {group['period_end']}). "
                    f"Indicators: {indicator_codes}"
                )
                for agg in group['aggregates']:
                    if agg.status == 'flagged':
                        continue
                    already = Flag.objects.filter(
                        content_type='aggregate',
                        object_id=agg.id,
                        flag_type='data_quality',
                        status='open',
                    ).exists()
                    if already:
                        continue
                    Flag.objects.create(
                        flag_type='data_quality',
                        status='open',
                        priority='high',
                        title=f"Copy-paste pattern: {agg.indicator.code or agg.indicator.name}",
                        description=cluster_desc,
                        content_type='aggregate',
                        object_id=agg.id,
                        organization=org,
                        created_by=request.user,
                    )
                    agg.status = 'flagged'
                    agg.save(update_fields=['status', 'updated_at'])
                    created_flags += 1
                    flagged_aggs += 1

        return Response({
            'groups_detected': len(flagged_groups),
            'flags_created': created_flags,
            'aggregates_flagged': flagged_aggs,
        })

