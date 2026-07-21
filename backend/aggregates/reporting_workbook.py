"""SESIGO Reporting Workbook generator + importer.

This module powers a familiar, donor-style reporting workbook (similar to the
NAHPA reporting templates) so that data submitters never touch technical
columns such as ``indicator_id`` / ``project_id`` / ``organization_id``.

The design goal is a *reliable* round-trip. Rather than fuzzy-matching sheet
names and indicator titles (the legacy template filler in ``views.py``), every
generated workbook carries two hidden machine sheets:

  * ``Metadata``  — project / organization / quarter / version, etc.
  * ``_cellmap``  — for every numeric input cell, the exact JSON path
                    (``primary`` / ``secondary`` / ``band``) it maps to.

Because of ``_cellmap`` the importer reconstructs the canonical aggregate
``value`` payload deterministically, regardless of how many disaggregation
dimensions an indicator has. The visible ``Reporting Form`` sheet mirrors the
on-screen ``AggregateMatrixTable``: its columns are derived from each
indicator's ``aggregate_disaggregation_config`` (the "indicator matrix").

The ``value`` shape produced here matches what the capture UI submits via
``buildEntryMatrixPayload``::

    {"disaggregates": {primary: {secondary: {band: number}}}, "total": number}

so existing dashboards, the review queue and analytics read it unchanged.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.protection import WorkbookProtection

# Bump the minor version for backwards-compatible layout tweaks; bump the major
# version (and gate the importer) for breaking changes to the cell map schema.
WORKBOOK_VERSION = "sesigo-reporting-workbook/1"
# Highest cell-map schema major version this importer understands. A workbook
# stamped with a HIGHER major (a future breaking format) is rejected on import
# rather than silently mis-parsed. Current workbooks are major 1 → always accepted.
SUPPORTED_WORKBOOK_MAJOR = int(WORKBOOK_VERSION.rsplit("/", 1)[-1].split(".")[0])


def _workbook_major(version: str):
    """Extract the integer major version from a workbook version string, or None."""
    tail = str(version or "").rsplit("/", 1)[-1]
    try:
        return int(str(tail).split(".")[0])
    except (TypeError, ValueError):
        return None

SHEET_INSTRUCTIONS = "Instructions"
SHEET_FORM = "Reporting Form"
SHEET_META = "Metadata"
SHEET_CELLMAP = "_cellmap"

MATRIX_SEXES = ["Male", "Female"]
NO_BAND = "Value"
ALL_PRIMARY = "All"

# Workbook protection password — deters casual structural edits / un-hiding the
# machine sheets. The importer reads protected workbooks without it.
PROTECTION_PASSWORD = "sesigo"

# Reporting Form column layout — mirrors the NAHPA consolidated reporting form:
# indicator | key-population | sex | age-band columns… | Sub-total | TOTAL | AYP
COL_NAME = 1        # A  (indicator name, merged down the block)
COL_KP = 2          # B  (primary disaggregate: key population / message / role)
COL_SEX = 3         # C  (secondary disaggregate: sex)
COL_BAND_START = 4  # D  (age-band columns, or a single value column)

AYP_LABEL = "AYP (10-24)"

# ── Styling ──────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
_HEADCELL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_LABEL_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
_INPUT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
_SUBTOTAL_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
_GRANDTOTAL_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
_WHITE_BOLD = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Locked vs unlocked: by default Excel treats every cell as locked, but the lock
# only bites once sheet protection is on. We unlock the numeric input cells so
# users can type there; everything else (labels, headers, auto-sum formulas)
# stays locked.
_LOCKED = Protection(locked=True)
_UNLOCKED = Protection(locked=False)


# ── Quarter / period helpers ─────────────────────────────────────────────────
def quarter_period_range(quarter: int, fiscal_start_year: int) -> tuple[date, date]:
    """Return (period_start, period_end) for a Botswana fiscal quarter.

    Fiscal year runs Apr 1 → Mar 31. Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec,
    Q4=Jan-Mar (of the following calendar year). ``fiscal_start_year`` is the
    calendar year the fiscal year starts in (e.g. 2025 for FY 2025/26).
    """
    q = int(quarter)
    y = int(fiscal_start_year)
    if q == 1:
        return date(y, 4, 1), date(y, 6, 30)
    if q == 2:
        return date(y, 7, 1), date(y, 9, 30)
    if q == 3:
        return date(y, 10, 1), date(y, 12, 31)
    if q == 4:
        return date(y + 1, 1, 1), date(y + 1, 3, 31)
    raise ValueError(f"Quarter must be 1-4, received {quarter!r}")


def fiscal_year_label(fiscal_start_year: int) -> str:
    y = int(fiscal_start_year)
    return f"{y}/{str(y + 1)[-2:]}"


def quarter_label(quarter: int, fiscal_start_year: int) -> str:
    return f"Q{int(quarter)} {fiscal_year_label(fiscal_start_year)}"


# ── Quarter-completion rule (reporting eligibility) ──────────────────────────
# An organisation may only report on a period AFTER that period has fully
# elapsed — a period must represent what is being reported on, not when data is
# entered. So Q1 (Apr–Jun) may only open on 1 Jul, Q2 (Jul–Sep) on 1 Oct, etc.
# These helpers are the single source of truth for that rule; the aggregate
# write path (all HTTP submit endpoints + the overwrite CLI) enforces it.

_QUARTER_START_MONTHS = {4: 1, 7: 2, 10: 3, 1: 4}


def _as_date(value):
    """Coerce a ``date`` / ``datetime`` / ISO string (as read from a serializer
    or an openpyxl cell) to a plain ``date``, or ``None`` if it cannot be."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except (TypeError, ValueError):
        return None


def quarter_of_period(period_start: date, period_end: date):
    """Reverse-map an exact fiscal-quarter ``[start, end]`` window to
    ``(quarter, fiscal_start_year)``, or ``None`` when the dates are not a
    canonical quarter (e.g. a monthly or yearly period)."""
    period_start = _as_date(period_start)
    period_end = _as_date(period_end)
    if period_start is None or period_end is None:
        return None
    q = _QUARTER_START_MONTHS.get(period_start.month)
    if q is None:
        return None
    fy = period_start.year if q != 4 else period_start.year - 1
    if quarter_period_range(q, fy) == (period_start, period_end):
        return q, fy
    return None


def earliest_reporting_open_date(period_end: date) -> date:
    """The first day reporting for a period may open — the day AFTER the period
    fully elapses (Q1 ends 30 Jun → opens 1 Jul)."""
    return _as_date(period_end) + timedelta(days=1)


def period_has_fully_elapsed(period_end: date, today: date | None = None) -> bool:
    """True once the reporting period has ended (``today`` is strictly after
    ``period_end``), i.e. reporting is eligible to open."""
    period_end = _as_date(period_end)
    if period_end is None:
        return True
    today = _as_date(today) or date.today()
    return today >= earliest_reporting_open_date(period_end)


def _format_open_date(open_date: date) -> str:
    # "1 October 2025" — no zero-padded day, and %-d is not portable.
    return f"{open_date.day} {open_date:%B %Y}"


def reporting_not_yet_eligible_message(period_start: date, period_end: date) -> str:
    """The exact rejection message for an early-reporting attempt."""
    period_start = _as_date(period_start)
    period_end = _as_date(period_end)
    open_str = _format_open_date(earliest_reporting_open_date(period_end))
    info = quarter_of_period(period_start, period_end)
    if info is not None:
        q, _fy = info
        return (
            f"Q{q} reporting cannot be opened yet because the reporting quarter "
            f"has not ended. Q{q} may only open from {open_str}."
        )
    return (
        "Reporting for this period cannot be opened yet because the period has "
        f"not ended. It may only open from {open_str}."
    )


# ── Yearly / monthly periods (period_type selector) ─────────────────────────
import calendar as _calendar  # noqa: E402

_MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]


def year_period_range(fiscal_start_year: int) -> tuple[date, date]:
    """Full Botswana fiscal year: Apr 1 → Mar 31 of the next calendar year."""
    y = int(fiscal_start_year)
    return date(y, 4, 1), date(y + 1, 3, 31)


def year_label(fiscal_start_year: int) -> str:
    return f"FY {fiscal_year_label(fiscal_start_year)}"


def month_period_range(month: int, calendar_year: int) -> tuple[date, date]:
    """First and last day of a calendar month."""
    m, y = int(month), int(calendar_year)
    if not 1 <= m <= 12:
        raise ValueError(f"Month must be 1-12, received {month!r}")
    return date(y, m, 1), date(y, m, _calendar.monthrange(y, m)[1])


def month_label(month: int, calendar_year: int) -> str:
    return f"{_MONTH_NAMES[int(month)]} {int(calendar_year)}"


def fiscal_year_of_month(month: int, calendar_year: int) -> int:
    """The fiscal start year that contains a given calendar month (FY = Apr-Mar)."""
    m, y = int(month), int(calendar_year)
    return y if m >= 4 else y - 1


def resolve_period(period_type: str, *, quarter=None, month=None, fiscal_start_year=None,
                   calendar_year=None) -> tuple[date, date, str]:
    """Return (period_start, period_end, label) for the requested period type."""
    pt = (period_type or "quarter").strip().lower()
    if pt in ("year", "yearly", "annual"):
        s, e = year_period_range(fiscal_start_year)
        return s, e, year_label(fiscal_start_year)
    if pt in ("month", "monthly"):
        cy = int(calendar_year if calendar_year is not None else fiscal_start_year)
        s, e = month_period_range(month, cy)
        return s, e, month_label(month, cy)
    s, e = quarter_period_range(quarter, fiscal_start_year)
    return s, e, quarter_label(quarter, fiscal_start_year)


_QUARTER_LABEL_RE = re.compile(
    r"^Q([1-4])\s+(\d{4})(?:\s*/\s*(?:\d{2}|\d{4}))?$", re.IGNORECASE
)


def parse_quarter_label(label: str) -> tuple[int, int] | None:
    """Parse ``"Q3 2025/26"`` (or ``"Q3 2025"``) → (quarter, fiscal_start_year)."""
    m = _QUARTER_LABEL_RE.match(str(label or "").strip())
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def period_to_quarter(period_start: date, period_end: date) -> tuple[int, int] | None:
    """Inverse of :func:`quarter_period_range`, used when only dates are known."""
    for q in (1, 2, 3, 4):
        for fy in (period_start.year, period_start.year - 1):
            start, end = quarter_period_range(q, fy)
            if start == period_start and end == period_end:
                return q, fy
    return None


# ── Indicator matrix resolution (ports getAggregateEntryMatrixConfig) ─────────
def _norm(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def normalize_band_label(label) -> str:
    """Canonicalise an age-band label's *whitespace* without changing its range.

    Partner workbooks sometimes store the same age band under inconsistent
    spellings — ``"0 - 4"``/``"0-4"``, ``"5- 10"``/``"5-10"``, ``"30 -39"``/
    ``"30-39"`` — which makes one band render as two near-identical columns in the
    reporting workbook. This collapses the spaces *around the hyphen* and any
    repeated internal whitespace so those spellings map to a single band. It does
    NOT alter the numbers in a band or the band boundaries (e.g. ``"65+"`` stays
    ``"65+"``, ``"18-24"`` stays ``"18-24"``), so intentionally different age
    ranges (e.g. an indicator using social-media age ranges) are untouched.
    """
    text = str(label or "").strip()
    text = re.sub(r"\s*-\s*", "-", text)  # "0 - 4" / "5- 10" / "30 -39" -> "0-4" / "5-10" / "30-39"
    text = re.sub(r"\s+", " ", text)       # collapse any other doubled spaces
    return text


def _dim_tokens(dim: dict) -> set[str]:
    # Token-based so "messages" never matches "age", etc.
    return set(_norm(dim.get("key")).split()) | set(_norm(dim.get("label")).split())


def _is_age_dimension(dim: dict) -> bool:
    return "age" in _dim_tokens(dim)


def _is_sex_dimension(dim: dict) -> bool:
    tokens = _dim_tokens(dim)
    return "sex" in tokens or "gender" in tokens


@dataclass
class MatrixConfig:
    has_disaggregates: bool
    primary_label: str
    primary_values: list[str]
    secondary_label: str
    secondary_values: list[str]
    band_label: str
    band_values: list[str]


def resolve_matrix_config(indicator) -> MatrixConfig:
    """Resolve an indicator's disaggregation matrix exactly like the capture UI.

    DISAGGREGATION CONTENT IS 100% UI-CONFIGURED: the dimensions present, their
    values, their labels and their enabled state come solely from the indicator's
    ``aggregate_disaggregation_config`` (the single source of truth). Nothing here
    invents dimensions or values.

    LAYOUT IS A FIXED REPORTING CONVENTION (intentionally NOT per-indicator
    configurable, by maintainer decision 2026-06-25): to keep one stable NAHPA
    workbook shape — and to keep the stored ``(primary, secondary, band)`` key
    mapping consistent with the ~16k existing aggregates — the axes are assigned by
    role, not by config order:

      * an "age" dimension is the band (column) axis;
      * a "sex" dimension is preferred as the secondary (row) axis;
      * any other dimension(s) fill primary then secondary.

    The AYP (10-24) column and TOTAL MALE/FEMALE rows are part of this fixed format.
    Mirrors the frontend ``getAggregateEntryMatrixConfig`` so the preview, capture
    grid, workbook and review queue are identical. Changing the axis mapping would
    require migrating existing aggregate keys and is out of scope here.
    """
    # Percentage indicators RESOLVE AS A VALUE, not a single rate: their config
    # holds the disaggregated COUNT structure (the numerator) and the % is computed
    # downstream from this indicator's achieved ÷ the denominator indicator's
    # achieved. So percentages use their aggregate_disaggregation_config like any
    # count indicator (no single-value special-case).
    config = getattr(indicator, "aggregate_disaggregation_config", None) or {}
    dimensions = config.get("dimensions") if isinstance(config, dict) else None
    enabled = bool(config.get("enabled", True)) if isinstance(config, dict) else False
    dimensions = [d for d in (dimensions or []) if isinstance(d, dict)]

    if not enabled or not dimensions:
        return MatrixConfig(False, "Indicator", [], "Category", [], "Value", [])

    age_dim = next((d for d in dimensions if _is_age_dimension(d)), None)
    sex_dim = next((d for d in dimensions if _is_sex_dimension(d)), None)
    non_special = [d for d in dimensions if not _is_age_dimension(d) and not _is_sex_dimension(d)]

    primary_dim = non_special[0] if non_special else None
    secondary_dim = sex_dim or (non_special[1] if len(non_special) > 1 else None)

    def _values(dim, fallback):
        vals = [str(v) for v in (dim.get("values") or [])] if dim else []
        return vals or fallback

    # Values come straight from the indicator's config — never a hardcoded list
    # (no Male/Female default). A configured dimension always has values.
    primary_values = _values(primary_dim, [ALL_PRIMARY])
    secondary_values = _values(secondary_dim, [ALL_PRIMARY]) if secondary_dim else [ALL_PRIMARY]
    # Normalise band-label whitespace and drop the duplicates it produces so a
    # band stored as both "0 - 4" and "0-4" renders as one column (not two).
    band_values = list(dict.fromkeys(normalize_band_label(b) for b in _values(age_dim, [NO_BAND])))

    return MatrixConfig(
        has_disaggregates=True,
        primary_label=(primary_dim or {}).get("label") or "Category",
        primary_values=primary_values,
        secondary_label=(secondary_dim or {}).get("label") or "Category",
        secondary_values=secondary_values,
        band_label=(age_dim or {}).get("label") or "Value",
        band_values=band_values,
    )


def iter_cells(cfg: MatrixConfig):
    """Yield every (primary, secondary, band) input-cell coordinate."""
    for primary in cfg.primary_values:
        for secondary in cfg.secondary_values:
            for band in cfg.band_values:
                yield primary, secondary, band


# ── value <-> cells ──────────────────────────────────────────────────────────
def _to_number(raw):
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return raw
    text = str(raw).strip().replace(",", "")
    if not text:
        return None
    try:
        num = float(text)
    except ValueError:
        return None
    return int(num) if num.is_integer() else num


def build_value_from_cells(cells: dict[tuple[str, str, str], float], cfg: MatrixConfig):
    """Reconstruct the canonical aggregate ``value`` from filled-in cells.

    ``cells`` maps (primary, secondary, band) -> number. Returns the nested
    ``{"disaggregates": ..., "total": ...}`` payload, or ``None`` when nothing
    was entered for the indicator.
    """
    numeric = {k: _to_number(v) for k, v in cells.items()}
    if not any(v is not None for v in numeric.values()):
        return None

    if not cfg.has_disaggregates:
        # Single value indicators are stored as a plain total.
        total = next((v for v in numeric.values() if v is not None), 0)
        return {"total": total}

    disaggregates: dict = {}
    total = 0
    for (primary, secondary, band), value in numeric.items():
        amount = value or 0
        disaggregates.setdefault(primary, {}).setdefault(secondary, {})[band] = amount
        total += amount
    return {"disaggregates": disaggregates, "total": total}


def extract_cells_from_value(value, cfg: MatrixConfig) -> dict[tuple[str, str, str], float]:
    """Best-effort inverse of :func:`build_value_from_cells` for pre-filling.

    Reads an existing aggregate ``value`` into the (primary, secondary, band)
    grid so the "Download Existing Data Workbook" round-trips.
    """
    cells: dict[tuple[str, str, str], float] = {}
    if value is None:
        return cells

    if not cfg.has_disaggregates:
        total = None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            total = value
        elif isinstance(value, dict):
            total = value.get("total")
        if total is not None:
            cells[(ALL_PRIMARY, ALL_PRIMARY, NO_BAND)] = total
        return cells

    disaggregates = value.get("disaggregates") if isinstance(value, dict) else None
    if isinstance(disaggregates, dict):
        for primary, secondary_map in disaggregates.items():
            if not isinstance(secondary_map, dict):
                continue
            for secondary, band_map in secondary_map.items():
                if isinstance(band_map, dict):
                    for band, amount in band_map.items():
                        # Normalise band whitespace and merge duplicates (e.g. a
                        # value stored under "0 - 4" lands in the same "0-4" cell
                        # as one stored under "0-4"), summing to preserve totals.
                        key = (str(primary), str(secondary), normalize_band_label(band))
                        existing = cells.get(key)
                        if isinstance(existing, (int, float)) and isinstance(amount, (int, float)):
                            cells[key] = existing + amount
                        else:
                            cells[key] = amount
                else:
                    cells[(str(primary), str(secondary), NO_BAND)] = band_map

    # Legacy {"male": n, "female": n} payloads for sex-only indicators.
    if not disaggregates and isinstance(value, dict):
        for sex in MATRIX_SEXES:
            amount = value.get(sex.lower())
            if amount is not None:
                cells[(ALL_PRIMARY, sex, NO_BAND)] = amount
    return cells


# ── Workbook generation ──────────────────────────────────────────────────────
@dataclass
class IndicatorPlan:
    indicator: object
    config: MatrixConfig
    target: float | None
    existing_cells: dict[tuple[str, str, str], float] = field(default_factory=dict)
    # Optional section heading this indicator sits under (from a WorkbookLayout).
    # Empty string means "no section" — the workbook renders no heading row.
    section: str = ""


def generate_workbook(
    *,
    project,
    organization,
    quarter: int,
    fiscal_start_year: int,
    indicator_plans: list[IndicatorPlan],
    generated_by: str = "",
    with_data: bool = False,
    period_start: date | None = None,
    period_end: date | None = None,
    period_label: str | None = None,
) -> BytesIO:
    """Build the multi-sheet reporting workbook and return it as a BytesIO.

    By default the period is the given fiscal ``quarter``; pass
    ``period_start``/``period_end``/``period_label`` to render a yearly or
    monthly workbook instead (see resolve_period)."""
    if period_start is None or period_end is None:
        period_start, period_end = quarter_period_range(quarter, fiscal_start_year)
    period_title = period_label or quarter_label(quarter, fiscal_start_year)

    wb = Workbook()
    form = wb.active
    form.title = SHEET_FORM
    meta = wb.create_sheet(SHEET_META)
    cellmap = wb.create_sheet(SHEET_CELLMAP)
    instructions = wb.create_sheet(SHEET_INSTRUCTIONS)

    cellmap_rows: list[list] = [
        ["indicator_id", "indicator_code", "kind", "primary", "secondary", "band", "coordinate"]
    ]

    # ── Reporting Form ──
    form.sheet_view.showGridLines = False
    form.column_dimensions[get_column_letter(COL_NAME)].width = 34
    form.column_dimensions[get_column_letter(COL_KP)].width = 16
    form.column_dimensions[get_column_letter(COL_SEX)].width = 10

    # Numbers-only validation shared by every input cell (whole numbers >= 0).
    number_validation = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True,
    )
    number_validation.showErrorMessage = True
    number_validation.errorTitle = "Numbers only"
    number_validation.error = "Please enter a whole number (0 or more)."
    number_validation.promptTitle = "Enter a number"
    number_validation.prompt = "Type a whole number of people/events for this cell."
    form.add_data_validation(number_validation)
    # Collect input-cell coordinates here; bind them to the validation in one
    # shot after all blocks are written (see _input_cell).
    number_validation._pending_coords = []

    title = f"{project.name} — {organization.name}"
    form.cell(row=1, column=COL_NAME, value=title).font = Font(bold=True, size=14)
    form.cell(row=2, column=COL_NAME, value=f"{period_title}  ({period_start} to {period_end})").font = _BOLD
    form.cell(
        row=3, column=COL_NAME,
        value="Enter whole numbers in the light-blue cells only. Totals fill in automatically. The sheet is protected; do not unprotect, rename or delete sheets.",
    ).font = Font(italic=True, color="808080")

    # Widest band-column count across all indicators → aligns every TOTAL column.
    band_grid = max(
        (len(p.config.band_values) for p in indicator_plans
         if getattr(p.config, "has_disaggregates", False)),
        default=0,
    )

    row = 5
    current_section = None
    for plan in indicator_plans:
        section = (getattr(plan, "section", "") or "").strip()
        if section and section != current_section:
            row = _write_section_heading(form, row, section)
            current_section = section
        row = _write_indicator_block(
            form, row, plan, cellmap_rows, number_validation, with_data=with_data,
            band_grid=band_grid,
        )
        row += 1  # spacer

    # Bind every collected input cell to the numeric validation in a single
    # assignment (one MultiCellRange build instead of thousands).
    if number_validation._pending_coords:
        number_validation.sqref = " ".join(number_validation._pending_coords)
    del number_validation._pending_coords

    # Lock the sheet: only the unlocked light-blue input cells are editable;
    # labels, headers and auto-sum formulas stay locked.
    form.protection.sheet = True
    form.protection.password = PROTECTION_PASSWORD
    form.protection.formatColumns = False
    form.protection.formatRows = False

    # ── Metadata (hidden) ──
    meta_pairs = [
        ("workbook_version", WORKBOOK_VERSION),
        ("project_id", project.id),
        ("project_code", getattr(project, "code", "") or ""),
        ("project_name", project.name),
        ("organization_id", organization.id),
        ("organization_code", getattr(organization, "code", "") or ""),
        ("organization_name", organization.name),
        ("quarter", quarter),
        ("fiscal_start_year", fiscal_start_year),
        ("fiscal_year", fiscal_year_label(fiscal_start_year)),
        ("quarter_label", period_title),
        ("period_label", period_title),
        ("period_start", period_start.isoformat()),
        ("period_end", period_end.isoformat()),
        ("is_training", "true" if getattr(project, "is_training", False) else "false"),
        ("generated_at", datetime.utcnow().isoformat() + "Z"),
        ("generated_by", generated_by or ""),
        ("indicator_count", len(indicator_plans)),
    ]
    meta.cell(row=1, column=1, value="key").font = _BOLD
    meta.cell(row=1, column=2, value="value").font = _BOLD
    for i, (key, val) in enumerate(meta_pairs, start=2):
        meta.cell(row=i, column=1, value=key)
        meta.cell(row=i, column=2, value=val)
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 48
    meta.sheet_state = "hidden"

    # ── _cellmap (hidden) ──
    for r, values in enumerate(cellmap_rows, start=1):
        for c, val in enumerate(values, start=1):
            cellmap.cell(row=r, column=c, value=val)
    cellmap.sheet_state = "veryHidden"

    # ── Instructions ──
    _write_instructions(instructions, project, organization, quarter, fiscal_start_year)

    # Keep the form active/selected on open.
    wb.active = wb.sheetnames.index(SHEET_FORM)

    # Lock workbook structure so the hidden Metadata/_cellmap sheets cannot be
    # unhidden, reordered or deleted in Excel (the importer reads them anyway).
    wb.security = WorkbookProtection(workbookPassword=PROTECTION_PASSWORD, lockStructure=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


SHEET_TOTAL = "TOTAL (Coordinator)"


def _safe_sheet_title(name: str, used: set) -> str:
    """Excel-safe, unique sheet title (<=31 chars, no : \\ / ? * [ ])."""
    title = re.sub(r"[:\\/?*\[\]]", " ", str(name or "Sheet")).strip()[:31] or "Sheet"
    base, n = title, 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def _write_form_sheet(wb, title, *, org_name, project, quarter, fiscal_start_year,
                      plans, cellmap_rows, with_data=False, provider_factory=None, index=None,
                      period_start=None, period_end=None, period_label=None):
    """Write one reporting-form sheet (the layout used by generate_workbook) into
    an existing workbook, and return the sheet. ``provider_factory(indicator)``
    optionally yields a formula_provider so the sheet's input cells become
    cross-sheet rollup formulas (used for the coordinator TOTAL sheet)."""
    ws = wb.create_sheet(title=title) if index is None else wb.create_sheet(title=title, index=index)
    if period_start is None or period_end is None:
        period_start, period_end = quarter_period_range(quarter, fiscal_start_year)
    period_title = period_label or quarter_label(quarter, fiscal_start_year)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions[get_column_letter(COL_NAME)].width = 34
    ws.column_dimensions[get_column_letter(COL_KP)].width = 16
    ws.column_dimensions[get_column_letter(COL_SEX)].width = 10

    dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv.showErrorMessage = True
    dv.errorTitle = "Numbers only"
    dv.error = "Please enter a whole number (0 or more)."
    ws.add_data_validation(dv)
    dv._pending_coords = []

    ws.cell(row=1, column=COL_NAME, value=f"{project.name} — {org_name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=COL_NAME,
            value=f"{period_title}  ({period_start} to {period_end})").font = _BOLD
    note = ("Coordinator rollup — totals are summed automatically from the sub sheets; do not edit."
            if provider_factory is not None else
            "Enter whole numbers in the light-blue cells only. Totals fill in automatically.")
    ws.cell(row=3, column=COL_NAME, value=note).font = Font(italic=True, color="808080")

    band_grid = max(
        (len(p.config.band_values) for p in plans
         if getattr(p.config, "has_disaggregates", False)),
        default=0,
    )

    row = 5
    current_section = None
    for plan in plans:
        section = (getattr(plan, "section", "") or "").strip()
        if section and section != current_section:
            row = _write_section_heading(ws, row, section)
            current_section = section
        fp = provider_factory(plan.indicator) if provider_factory else None
        # band_grid is intentionally NOT used for fixed Sub-total/TOTAL/AYP
        # placement: pinning them to the widest indicator's column left empty
        # gap columns between the age bands and the TOTAL for narrower / no-age
        # indicators (the "split"). Each indicator block carries its own header,
        # so placing Sub-total/TOTAL/AYP immediately after its own age bands
        # reads cleanly with no gap.
        row = _write_indicator_block(ws, row, plan, cellmap_rows, dv, with_data=with_data,
                                     formula_provider=fp, band_grid=None)
        row += 1

    if dv._pending_coords:
        dv.sqref = " ".join(dv._pending_coords)
    del dv._pending_coords

    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    return ws


# ── Formula cached-value evaluation ──────────────────────────────────────────
# openpyxl writes formula cells as ``<f>…</f>`` with NO cached ``<v>`` (or an
# empty ``<v/>``). Excel tolerates that for same-sheet formulas, but for the
# coordinator/consolidated TOTAL sheets' *cross-sheet* ``=SUM('Sub'!…)`` rollups
# it reports "we found a problem with some content" and silently strips them
# ("Repaired"). It also means every formula cell reads as blank until Excel
# recalculates — so in Protected View (which does NOT recalc), non-Excel viewers
# (Google Sheets, LibreOffice, macOS Preview) and in-browser previews, the whole
# TOTAL sheet shows zeros even though the underlying data is correct.
#
# We fix both by (1) setting ``fullCalcOnLoad`` on the workbook so Excel force-
# recomputes on open, and (2) writing the REAL computed total as each formula
# cell's cached ``<v>`` so the numbers are already correct before any recalc.


def _split_top_level_commas(text: str) -> list[str]:
    """Split ``SUM(...)`` arguments on top-level commas, respecting the single
    quotes that wrap sheet names (which may themselves contain commas)."""
    args, buf, in_quote = [], [], False
    for ch in text:
        if ch == "'":
            in_quote = not in_quote
            buf.append(ch)
        elif ch == "," and not in_quote:
            args.append("".join(buf))
            buf = []
        else:
            buf.append(ch)
    if buf:
        args.append("".join(buf))
    return args


def _parse_cell_ref(current_title: str, ref: str) -> tuple[str, str]:
    """Resolve ``'Sheet Name'!A1`` / ``Sheet!A1`` / ``A1`` to ``(title, coord)``."""
    title, coord = current_title, ref
    if "!" in ref:
        sheet, coord = ref.rsplit("!", 1)
        sheet = sheet.strip()
        if sheet.startswith("'") and sheet.endswith("'"):
            sheet = sheet[1:-1].replace("''", "'")
        title = sheet
    return title, coord.strip()


def _evaluate_workbook_values(wb) -> dict[tuple[str, str], float]:
    """Compute the numeric result of every formula cell in ``wb``.

    The reporting workbooks only ever emit SUM-of-cells formulas: ``=SUM(range)``,
    ``=SUM('Sheet'!A1,'Sheet'!A2,…)`` and ``=A1+B1+…`` (plus bare ``0``). That is a
    pure DAG of additions over literal input cells, so a small fixed-point pass
    resolves it exactly the way Excel would — no external formula engine needed.
    Returns ``{(sheet_title, coord): value}`` for formula cells only.
    """
    from openpyxl.utils import range_boundaries, get_column_letter

    literals: dict[tuple[str, str], float] = {}
    formulas: dict[tuple[str, str], str] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if isinstance(value, str) and value.startswith("="):
                    formulas[(ws.title, cell.coordinate)] = value[1:]
                elif isinstance(value, (int, float)) and not isinstance(value, bool):
                    literals[(ws.title, cell.coordinate)] = float(value)

    computed = dict(literals)

    def _lookup(title: str, coord: str):
        """(known, value): a blank cell counts as 0; a not-yet-evaluated formula
        cell is 'unknown' so its dependants are deferred to a later pass."""
        key = (title, coord)
        if key in computed:
            return True, computed[key]
        if key in formulas:
            return False, None
        return True, 0.0

    def _eval(title: str, formula: str):
        expr = formula.strip()
        upper = expr.upper()
        if upper.startswith("SUM(") and expr.endswith(")"):
            total = 0.0
            for arg in _split_top_level_commas(expr[4:-1]):
                arg = arg.strip()
                if not arg:
                    continue
                if ":" in arg:
                    rng_title, rng = title, arg
                    if "!" in arg:
                        sheet, rng = arg.rsplit("!", 1)
                        sheet = sheet.strip()
                        if sheet.startswith("'") and sheet.endswith("'"):
                            sheet = sheet[1:-1].replace("''", "'")
                        rng_title = sheet
                    min_col, min_row, max_col, max_row = range_boundaries(rng)
                    for rr in range(min_row, max_row + 1):
                        for cc in range(min_col, max_col + 1):
                            known, val = _lookup(rng_title, f"{get_column_letter(cc)}{rr}")
                            if not known:
                                return None
                            total += val or 0
                else:
                    ref_title, ref_coord = _parse_cell_ref(title, arg)
                    known, val = _lookup(ref_title, ref_coord)
                    if not known:
                        return None
                    total += val or 0
            return total
        if "+" in expr:
            total = 0.0
            for part in expr.split("+"):
                part = part.strip()
                if not part:
                    continue
                ref_title, ref_coord = _parse_cell_ref(title, part)
                known, val = _lookup(ref_title, ref_coord)
                if not known:
                    return None
                total += val or 0
            return total
        try:
            return float(expr)
        except ValueError:
            ref_title, ref_coord = _parse_cell_ref(title, expr)
            known, val = _lookup(ref_title, ref_coord)
            return val if known else None

    pending = dict(formulas)
    # Dependency depth is tiny (input → sub-total → total); bound the passes by the
    # formula count so a malformed/circular ref can never spin forever.
    for _ in range(len(formulas) + 2):
        if not pending:
            break
        progressed = False
        for key, formula in list(pending.items()):
            result = _eval(key[0], formula)
            if result is not None:
                computed[key] = result
                del pending[key]
                progressed = True
        if not progressed:
            break
    return {key: computed[key] for key in formulas if key in computed}


# Matches one formula cell — captures the opening ``<c r="COORD" …>`` tag, the
# ``<f>…</f>`` element and any existing cached ``<v>`` so we can replace the value.
_FORMULA_CELL_RE = re.compile(
    r'(<c\b[^>]*\br="([A-Z]+[0-9]+)"[^>]*>)(<f\b[^>]*>.*?</f>)'
    r'(?:<v\b[^>]*>.*?</v>|<v\s*/>)?(</c>)',
    re.DOTALL,
)


def _inject_formula_cache(buf: BytesIO, value_map: dict | None = None) -> BytesIO:
    """Give every formula cell an explicit cached ``<v>`` value.

    ``value_map`` maps ``(sheet_title, coord) -> number`` (from
    :func:`_evaluate_workbook_values`); each formula cell is stamped with its real
    total so the workbook reads correctly even before Excel recalculates. Cells
    missing from the map fall back to ``0`` (structurally complete, so Excel still
    stops "repairing" the file). Only worksheet parts are rewritten; label/string
    cells have no ``<f>`` and are left untouched.
    """
    import zipfile
    from xml.sax.saxutils import unescape

    value_map = value_map or {}
    src = zipfile.ZipFile(buf, "r")

    # Map each worksheet part (xl/worksheets/sheetN.xml) to its sheet title via
    # workbook.xml + its rels, so per-cell values land on the right sheet.
    path_to_title: dict[str, str] = {}
    try:
        wb_xml = src.read("xl/workbook.xml").decode("utf-8")
        rels_xml = src.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rid_to_path: dict[str, str] = {}
        for rel in re.finditer(r"<Relationship\b[^>]*?/>", rels_xml):
            el = rel.group(0)
            rid = re.search(r'Id="([^"]+)"', el)
            target = re.search(r'Target="([^"]+)"', el)
            if rid and target:
                path = target.group(1).lstrip("/")
                if not path.startswith("xl/"):
                    path = "xl/" + path
                rid_to_path[rid.group(1)] = path
        for sheet in re.finditer(r"<sheet\b[^>]*?/>", wb_xml):
            el = sheet.group(0)
            name = re.search(r'name="([^"]+)"', el)
            rid = re.search(r'r:id="([^"]+)"', el)
            if name and rid and rid.group(1) in rid_to_path:
                path_to_title[rid_to_path[rid.group(1)]] = unescape(name.group(1))
    except KeyError:
        pass

    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in src.namelist():
            data = src.read(name)
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                title = path_to_title.get(name)

                def _repl(match, _title=title):
                    coord = match.group(2)
                    val = value_map.get((_title, coord)) if _title is not None else None
                    if val is None:
                        cached = "0"
                    else:
                        cached = str(int(val)) if float(val).is_integer() else repr(float(val))
                    return f"{match.group(1)}{match.group(3)}<v>{cached}</v>{match.group(4)}"

                data = _FORMULA_CELL_RE.sub(_repl, data.decode("utf-8")).encode("utf-8")
            zout.writestr(name, data)
    out.seek(0)
    return out


def generate_coordinator_workbook(*, project, coordinator, sub_specs, coordinator_plans,
                                  quarter: int, fiscal_start_year: int, generated_by: str = "",
                                  period_start=None, period_end=None, period_label=None,
                                  with_data: bool = False) -> BytesIO:
    """One workbook for a coordinator: a reporting-form sheet per sub-organisation
    plus a leading TOTAL sheet whose cells SUM the matching sub cells live.

    sub_specs: list of (organization, plans) for each sub-organisation.
    coordinator_plans: plans for the union of indicators (the TOTAL sheet layout).
    Standardised age bands are what make the cross-sheet sums line up.
    """
    from collections import defaultdict

    if period_start is None or period_end is None:
        period_start, period_end = quarter_period_range(quarter, fiscal_start_year)
    period_title = period_label or quarter_label(quarter, fiscal_start_year)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we add our own
    cellmap_rows = [["indicator_id", "indicator_code", "kind", "primary", "secondary", "band", "coordinate"]]
    used_titles: set = set()

    # (indicator_id, kind, primary, secondary, band) -> ["'Sheet'!Coord", ...]
    ref_index: dict = defaultdict(list)
    sub_titles = []
    for org, plans in sub_specs:
        title = _safe_sheet_title(org.name, used_titles)
        sub_titles.append((org, title))
        start = len(cellmap_rows)
        _write_form_sheet(wb, title, org_name=org.name, project=project, quarter=quarter,
                          fiscal_start_year=fiscal_start_year, plans=plans, cellmap_rows=cellmap_rows,
                          period_start=period_start, period_end=period_end, period_label=period_title,
                          with_data=with_data)
        for r in cellmap_rows[start:]:
            ind_id, _code, kind, primary, secondary, band, coord = r
            if kind in ("cell", "total"):
                sheet, cell = coord.split("!", 1)
                ref_index[(ind_id, kind, str(primary), str(secondary), str(band))].append(f"'{sheet}'!{cell}")

    def _make_provider(indicator_id):
        def provider(kind, primary, secondary, band):
            refs = ref_index.get((indicator_id, kind, str(primary), str(secondary), str(band)))
            return ("=SUM(" + ",".join(refs) + ")") if refs else 0
        return provider

    # TOTAL sheet goes LAST (after all the data sheets), mirroring the partner
    # template the coordinators already use.
    total_title = _safe_sheet_title(SHEET_TOTAL, used_titles)
    _write_form_sheet(wb, total_title, org_name=f"{coordinator.name} (All data sheets)",
                      project=project, quarter=quarter, fiscal_start_year=fiscal_start_year,
                      plans=coordinator_plans, cellmap_rows=cellmap_rows,
                      provider_factory=lambda ind: _make_provider(ind.id),
                      period_start=period_start, period_end=period_end, period_label=period_title)

    # Metadata + cellmap + instructions
    meta = wb.create_sheet(SHEET_META)
    meta_pairs = [
        ("workbook_version", WORKBOOK_VERSION),
        ("workbook_kind", "coordinator_rollup"),
        ("project_id", project.id), ("project_name", project.name),
        ("coordinator_id", coordinator.id), ("coordinator_name", coordinator.name),
        ("quarter", quarter), ("fiscal_start_year", fiscal_start_year),
        ("quarter_label", period_title), ("period_label", period_title),
        ("period_start", period_start.isoformat()), ("period_end", period_end.isoformat()),
        ("is_training", "true" if getattr(project, "is_training", False) else "false"),
        ("generated_at", datetime.utcnow().isoformat() + "Z"), ("generated_by", generated_by or ""),
        ("sub_count", len(sub_specs)),
        ("sub_sheets", "; ".join(t for _o, t in sub_titles)),
    ]
    meta.cell(row=1, column=1, value="key").font = _BOLD
    meta.cell(row=1, column=2, value="value").font = _BOLD
    for i, (k, v) in enumerate(meta_pairs, start=2):
        meta.cell(row=i, column=1, value=k)
        meta.cell(row=i, column=2, value=v)
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 60
    meta.sheet_state = "hidden"

    cellmap = wb.create_sheet(SHEET_CELLMAP)
    for r, values in enumerate(cellmap_rows, start=1):
        for c, val in enumerate(values, start=1):
            cellmap.cell(row=r, column=c, value=val)
    cellmap.sheet_state = "veryHidden"

    wb.active = 0  # open on the first data sheet (coordinator / first sub)
    wb.security = WorkbookProtection(workbookPassword=PROTECTION_PASSWORD, lockStructure=True)
    wb.calculation.fullCalcOnLoad = True  # force Excel to recompute the rollups on open
    # Stamp each formula cell with its real computed total BEFORE saving-out, so the
    # cross-sheet TOTAL sheet reads correctly even in Protected View / non-Excel
    # viewers (and Excel stops "repairing" the file).
    cached_values = _evaluate_workbook_values(wb)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf = _inject_formula_cache(buf, cached_values)
    return buf


def generate_bonaso_workbook(*, project, org_label, coordinator_groups, grand_plans,
                             quarter: int, fiscal_start_year: int, generated_by: str = "",
                             period_start=None, period_end=None, period_label=None,
                             with_data: bool = False) -> BytesIO:
    """Whole-programme (BONASO / NAHPA) workbook: for every coordinator, its data
    sheets AS-IS plus that coordinator's TOTAL sheet, then a final GRAND TOTAL
    sheet that consolidates the shared indicators across ALL coordinators.

    coordinator_groups: list of (coordinator, sub_specs, coordinator_plans) — each
    exactly what generate_coordinator_workbook consumes for one coordinator.
    grand_plans: plans (union of shared indicators) for the GRAND TOTAL sheet.
    Every TOTAL sums its sub cells live via cross-sheet formulas.
    """
    from collections import defaultdict

    if period_start is None or period_end is None:
        period_start, period_end = quarter_period_range(quarter, fiscal_start_year)
    period_title = period_label or quarter_label(quarter, fiscal_start_year)

    wb = Workbook()
    wb.remove(wb.active)
    cellmap_rows = [["indicator_id", "indicator_code", "kind", "primary", "secondary", "band", "coordinate"]]
    used_titles: set = set()
    global_ref: dict = defaultdict(list)  # every sub cell → GRAND TOTAL
    total_subs = 0

    def _provider_for(ref_index):
        def factory(ind):
            def provider(kind, primary, secondary, band):
                refs = ref_index.get((ind.id, kind, str(primary), str(secondary), str(band)))
                return ("=SUM(" + ",".join(refs) + ")") if refs else 0
            return provider
        return factory

    for coordinator, sub_specs, coord_plans in coordinator_groups:
        coord_ref: dict = defaultdict(list)  # this coordinator's subs → its TOTAL
        for org, plans in sub_specs:
            title = _safe_sheet_title(org.name, used_titles)
            start = len(cellmap_rows)
            _write_form_sheet(wb, title, org_name=org.name, project=project, quarter=quarter,
                              fiscal_start_year=fiscal_start_year, plans=plans, cellmap_rows=cellmap_rows,
                              period_start=period_start, period_end=period_end, period_label=period_title,
                              with_data=with_data)
            total_subs += 1
            for r in cellmap_rows[start:]:
                ind_id, _code, kind, primary, secondary, band, coord = r
                if kind in ("cell", "total"):
                    sheet, cell = coord.split("!", 1)
                    ref = f"'{sheet}'!{cell}"
                    key = (ind_id, kind, str(primary), str(secondary), str(band))
                    coord_ref[key].append(ref)
                    global_ref[key].append(ref)
        ctitle = _safe_sheet_title(f"{coordinator.name} TOTAL", used_titles)
        _write_form_sheet(wb, ctitle, org_name=f"{coordinator.name} (coordinator total)",
                          project=project, quarter=quarter, fiscal_start_year=fiscal_start_year,
                          plans=coord_plans, cellmap_rows=cellmap_rows,
                          provider_factory=_provider_for(coord_ref),
                          period_start=period_start, period_end=period_end, period_label=period_title)

    gtitle = _safe_sheet_title("GRAND TOTAL", used_titles)
    _write_form_sheet(wb, gtitle, org_name=f"{org_label} — GRAND TOTAL (all coordinators)",
                      project=project, quarter=quarter, fiscal_start_year=fiscal_start_year,
                      plans=grand_plans, cellmap_rows=cellmap_rows,
                      provider_factory=_provider_for(global_ref),
                      period_start=period_start, period_end=period_end, period_label=period_title)

    meta = wb.create_sheet(SHEET_META)
    meta_pairs = [
        ("workbook_version", WORKBOOK_VERSION), ("workbook_kind", "bonaso_rollup"),
        ("project_id", project.id), ("project_name", project.name),
        ("coordinator_count", len(coordinator_groups)), ("sub_count", total_subs),
        ("quarter", quarter), ("fiscal_start_year", fiscal_start_year),
        ("quarter_label", period_title), ("period_label", period_title),
        ("period_start", period_start.isoformat()), ("period_end", period_end.isoformat()),
        ("is_training", "true" if getattr(project, "is_training", False) else "false"),
        ("generated_at", datetime.utcnow().isoformat() + "Z"), ("generated_by", generated_by or ""),
    ]
    meta.cell(row=1, column=1, value="key").font = _BOLD
    meta.cell(row=1, column=2, value="value").font = _BOLD
    for i, (k, v) in enumerate(meta_pairs, start=2):
        meta.cell(row=i, column=1, value=k)
        meta.cell(row=i, column=2, value=v)
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 60
    meta.sheet_state = "hidden"

    cellmap = wb.create_sheet(SHEET_CELLMAP)
    for r, values in enumerate(cellmap_rows, start=1):
        for c, val in enumerate(values, start=1):
            cellmap.cell(row=r, column=c, value=val)
    cellmap.sheet_state = "veryHidden"

    wb.active = 0
    wb.security = WorkbookProtection(workbookPassword=PROTECTION_PASSWORD, lockStructure=True)
    wb.calculation.fullCalcOnLoad = True  # force Excel to recompute the rollups on open
    cached_values = _evaluate_workbook_values(wb)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf = _inject_formula_cache(buf, cached_values)
    return buf


def generate_consolidated_workbook(*, project, org_label, coordinator_specs, grand_plans,
                                   quarter: int, fiscal_start_year: int, generated_by: str = "",
                                   period_start=None, period_end=None, period_label=None) -> BytesIO:
    """Consolidated (BONASO / NAHPA) workbook: one sheet per coordinator showing
    that coordinator's summed indicator matrix, then a GRAND TOTAL sheet that
    consolidates every coordinator.

    Unlike :func:`generate_bonaso_workbook`, the per-coordinator and grand cells
    are pre-summed server-side and written as STATIC values (no per-sub sheets, no
    cross-sheet formulas), so the file stays small and generates in seconds even
    for large programmes — safe to build on the request path.

    coordinator_specs: list of (coordinator, plans) where each plan's
    ``existing_cells`` already holds that coordinator's summed values.
    grand_plans: plans whose ``existing_cells`` hold the programme-wide sums.
    """
    if period_start is None or period_end is None:
        period_start, period_end = quarter_period_range(quarter, fiscal_start_year)
    period_title = period_label or quarter_label(quarter, fiscal_start_year)

    wb = Workbook()
    wb.remove(wb.active)
    cellmap_rows = [["indicator_id", "indicator_code", "kind", "primary", "secondary", "band", "coordinate"]]
    used_titles: set = set()
    coord_titles = []

    for coordinator, plans in coordinator_specs:
        title = _safe_sheet_title(f"{coordinator.name} TOTAL", used_titles)
        coord_titles.append(title)
        _write_form_sheet(wb, title, org_name=f"{coordinator.name} (all sub-grantees)",
                          project=project, quarter=quarter, fiscal_start_year=fiscal_start_year,
                          plans=plans, cellmap_rows=cellmap_rows,
                          period_start=period_start, period_end=period_end, period_label=period_title,
                          with_data=True)

    gtitle = _safe_sheet_title("GRAND TOTAL", used_titles)
    _write_form_sheet(wb, gtitle, org_name=f"{org_label} — GRAND TOTAL (all coordinators)",
                      project=project, quarter=quarter, fiscal_start_year=fiscal_start_year,
                      plans=grand_plans, cellmap_rows=cellmap_rows,
                      period_start=period_start, period_end=period_end, period_label=period_title,
                      with_data=True)

    meta = wb.create_sheet(SHEET_META)
    meta_pairs = [
        ("workbook_version", WORKBOOK_VERSION), ("workbook_kind", "consolidated_rollup"),
        ("project_id", project.id), ("project_name", project.name),
        ("coordinator_count", len(coordinator_specs)),
        ("coordinator_sheets", "; ".join(coord_titles)),
        ("quarter", quarter), ("fiscal_start_year", fiscal_start_year),
        ("quarter_label", period_title), ("period_label", period_title),
        ("period_start", period_start.isoformat()), ("period_end", period_end.isoformat()),
        ("is_training", "true" if getattr(project, "is_training", False) else "false"),
        ("generated_at", datetime.utcnow().isoformat() + "Z"), ("generated_by", generated_by or ""),
    ]
    meta.cell(row=1, column=1, value="key").font = _BOLD
    meta.cell(row=1, column=2, value="value").font = _BOLD
    for i, (k, v) in enumerate(meta_pairs, start=2):
        meta.cell(row=i, column=1, value=k)
        meta.cell(row=i, column=2, value=v)
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 60
    meta.sheet_state = "hidden"

    cellmap = wb.create_sheet(SHEET_CELLMAP)
    for r, values in enumerate(cellmap_rows, start=1):
        for c, val in enumerate(values, start=1):
            cellmap.cell(row=r, column=c, value=val)
    cellmap.sheet_state = "veryHidden"

    wb.active = 0
    wb.security = WorkbookProtection(workbookPassword=PROTECTION_PASSWORD, lockStructure=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Static values only — no cross-sheet formulas, so no formula-cache injection needed.
    return buf


_AYP_RANGE_RE = re.compile(r"^\s*(\d{1,2})\s*-\s*(\d{1,2})\s*$")
_AYP_SINGLE_RE = re.compile(r"^\s*(\d{1,2})\s*$")


def _band_in_ayp(band: str) -> bool:
    """True when an age band falls entirely within the AYP window (10-24)."""
    m = _AYP_RANGE_RE.match(str(band))
    if m:
        return int(m.group(1)) >= 10 and int(m.group(2)) <= 24
    m = _AYP_SINGLE_RE.match(str(band))
    if m:
        return 10 <= int(m.group(1)) <= 24
    return False


def _style(cell, *, fill=None, font=None, align=_CENTER, locked=True):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    cell.alignment = align
    cell.border = _BORDER
    cell.protection = _LOCKED if locked else _UNLOCKED
    return cell


def _merge(ws, r1, c1, r2, c2, value, *, fill=None, font=None, align=_CENTER):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            _style(ws.cell(row=rr, column=cc), fill=fill, font=font, align=align)
    top_left = ws.cell(row=r1, column=c1, value=value)
    _style(top_left, fill=fill, font=font, align=align)
    return top_left


def _input_cell(ws, row, col, dv, value=None):
    cell = _style(ws.cell(row=row, column=col), fill=_INPUT_FILL, locked=False)
    if value is not None:
        cell.value = value
    # Defer validation binding: collecting coordinates and setting ``dv.sqref``
    # once (in generate_workbook) avoids openpyxl rebuilding the MultiCellRange
    # on every single ``dv.add`` — that per-cell rebuild is O(n^2) and was the
    # dominant cost when a form has thousands of input cells (6s+ -> sub-second).
    coll = getattr(dv, "_pending_coords", None)
    if coll is not None:
        coll.append(cell.coordinate)
    else:
        dv.add(cell)
    return cell


def _value_cell(ws, row, col, dv, *, formula_provider, kind, primary, secondary, band):
    """An editable blue input cell, or — for the coordinator TOTAL sheet — a
    locked cross-sheet rollup formula cell. ``formula_provider`` returns the
    Excel formula string for this (kind, primary, secondary, band), or None."""
    if formula_provider is not None:
        formula = formula_provider(kind, primary, secondary, band)
        cell = _style(ws.cell(row=row, column=col, value=formula if formula is not None else 0),
                      fill=_SUBTOTAL_FILL, locked=True)
        return cell
    return _input_cell(ws, row, col, dv)


SECTION_HEADING_SPAN = 8  # columns a section heading row spans (A..H)


def _write_section_heading(ws, start_row, title: str) -> int:
    """Write a full-width section heading row (e.g. "HIV Testing"). Returns the
    next free row. Headings carry no input cells, so they never affect the
    cellmap or the import round-trip."""
    _merge(ws, start_row, COL_NAME, start_row, SECTION_HEADING_SPAN, str(title).strip(),
           fill=_SECTION_FILL, font=_WHITE_BOLD, align=_LEFT)
    return start_row + 1


def _write_indicator_block(ws, start_row, plan: IndicatorPlan, cellmap_rows, dv, *, with_data, formula_provider=None, band_grid=None):
    """Write one indicator in the NAHPA consolidated reporting layout.

    Columns: indicator name | key population | sex | age bands… | Sub-total |
    TOTAL (per key population) | AYP (10-24). Followed by a Sub-total row and
    (for sex-disaggregated indicators) TOTAL MALE / TOTAL FEMALE rows. Every
    cell is bordered; only the light-blue age cells are unlocked + numeric.

    ``band_grid`` (the widest band-column count across all indicators on the
    sheet) aligns the Sub-total/TOTAL/AYP columns to FIXED positions so every
    indicator's TOTAL lines up in one column regardless of its own band count.
    When None, each block sizes those columns to its own bands (legacy layout).
    """
    indicator = plan.indicator
    cfg = plan.config
    # Fixed (aligned) Sub-total / TOTAL / AYP columns, shared by every block.
    aligned = band_grid is not None and band_grid > 0
    fixed_sub_col = (COL_BAND_START + band_grid) if aligned else None
    fixed_total_col = (fixed_sub_col + 1) if aligned else None
    fixed_ayp_col = (fixed_total_col + 1) if aligned else None
    name = (getattr(indicator, "name", "") or "").strip()
    try:
        category = (indicator.get_category_display() or "").upper()
    except Exception:
        category = "REPORTING"

    def _record(kind, primary, secondary, band, cell):
        cellmap_rows.append([
            indicator.id, getattr(indicator, "code", "") or "", kind,
            primary, secondary, band, f"{ws.title}!{cell.coordinate}",
        ])

    # ── Plain (no disaggregation): name | ONE merged full-width value cell. ──
    # The single value is one merged, bordered cell spanning the whole data area
    # (COL_SEX .. last data column), so plain indicators read flat and their
    # borders line up with the disaggregated grid. No "Total"/"Count" label.
    if not cfg.has_disaggregates:
        _merge(ws, start_row, COL_NAME, start_row, COL_KP, name or category,
               fill=_LABEL_FILL, font=_BOLD, align=_LEFT)
        anchor = COL_SEX
        end_col = fixed_ayp_col if aligned else (COL_BAND_START + 6)
        cell = _value_cell(ws, start_row, anchor, dv, formula_provider=formula_provider,
                           kind="total", primary=ALL_PRIMARY, secondary=ALL_PRIMARY, band=NO_BAND)
        if with_data and formula_provider is None:
            amount = plan.existing_cells.get((ALL_PRIMARY, ALL_PRIMARY, NO_BAND))
            if amount is not None:
                cell.value = amount
        if end_col > anchor:
            ws.merge_cells(start_row=start_row, start_column=anchor,
                           end_row=start_row, end_column=end_col)
            fill = _SUBTOTAL_FILL if formula_provider is not None else _INPUT_FILL
            locked = formula_provider is not None
            for cc in range(anchor, end_col + 1):
                covered = ws.cell(row=start_row, column=cc)
                covered.border = _BORDER
                covered.fill = fill
                covered.alignment = _CENTER
                covered.protection = _LOCKED if locked else _UNLOCKED
        _record("total", "", "", "", cell)
        return start_row + 1

    bands = cfg.band_values
    has_age = bands != [NO_BAND]
    primary_vals = cfg.primary_values
    sec_vals = cfg.secondary_values
    has_kp = primary_vals != [ALL_PRIMARY]
    has_sex = sec_vals != [ALL_PRIMARY]
    is_sex = has_sex and all(v in MATRIX_SEXES for v in sec_vals)

    # Primary-only count indicator (a disaggregate list but no sex and no age):
    # collapse the single value into the sex slot (col C) so we don't emit a
    # redundant "Count" in BOTH the sex column and the value column. The importer
    # is cellmap-driven, so moving the input cell here keeps the round-trip intact.
    single_count = not has_sex and not has_age

    n_bands = len(bands)
    first_band = COL_SEX if single_count else COL_BAND_START
    last_band = first_band + n_bands - 1
    if aligned:
        # Sub-total/TOTAL/AYP sit in shared fixed columns so all TOTALs align.
        col_sub = fixed_sub_col if has_age else None
        col_total = fixed_total_col
        col_ayp = fixed_ayp_col if has_age else None
    else:
        col_sub = (last_band + 1) if has_age else None
        col_total = (col_sub if col_sub else last_band) + 1
        col_ayp = (col_total + 1) if has_age else None
    last_col = col_ayp or col_total

    for c in range(first_band, last_band + 1):
        ws.column_dimensions[get_column_letter(c)].width = 7
    for c in (col_sub, col_total, col_ayp):
        if c:
            ws.column_dimensions[get_column_letter(c)].width = 10

    # ── Header bar ──
    hr = start_row
    _merge(ws, hr, COL_NAME, hr, COL_KP, category, fill=_HEADER_FILL, font=_WHITE_BOLD, align=_LEFT)
    # The secondary column header must name the actual disaggregation dimension
    # from the indicator's config (e.g. "Service Provider Type", "Key Population"),
    # NOT a hardcoded "SEX". Config is the single source of truth for labels too.
    if has_sex:
        secondary_header = cfg.secondary_label or "Category"
    elif has_age:
        secondary_header = cfg.band_label or "Age Range"
    else:
        secondary_header = "Count"
    # For a primary-only count indicator col C IS the value column (labelled
    # "Count" by the band loop below), so skip the separate secondary header here.
    if not single_count:
        _style(ws.cell(row=hr, column=COL_SEX, value=secondary_header),
               fill=_HEADCELL_FILL, font=_BOLD)
    band_cols: dict[str, int] = {}
    for i, band in enumerate(bands):
        c = first_band + i
        _style(ws.cell(row=hr, column=c, value=band if has_age else "Count"),
               fill=_HEADCELL_FILL, font=_BOLD)
        band_cols[band] = c
    if col_sub:
        _style(ws.cell(row=hr, column=col_sub, value="Sub - total"), fill=_HEADCELL_FILL, font=_BOLD)
    _style(ws.cell(row=hr, column=col_total, value="TOTAL"), fill=_HEADCELL_FILL, font=_BOLD)
    if col_ayp:
        _style(ws.cell(row=hr, column=col_ayp, value=AYP_LABEL), fill=_HEADCELL_FILL, font=_BOLD)

    # ── Data rows ──
    data_start = hr + 1
    row = data_start
    male_subs: list[str] = []
    female_subs: list[str] = []
    for primary in primary_vals:
        kp_first = row
        for secondary in sec_vals:
            if has_sex:
                _style(ws.cell(row=row, column=COL_SEX, value=secondary), fill=_LABEL_FILL, font=_BOLD)
            elif not single_count:
                _style(ws.cell(row=row, column=COL_SEX, value="Count"), fill=_LABEL_FILL, font=_BOLD)
            for band in bands:
                cell = _value_cell(ws, row, band_cols[band], dv, formula_provider=formula_provider,
                                   kind="cell", primary=primary, secondary=secondary, band=band)
                if with_data and formula_provider is None:
                    amount = plan.existing_cells.get((primary, secondary, band))
                    if amount is not None:
                        cell.value = amount
                _record("cell", primary, secondary, band, cell)
            if col_sub:
                sub_ref = f"{get_column_letter(col_sub)}{row}"
                _style(ws.cell(row=row, column=col_sub,
                               value=f"=SUM({get_column_letter(first_band)}{row}:{get_column_letter(last_band)}{row})"),
                       fill=_SUBTOTAL_FILL, font=_BOLD)
                if is_sex and secondary == "Male":
                    male_subs.append(sub_ref)
                elif is_sex and secondary == "Female":
                    female_subs.append(sub_ref)
            if col_ayp:
                ayp_refs = [f"{get_column_letter(band_cols[b])}{row}" for b in bands if _band_in_ayp(b)]
                ws_ayp = ws.cell(row=row, column=col_ayp,
                                 value=("=" + "+".join(ayp_refs)) if ayp_refs else 0)
                _style(ws_ayp, fill=_SUBTOTAL_FILL, font=_BOLD)
            row += 1
        kp_last = row - 1
        # Key-population label (col B), merged across its sex rows.
        _merge(ws, kp_first, COL_KP, kp_last, COL_KP, primary if has_kp else "All",
               fill=_LABEL_FILL, font=_BOLD, align=_CENTER)
        # TOTAL per key population (sum of its rows' sub-totals), merged.
        if col_sub:
            total_formula = f"=SUM({get_column_letter(col_sub)}{kp_first}:{get_column_letter(col_sub)}{kp_last})"
        else:
            total_formula = (
                f"=SUM({get_column_letter(first_band)}{kp_first}:{get_column_letter(last_band)}{kp_last})"
                if has_age else
                f"=SUM({get_column_letter(first_band)}{kp_first}:{get_column_letter(first_band)}{kp_last})"
            )
        _merge(ws, kp_first, col_total, kp_last, col_total, total_formula,
               fill=_SUBTOTAL_FILL, font=_BOLD)
    data_end = row - 1

    # Indicator name down the left, spanning the whole block.
    _merge(ws, data_start, COL_NAME, data_end, COL_NAME, name, fill=_HEADCELL_FILL, font=_BOLD, align=_LEFT)

    # ── Sub-total row (column sums) ──
    # Primary-only count blocks keep col C for the value sum, so the "Sub - total"
    # label merges A:B; everything else merges A:C as before.
    _sub_label_end = COL_KP if single_count else COL_SEX
    _merge(ws, row, COL_NAME, row, _sub_label_end, "Sub - total", fill=_SUBTOTAL_FILL, font=_BOLD)
    for band in bands:
        c = band_cols[band]
        _style(ws.cell(row=row, column=c,
                       value=f"=SUM({get_column_letter(c)}{data_start}:{get_column_letter(c)}{data_end})"),
               fill=_SUBTOTAL_FILL, font=_BOLD)
    if col_sub:
        _style(ws.cell(row=row, column=col_sub,
                       value=f"=SUM({get_column_letter(col_sub)}{data_start}:{get_column_letter(col_sub)}{data_end})"),
               fill=_SUBTOTAL_FILL, font=_BOLD)
    _style(ws.cell(row=row, column=col_total,
                   value=f"=SUM({get_column_letter(first_band)}{data_start}:{get_column_letter(last_band)}{data_end})"),
           fill=_SUBTOTAL_FILL, font=_BOLD)
    if col_ayp:
        _style(ws.cell(row=row, column=col_ayp,
                       value=f"=SUM({get_column_letter(col_ayp)}{data_start}:{get_column_letter(col_ayp)}{data_end})"),
               fill=_SUBTOTAL_FILL, font=_BOLD)
    row += 1

    # ── TOTAL MALE / TOTAL FEMALE rows (sex-disaggregated indicators only) ──
    if is_sex:
        for label, refs in (("TOTAL MALE", male_subs), ("TOTAL FEMALE", female_subs)):
            _merge(ws, row, COL_NAME, row, COL_SEX, label, fill=_GRANDTOTAL_FILL, font=_WHITE_BOLD)
            value = ("=" + "+".join(refs)) if refs else 0
            _merge(ws, row, first_band, row, last_col, value, fill=_GRANDTOTAL_FILL, font=_WHITE_BOLD)
            row += 1

    return row


def _write_instructions(ws, project, organization, quarter, fiscal_start_year):
    ws.column_dimensions["A"].width = 100
    lines = [
        ("How to complete this workbook", 14, True),
        ("", 11, False),
        (f"Project: {project.name} ({getattr(project, 'code', '') or '—'})", 11, True),
        (f"Organization: {organization.name}", 11, True),
        (f"Reporting period: {quarter_label(quarter, fiscal_start_year)}", 11, True),
        ("", 11, False),
        ("1. Go to the 'Reporting Form' sheet.", 11, False),
        ("2. Enter your counts in the yellow input cells only.", 11, False),
        ("3. Each indicator shows the disaggregation breakdown it requires (e.g. sex, age, key population).", 11, False),
        ("4. Leave cells blank where you have no data. Blank indicators are skipped on upload.", 11, False),
        ("5. Totals are calculated automatically — do not overwrite Total columns.", 11, False),
        ("6. Do NOT rename, reorder or delete rows, columns or sheets.", 11, False),
        ("7. The hidden 'Metadata' and '_cellmap' sheets identify this workbook — leave them untouched.", 11, False),
        ("", 11, False),
        ("Validation rules", 12, True),
        ("• Values must be whole, non-negative numbers.", 11, False),
        ("• Percentage indicators cannot exceed 100.", 11, False),
        ("• Re-download a fresh workbook if you change project, organization or quarter.", 11, False),
        ("", 11, False),
        ("Uploading", 12, True),
        ("Return to the Aggregates page and choose 'Upload Workbook'. The system reads the project,", 11, False),
        ("organization and quarter automatically from this file — you never enter ID columns.", 11, False),
    ]
    for i, (text, size, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=size)
        cell.alignment = _LEFT
    ws.sheet_view.showGridLines = False


# ── Workbook parsing / validation ────────────────────────────────────────────
class WorkbookError(Exception):
    """Raised with a list of friendly validation messages."""

    def __init__(self, messages: list[str]):
        self.messages = messages
        super().__init__("; ".join(messages))


@dataclass
class ParsedIndicator:
    indicator_id: int
    indicator_code: str
    value: dict


@dataclass
class ParsedWorkbook:
    metadata: dict
    indicators: list[ParsedIndicator]
    warnings: list[str] = field(default_factory=list)


def parse_workbook(file_obj) -> ParsedWorkbook:
    """Parse + validate an uploaded reporting workbook.

    Raises :class:`WorkbookError` with friendly, user-facing messages when the
    workbook is not a recognisable SESIGO reporting workbook.
    """
    try:
        wb = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception:
        raise WorkbookError([
            "We could not open this file as an Excel workbook.",
            "Please download a fresh workbook and try again.",
        ])

    try:
        missing_sheets = [s for s in (SHEET_META, SHEET_FORM, SHEET_CELLMAP) if s not in wb.sheetnames]
        if missing_sheets:
            raise WorkbookError([
                "This file is not a SESIGO reporting workbook.",
                "Missing: " + ", ".join(missing_sheets),
                "Please use 'Download Workbook' to get a valid template.",
            ])

        metadata = _read_metadata(wb[SHEET_META])
        problems: list[str] = []

        version = str(metadata.get("workbook_version") or "")
        if not version.startswith("sesigo-reporting-workbook/"):
            problems.append("Unrecognised workbook version (this file was not generated by SESIGO).")
        else:
            major = _workbook_major(version)
            if major is not None and major > SUPPORTED_WORKBOOK_MAJOR:
                problems.append(
                    f"This workbook is a newer format (version {major}) than this system "
                    f"supports (version {SUPPORTED_WORKBOOK_MAJOR}). Please download a fresh "
                    "workbook and re-enter your data."
                )
        for key, label in (("project_id", "Project"), ("organization_id", "Organization"), ("quarter", "Quarter")):
            if not metadata.get(key):
                problems.append(f"Missing {label} information.")
        if problems:
            raise WorkbookError(["Workbook Validation Failed", *problems,
                                 "Please download a fresh workbook and try again."])

        form = wb[SHEET_FORM]
        indicators = _read_cellmap_values(wb[SHEET_CELLMAP], form)
        return ParsedWorkbook(metadata=metadata, indicators=indicators)
    finally:
        wb.close()


def _read_metadata(ws) -> dict:
    metadata: dict = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if not row:
            continue
        key, value = (row[0], row[1] if len(row) > 1 else None)
        if key is None:
            continue
        metadata[str(key).strip()] = value
    # Coerce the numeric identifiers we rely on.
    for key in ("project_id", "organization_id", "quarter", "fiscal_start_year"):
        if key in metadata and metadata[key] is not None:
            try:
                metadata[key] = int(metadata[key])
            except (TypeError, ValueError):
                pass
    return metadata


def _read_cellmap_values(cellmap_ws, form_ws) -> list[ParsedIndicator]:
    # Snapshot the form once (read_only sheets are forward-only).
    form_values: dict[str, object] = {}
    for row in form_ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is not None:
                form_values[cell.coordinate] = cell.value

    grouped: dict[int, dict] = {}
    order: list[int] = []
    rows = list(cellmap_ws.iter_rows(min_row=2, values_only=True))
    for raw in rows:
        if not raw or raw[0] is None:
            continue
        indicator_id, code, kind, primary, secondary, band, coordinate = (list(raw) + [None] * 7)[:7]
        try:
            indicator_id = int(indicator_id)
        except (TypeError, ValueError):
            continue
        coord = str(coordinate or "")
        if "!" in coord:
            coord = coord.split("!", 1)[1]
        cell_value = form_values.get(coord)
        if indicator_id not in grouped:
            grouped[indicator_id] = {"code": str(code or ""), "total": None, "cells": {}}
            order.append(indicator_id)
        bucket = grouped[indicator_id]
        if kind == "total":
            bucket["total"] = cell_value
        else:
            bucket["cells"][(str(primary), str(secondary), str(band))] = cell_value

    parsed: list[ParsedIndicator] = []
    for indicator_id in order:
        bucket = grouped[indicator_id]
        value = _bucket_to_value(bucket)
        if value is None:
            continue  # nothing entered for this indicator → skip
        parsed.append(ParsedIndicator(indicator_id, bucket["code"], value))
    return parsed


def _bucket_to_value(bucket) -> dict | None:
    cells = bucket["cells"]
    if cells:
        numeric = {k: _to_number(v) for k, v in cells.items()}
        if not any(v is not None for v in numeric.values()):
            return None
        disaggregates: dict = {}
        total = 0
        for (primary, secondary, band), value in numeric.items():
            amount = value or 0
            disaggregates.setdefault(primary, {}).setdefault(secondary, {})[band] = amount
            total += amount
        return {"disaggregates": disaggregates, "total": total}
    total = _to_number(bucket["total"])
    if total is None:
        return None
    return {"total": total}
