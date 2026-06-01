import argparse
import json
import os
import re
import sys
from pathlib import Path

DEFAULT_BACKEND_ROOT = Path("/app") if Path("/app/manage.py").exists() else Path(r"C:\Projects\django_backend")
BACKEND_ROOT = Path(os.environ.get("BONASO_DJANGO_ROOT", str(DEFAULT_BACKEND_ROOT)))
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django  # noqa: E402

django.setup()

from django.db import transaction  # noqa: E402

from aggregates.models import Aggregate  # noqa: E402
from indicators.models import Indicator, IndicatorAlias  # noqa: E402
from indicator_import_aliases import canonical_resolution_aliases, preferred_duplicate_rank  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from organizations.models import Organization  # noqa: E402
from projects.models import Project, ProjectIndicator  # noqa: E402
from uploads.management.commands.import_reporting_workbook_overwrite import (  # noqa: E402
    AGE_COLUMNS,
    IndicatorResolver,
    build_disaggregation_config,
    build_ordered_sub_labels,
    canonical_indicator_name,
    extract_section_value,
    get_age_band_mapping,
    merge_json_values,
    merge_disaggregation_configs,
    normalize_text,
    unique_indicator_code,
)


SECTION_INDEX_PATTERN = re.compile(r"^\d+[a-z]?$", re.IGNORECASE)
DEFAULT_PROJECT_CODE = "NAHPA2025/26"
DEFAULT_PERIOD_START = "2025-10-01"
DEFAULT_PERIOD_END = "2025-12-31"
DEFAULT_SHEET = "TOTAL"
DEFAULT_REPORT_PATH = Path("reports") / "cbo-oct-dec-batch-import.json"
SNAPSHOT_COLUMNS = ["B", "C", "E", "F", *AGE_COLUMNS, "AA"]
FILENAME_ORG_OVERRIDES = [
    ("apsa", "APSA"),
    ("atn", "ATN"),
    ("bona naledi", "BONA NALEDI"),
    ("bonmeh", "BONMEH"),
    ("bosasnet", "BOSASNet"),
    ("ncongo", "NCONGO"),
    ("chobe arts", "Chobe Arts"),
    ("maata", "MAATA"),
    ("home of hope", "Home of Hope"),
    ("hpp", "HPP"),
    ("journey of hope", "JOH"),
    ("makgabaneng keitsholofetse", "Keitsholofetse"),
    ("masego mental health", "Masego Mental Health"),
    ("sssg", "SSSG"),
    ("fighters support group", "TFSG"),
    ("just hope foundation", "Just Hope Foundation"),
    ("ultimate youth", "Ultimate Youth"),
    ("vmhf", "Valour Mental Health"),
]
CODE_RESOLUTION_OVERRIDES = [
    {
        "code": "16",
        "source_title": "Number of people screened for NCDs behavioural risk factors",
        "indicator_names": ["Number of people screened for NCDs behavioural risk factors (tobacco use)"],
    },
    {
        "code": "19",
        "source_title": "Number of people screened for NCDs behavioural risk factors",
        "indicator_names": ["Number of people screened for NCDs behavioural risk factors (alcohol use)"],
    },
    {
        "code": "15f",
        "source_title": "Number of individual counselling sessions conducted for youth and vulnerable groups",
        "indicator_names": ["Number of individual counselling sessions conducted for youth and vulnerable groups (15f)"],
    },
    {
        "code": "15h",
        "source_title": "Number of individual counselling sessions conducted for youth and vulnerable groups",
        "indicator_names": ["Number of individual counselling sessions conducted for youth and vulnerable groups (15h)"],
    },
    {
        "code": "37b",
        "source_title": "Number of support group meetings held",
        "indicator_names": ["Number of support group meetings held (tobacco/alcohol)"],
    },
    {
        "code": "37d",
        "source_title": "Number of new members/continuing",
        "indicator_names": ["Number of new members/continuing (tobacco/alcohol)"],
    },
]


def build_args():
    parser = argparse.ArgumentParser(
        description="Import a folder of monthly CBO reporting workbooks using the TOTAL sheet."
    )
    parser.add_argument("--folder", required=True)
    parser.add_argument("--project-code", default=DEFAULT_PROJECT_CODE)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--period-start", default=DEFAULT_PERIOD_START)
    parser.add_argument("--period-end", default=DEFAULT_PERIOD_END)
    parser.add_argument("--category", default="ncd")
    parser.add_argument("--rollup-parent", default="")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report-path", default=str(DEFAULT_REPORT_PATH))
    return parser.parse_args()


def get_code_resolution_overrides(code: str | None, title: str) -> list[str]:
    code_token = str(code or "").strip().lower()
    title_key = canonical_indicator_name(title)
    matches = []
    for entry in CODE_RESOLUTION_OVERRIDES:
        if entry["code"] != code_token:
            continue
        if canonical_indicator_name(entry["source_title"]) != title_key:
            continue
        matches.extend(entry["indicator_names"])
    return matches


def is_importable_section(title: str) -> bool:
    normalized = normalize_text(title)
    return normalized.startswith("number") or normalized.startswith("total number")


def indicator_sort_key(indicator: Indicator, project_indicator_ids: set[int]) -> tuple:
    return (
        preferred_duplicate_rank(indicator.id, canonical_indicator_name(indicator.name)),
        0 if indicator.id in project_indicator_ids else 1,
        0 if not indicator.code.upper().startswith("AUTO_") else 1,
        indicator.name.lower(),
    )


class ExactFirstIndicatorResolver:
    def __init__(self, project: Project):
        self.delegate = IndicatorResolver(project=project)
        self.project_indicator_ids = set(
            ProjectIndicator.objects.filter(project=project).values_list("indicator_id", flat=True)
        )
        self.exact_index: dict[str, list[Indicator]] = {}
        for indicator in Indicator.objects.all():
            for candidate in [indicator.name, *canonical_resolution_aliases(indicator.name)]:
                key = canonical_indicator_name(candidate)
                self.exact_index.setdefault(key, []).append(indicator)
        for alias in IndicatorAlias.objects.select_related('indicator').filter(is_active=True):
            for candidate in [alias.name, *canonical_resolution_aliases(alias.name)]:
                key = canonical_indicator_name(candidate)
                self.exact_index.setdefault(key, []).append(alias.indicator)

    def resolve(self, title: str, section_index: str | None = None) -> Indicator | None:
        for override_name in get_code_resolution_overrides(section_index, title):
            override_key = canonical_indicator_name(override_name)
            override_matches = self.exact_index.get(override_key, [])
            if override_matches:
                return sorted(
                    override_matches,
                    key=lambda item: indicator_sort_key(item, self.project_indicator_ids),
                )[0]

        for candidate in [title, *canonical_resolution_aliases(title)]:
            key = canonical_indicator_name(candidate)
            exact_matches = self.exact_index.get(key, [])
            if exact_matches:
                return sorted(
                    exact_matches,
                    key=lambda item: indicator_sort_key(item, self.project_indicator_ids),
                )[0]
        return self.delegate.resolve(title, section_index)

    def remember(self, indicator: Indicator):
        key = canonical_indicator_name(indicator.name)
        self.exact_index.setdefault(key, []).append(indicator)
        self.delegate.remember(indicator)


def normalize_filename(value: str) -> str:
    return normalize_text(Path(value).stem.replace("excel report", "report"))


def infer_category(title: str, fallback: str) -> str:
    token = normalize_text(title)
    if "sti" in token:
        return "sti"
    if "gbv" in token:
        return "gbv"
    if any(value in token for value in ("mental health", "counselling", "counseling", "psychosocial")):
        return "mental_health"
    if any(
        value in token
        for value in (
            "ncd",
            "diabetes",
            "blood glucose",
            "blood pressure",
            "bmi",
            "waist",
            "alcohol",
            "tobacco",
            "cancer",
        )
    ):
        return "ncd"
    if any(value in token for value in ("training", "mentored")):
        return "trainings"
    return fallback


def snapshot_worksheet_row(ws, row_number: int) -> dict:
    return {
        column: ws[f"{column}{row_number}"].value
        for column in SNAPSHOT_COLUMNS
    }


def parse_sections_from_worksheet(ws) -> list[dict]:
    sections = []
    current = None

    for row_number in range(1, ws.max_row + 1):
        snapshot = snapshot_worksheet_row(ws, row_number)
        index_value = str(snapshot.get("B") or "").strip()
        title = str(snapshot.get("C") or "").strip()
        is_section_start = bool(title) and bool(SECTION_INDEX_PATTERN.fullmatch(index_value))

        if is_section_start:
            if current:
                sections.append(current)
            current = {"index": index_value, "title": title, "rows": [snapshot]}
            continue

        if current:
            current["rows"].append(snapshot)

    if current:
        sections.append(current)

    return sections


def load_sheet_worksheet(workbook_path: Path, sheet_name: str):
    workbook = load_workbook(workbook_path, data_only=True, keep_links=False)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Sheet not found: {sheet_name}")
    return workbook, workbook[sheet_name]


def iter_workbooks(folder: Path):
    for path in sorted(folder.iterdir(), key=lambda item: item.name.lower()):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in {".xlsx", ".xls"}:
            continue
        yield path


def token_overlap_score(left: str, right: str) -> int:
    return len(set(left.split()) & set(right.split()))


def resolve_organization(workbook_path: Path, organizations_by_key: dict[str, Organization]) -> tuple[Organization | None, str]:
    normalized_name = normalize_filename(workbook_path.name)

    for alias, target_name in FILENAME_ORG_OVERRIDES:
        if alias in normalized_name:
            return organizations_by_key.get(normalize_text(target_name)), f"override:{alias}"

    exact_candidates = []
    for key, organization in organizations_by_key.items():
        if key in normalized_name or normalized_name in key:
            exact_candidates.append((len(key), organization))
    if exact_candidates:
        exact_candidates.sort(key=lambda item: (-item[0], item[1].name.lower()))
        return exact_candidates[0][1], "token-match"

    overlap_candidates = []
    for key, organization in organizations_by_key.items():
        overlap = token_overlap_score(normalized_name, key)
        if overlap:
            overlap_candidates.append((overlap, len(key.split()), organization))
    if overlap_candidates:
        overlap_candidates.sort(key=lambda item: (-item[0], item[1], item[2].name.lower()))
        best_overlap, _, best_org = overlap_candidates[0]
        if best_overlap >= 2:
            return best_org, "fuzzy-match"

    return None, "unmapped"


def normalize_json_value(value):
    if isinstance(value, dict):
        return {key: normalize_json_value(value[key]) for key in sorted(value.keys())}
    if isinstance(value, list):
        return [normalize_json_value(item) for item in value]
    return value


def upsert_section_aggregate(
    *,
    project: Project,
    organization: Organization,
    resolver: ExactFirstIndicatorResolver,
    section: dict,
    age_band_by_column: dict[str, str],
    fallback_category: str,
    dry_run: bool,
    workbook_name: str,
    sheet_name: str,
    period_start: str,
    period_end: str,
) -> dict:
    index = str(section["index"] or "").strip()
    title = str(section["title"] or "").strip()
    value, disaggregations = extract_section_value(section, age_band_by_column)

    indicator = resolver.resolve(title, index)
    created_indicator = False
    desired_sub_labels = build_ordered_sub_labels(disaggregations, value)
    desired_config = build_disaggregation_config(value, desired_sub_labels)

    if indicator is None:
        if dry_run:
            return {
                "index": index,
                "title": title,
                "action": "missing_indicator",
                "value": value,
            }

        indicator = Indicator.objects.create(
            name=title,
            code=unique_indicator_code(f"{project.code}-{index}-{title}"),
            type="number",
            category=infer_category(title, fallback_category),
            unit="people",
            sub_labels=desired_sub_labels,
            aggregate_disaggregation_config=desired_config,
        )
        resolver.remember(indicator)
        created_indicator = True
    else:
        update_fields = []
        if desired_sub_labels and not list(indicator.sub_labels or []):
            indicator.sub_labels = desired_sub_labels
            update_fields.append("sub_labels")
        merged_config = merge_disaggregation_configs(
            indicator.aggregate_disaggregation_config,
            desired_config,
        )
        if merged_config != dict(indicator.aggregate_disaggregation_config or {}):
            indicator.aggregate_disaggregation_config = merged_config
            update_fields.append("aggregate_disaggregation_config")
        if update_fields and not dry_run:
            indicator.save(update_fields=update_fields)

    if dry_run:
        return {
            "index": index,
            "title": title,
            "indicator_id": indicator.id if indicator else None,
            "indicator_name": indicator.name if indicator else None,
            "action": "would_create_indicator" if created_indicator else "would_import",
            "value": value,
        }

    project.organizations.add(organization)
    indicator.organizations.add(organization)
    ProjectIndicator.objects.get_or_create(project=project, indicator=indicator)

    aggregate, created_aggregate = Aggregate.objects.get_or_create(
        indicator=indicator,
        project=project,
        organization=organization,
        period_start=period_start,
        period_end=period_end,
        defaults={"value": value},
    )

    notes = f"Imported from {workbook_name} | sheet={sheet_name} | index={index}"
    action = "created"
    if created_aggregate:
        if hasattr(aggregate, "notes"):
            aggregate.notes = notes
            aggregate.save(update_fields=["notes"])
    else:
        update_fields = []
        if aggregate.value != value:
            aggregate.value = value
            update_fields.append("value")
        if hasattr(aggregate, "notes") and getattr(aggregate, "notes", None) != notes:
            aggregate.notes = notes
            update_fields.append("notes")
        if update_fields:
            aggregate.save(update_fields=update_fields)
            action = "updated"
        else:
            action = "unchanged"

    return {
        "index": index,
        "title": title,
        "indicator_id": indicator.id,
        "indicator_name": indicator.name,
        "aggregate_id": aggregate.id,
        "action": action,
        "value": value,
        "created_indicator": created_indicator,
    }


def verify_imported_sections(imported_sections: list[dict]) -> dict:
    summary = {"exact_matches": 0, "mismatches": 0, "details": []}
    for item in imported_sections:
        aggregate = Aggregate.objects.filter(id=item["aggregate_id"]).first()
        stored = aggregate.value if aggregate else None
        expected = item["value"]
        is_exact = normalize_json_value(stored) == normalize_json_value(expected)
        summary["details"].append(
            {
                "aggregate_id": item["aggregate_id"],
                "indicator_id": item["indicator_id"],
                "indicator_name": item["indicator_name"],
                "exact": is_exact,
            }
        )
        if is_exact:
            summary["exact_matches"] += 1
        else:
            summary["mismatches"] += 1
    return summary


def normalize_optional_name(value: str) -> str:
    return normalize_text(value or "")


def main():
    args = build_args()
    folder = Path(args.folder)
    if not folder.exists():
        raise SystemExit(f"Folder not found: {folder}")

    project = Project.objects.filter(code__iexact=args.project_code).first()
    if project is None:
        raise SystemExit(f"Project not found: {args.project_code}")

    organizations_by_key = {
        normalize_text(organization.name): organization
        for organization in Organization.objects.order_by("name")
    }
    resolver = ExactFirstIndicatorResolver(project=project)
    requested_rollup_parent = normalize_optional_name(args.rollup_parent)
    parent_rollups: dict[tuple[int, int], dict] = {}

    report = {
        "folder": str(folder),
        "sheet": args.sheet,
        "project": {"id": project.id, "code": project.code, "name": project.name},
        "period_start": args.period_start,
        "period_end": args.period_end,
        "dry_run": args.dry_run,
        "rollup_parent": args.rollup_parent,
        "summary": {
            "files_seen": 0,
            "files_processed": 0,
            "files_failed": 0,
            "files_unmapped": 0,
            "sections_seen": 0,
            "sections_importable": 0,
            "sections_skipped": 0,
            "missing_indicators": 0,
            "created_indicators": 0,
            "created_aggregates": 0,
            "updated_aggregates": 0,
            "unchanged_aggregates": 0,
            "exact_matches": 0,
            "mismatches": 0,
            "parent_rollup_aggregates_created": 0,
            "parent_rollup_aggregates_updated": 0,
            "parent_rollup_aggregates_unchanged": 0,
        },
        "files": [],
        "parent_rollups": [],
    }

    for workbook_path in iter_workbooks(folder):
        report["summary"]["files_seen"] += 1
        organization, mapping_reason = resolve_organization(workbook_path, organizations_by_key)
        file_report = {
            "workbook": str(workbook_path),
            "file_name": workbook_path.name,
            "mapping_reason": mapping_reason,
            "organization": None,
            "status": "pending",
            "summary": {
                "sections_seen": 0,
                "sections_importable": 0,
                "sections_skipped": 0,
                "missing_indicators": 0,
                "created_indicators": 0,
                "created_aggregates": 0,
                "updated_aggregates": 0,
                "unchanged_aggregates": 0,
                "exact_matches": 0,
                "mismatches": 0,
            },
            "skipped_sections": [],
            "missing_indicator_sections": [],
            "imported_sections": [],
            "verification": {"details": []},
        }

        if organization is None:
            file_report["status"] = "unmapped"
            file_report["error"] = "Organization could not be inferred from filename."
            report["summary"]["files_unmapped"] += 1
            report["summary"]["files_failed"] += 1
            report["files"].append(file_report)
            continue

        file_report["organization"] = {"id": organization.id, "name": organization.name}
        workbook = None

        try:
            workbook, worksheet = load_sheet_worksheet(workbook_path, args.sheet)
            sections = parse_sections_from_worksheet(worksheet)
            file_report["summary"]["sections_seen"] = len(sections)
            report["summary"]["sections_seen"] += len(sections)
            age_band_by_column = get_age_band_mapping(organization, None)

            with transaction.atomic():
                for section in sections:
                    index = str(section["index"] or "").strip()
                    title = str(section["title"] or "").strip()
                    if not is_importable_section(title):
                        file_report["summary"]["sections_skipped"] += 1
                        report["summary"]["sections_skipped"] += 1
                        file_report["skipped_sections"].append(
                            {"index": index, "title": title, "reason": "non_indicator_header"}
                        )
                        continue

                    file_report["summary"]["sections_importable"] += 1
                    report["summary"]["sections_importable"] += 1

                    result = upsert_section_aggregate(
                        project=project,
                        organization=organization,
                        resolver=resolver,
                        section=section,
                        age_band_by_column=age_band_by_column,
                        fallback_category=args.category,
                        dry_run=args.dry_run,
                        workbook_name=workbook_path.name,
                        sheet_name=args.sheet,
                        period_start=args.period_start,
                        period_end=args.period_end,
                    )

                    if result["action"] == "missing_indicator":
                        file_report["summary"]["missing_indicators"] += 1
                        report["summary"]["missing_indicators"] += 1
                        file_report["missing_indicator_sections"].append(result)
                        continue

                    if result.get("created_indicator"):
                        file_report["summary"]["created_indicators"] += 1
                        report["summary"]["created_indicators"] += 1

                    if result["action"] == "created":
                        file_report["summary"]["created_aggregates"] += 1
                        report["summary"]["created_aggregates"] += 1
                    elif result["action"] == "updated":
                        file_report["summary"]["updated_aggregates"] += 1
                        report["summary"]["updated_aggregates"] += 1
                    elif result["action"] == "unchanged":
                        file_report["summary"]["unchanged_aggregates"] += 1
                        report["summary"]["unchanged_aggregates"] += 1

                    file_report["imported_sections"].append(result)

                    if (
                        not args.dry_run
                        and organization.parent_id
                        and (
                            not requested_rollup_parent
                            or normalize_text(organization.parent.name) == requested_rollup_parent
                        )
                    ):
                        rollup_key = (organization.parent_id, result["indicator_id"])
                        if rollup_key in parent_rollups:
                            parent_rollups[rollup_key]["value"] = merge_json_values(
                                parent_rollups[rollup_key]["value"],
                                result["value"],
                            )
                            parent_rollups[rollup_key]["source_organizations"].add(organization.name)
                        else:
                            parent_rollups[rollup_key] = {
                                "parent": organization.parent,
                                "indicator_id": result["indicator_id"],
                                "indicator_name": result["indicator_name"],
                                "value": result["value"],
                                "source_organizations": {organization.name},
                            }

                if args.dry_run:
                    transaction.set_rollback(True)

            if not args.dry_run:
                verification = verify_imported_sections(file_report["imported_sections"])
                file_report["verification"] = verification
                file_report["summary"]["exact_matches"] = verification["exact_matches"]
                file_report["summary"]["mismatches"] = verification["mismatches"]
                report["summary"]["exact_matches"] += verification["exact_matches"]
                report["summary"]["mismatches"] += verification["mismatches"]

            file_report["status"] = "ok"
            report["summary"]["files_processed"] += 1
        except Exception as exc:
            file_report["status"] = "failed"
            file_report["error"] = f"{type(exc).__name__}: {exc}"
            report["summary"]["files_failed"] += 1
        finally:
            if "workbook" in locals():
                try:
                    workbook.close()
                except Exception:
                    pass

        report["files"].append(file_report)

    if not args.dry_run:
        for rollup in parent_rollups.values():
            parent = rollup["parent"]
            indicator = Indicator.objects.get(id=rollup["indicator_id"])
            project.organizations.add(parent)
            indicator.organizations.add(parent)
            aggregate, created = Aggregate.objects.update_or_create(
                indicator=indicator,
                project=project,
                organization=parent,
                period_start=args.period_start,
                period_end=args.period_end,
                defaults={"value": rollup["value"]},
            )
            action = "created" if created else "updated"
            if hasattr(aggregate, "notes"):
                aggregate.notes = (
                    f"Rolled up from child workbook imports in {Path(args.folder).name} | "
                    f"parentscope={parent.name}"
                )
                aggregate.save(update_fields=["notes"])
            if action == "created":
                report["summary"]["parent_rollup_aggregates_created"] += 1
            else:
                report["summary"]["parent_rollup_aggregates_updated"] += 1
            report["parent_rollups"].append(
                {
                    "aggregate_id": aggregate.id,
                    "organization": {"id": parent.id, "name": parent.name},
                    "indicator_id": indicator.id,
                    "indicator_name": indicator.name,
                    "action": action,
                    "source_organizations": sorted(rollup["source_organizations"]),
                }
            )

    report_path = Path(args.report_path)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(json.dumps(report["summary"], indent=2))
    print(f"Saved report: {report_path}")


if __name__ == "__main__":
    main()
