import argparse
import json
import os
import re
import sys
from collections import Counter
from pathlib import Path


BACKEND_ROOT = Path(os.environ.get("BONASO_DJANGO_ROOT", r"C:\Projects\django_backend"))
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django  # noqa: E402

django.setup()

from openpyxl import load_workbook  # noqa: E402

from aggregates.models import Aggregate  # noqa: E402
from indicators.models import Indicator, IndicatorAlias  # noqa: E402
from indicator_import_aliases import canonical_resolution_aliases, preferred_duplicate_rank  # noqa: E402
from organizations.models import Organization  # noqa: E402
from projects.models import Project, ProjectIndicator  # noqa: E402
from projects.project_indicator_links import ensure_project_indicator_link  # noqa: E402


DEFAULT_WORKBOOK = r"C:\Users\dekok\Downloads\2026 FINAL ANALYSIS May-March 03.02.2025 (1) (3).xlsx"
DEFAULT_SHEETS = ["TEBELOPELE", "BONEPWA", "BONELA"]
DEFAULT_PERIOD_START = "2025-10-01"
DEFAULT_PERIOD_END = "2025-12-31"
DEFAULT_PROJECT_CODE = "NAHPA2025/26"

DEFAULT_SUB_LABELS = ["Key Population", "Sex", "Age Range"]
AGE_SEX_HEADER_TOKEN = "age sex"
AGE_ONLY_HEADER_TOKEN = "age"
TOTAL_TOKENS = {"sub total", "subtotal", "total", "total male", "total female"}
SECTION_HEADER_TITLES = {
    "hiv prevention & control messages",
    "hiv prevention and control messages",
    "hiv prevention messages",
    "gender based violence",
    "commodity distribution",
    "stis",
    "ncds",
}
SHEET_ORG_ALIASES = {
    "gfc": "Gende Fountain",
    "gender fountain": "Gende Fountain",
    "gender fountain center": "Gende Fountain",
    "gender fountation centre": "Gende Fountain",
    "gende fountain": "Gende Fountain",
    "ovajua": "Ovajuha",
    "ovajhuha": "Ovajuha",
    "ovajuha": "Ovajuha",
    "inspired horizons": "Inpired Horizons",
    "inspired hozirons": "Inpired Horizons",
    "inpired horizons": "Inpired Horizons",
    "social dialogue": "Social Dialogue",
    "thabologo support group": "Thhabologo Support Group",
    "matlhogonolo charitable society": "Mathogonolo Charitable Society",
    "kebotlhokwa": "Kebotlhekwa",
    "leitlho la sechaba": "Leitho la Sechaba",
    "tozwimilidizha muti amuchile": "Tozwimilidzha Muti Amuchile Support Group",
    "ditsheganwe support group": "Ditshegwane Support Group",
    "guardian angels orphans society": "Guardian Angel Orphans Society",
    "bofaboneta": "BOFABONETHA",
    "botswana association of the dea": "Botswana Association for the Deaf",
    "botswana council of the disable": "BCD",
    "lesbians gays bisexuals of b": "LEGABIBO",
    "men for health gender justice": "Men for Health and Gender Justice Org.",
}
CODE_RESOLUTION_OVERRIDES = [
    {
        "code": "16",
        "source_title": "Number of people screened for NCDs behavioural risk factors",
        "indicator_names": ["Number of people screened for NCDs behavioural risk factors (tobacco use)"],
    },
    {
        "code": "19",
        "source_title": "Number of people screened for NCDs behavioural risk factors",
        "indicator_names": ["Number of people screened for NCDs behavioural risk factors (alcohol use)"],
    },
    {
        "code": "15f",
        "source_title": "Number of individual counselling sessions conducted for youth and vulnerable groups",
        "indicator_names": ["Number of individual counselling sessions conducted for youth and vulnerable groups (15f)"],
    },
    {
        "code": "15h",
        "source_title": "Number of individual counselling sessions conducted for youth and vulnerable groups",
        "indicator_names": ["Number of individual counselling sessions conducted for youth and vulnerable groups (15h)"],
    },
    {
        "code": "37b",
        "source_title": "Number of support group meetings held",
        "indicator_names": ["Number of support group meetings held (tobacco/alcohol)"],
    },
    {
        "code": "37d",
        "source_title": "Number of new members/continuing",
        "indicator_names": ["Number of new members/continuing (tobacco/alcohol)"],
    },
]

TITLE_RESOLUTION_ALIASES = {
    "number of people who report that their rights were violated and who sought redress": [
        "Number of people who sought redress.",
    ],
    "number of people who report that their rights were violated and who sought redness": [
        "Number of people who sought redress.",
    ],
    "number of people who reported collecting condoms for the repeated time": [
        "Number of people who reported collecting condoms for a repeated time.",
    ],
    "number of people who reported collecting condoms for the repeated time by age and sex": [
        "Number of people who reported collecting condoms for a repeated time.",
    ],
    "number of people who repeated collecting condoms for the repeated time": [
        "Number of people who reported collecting condoms for a repeated time.",
    ],
    "number of male condoms distributed": [
        "Number of condoms distributed",
    ],
    "number of stis referrals completed": [
        "Number of STI cases referrals completed",
    ],
    "number of plwhs provided with treatment literacy": [
        "Total Number of People Reached with ARV Based Prevention Messages",
    ],
    "number of plwh who tested positive for tb and are on treatment": [
        "Number of PLWH who tested positive for TB and are on treatment",
        "Number of people living with HIV who tested positive for TB and are on treatment",
    ],
    "number of people eligible for psychosocial services for gbv": [
        "Number of People eligible for psychosocial support on GBV.",
    ],
    "number of people referred for psychosocial services for gbv": [
        "Number of People referred for psychosocial support on GBV",
    ],
}
INDICATOR_TITLE_REGEX_REPLACEMENTS = [
    (r"\belligible\b", "eligible"),
    (r"\breffered\b", "referred"),
    (r"\brefferals\b", "referrals"),
    (r"\bpyschosocial\b", "psychosocial"),
    (r"\bpeopel\b", "people"),
    (r"\bbrailed condoms\b", "braille labelled condoms"),
    (r"\bPREP\b", "PrEP"),
    (r"\bEMTCT\b", "eMTCT"),
    (r"^Number of Total Number of ", "Total Number of "),
    (r"^Number of Number of ", "Number of "),
    (r"^Number of Number ", "Number "),
    (r"^Number eligible\b", "Number of people eligible"),
    (r"^Number referred\b", "Number of people referred"),
    (r"^Number linked\b", "Number of people linked"),
    (r"^Number provided with\b", "Number of people provided with"),
    (r"\bBY Age,\s*sex\b", ""),
]


def normalize(value):
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())).strip()


def clean_title(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip(" .")


def normalize_indicator_title(value):
    title = clean_title(value)
    for pattern, replacement in INDICATOR_TITLE_REGEX_REPLACEMENTS:
        title = re.sub(pattern, replacement, title, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", title).strip(" .")


def canonical_indicator_key(value):
    normalized = normalize(normalize_indicator_title(value))
    replacements = {
        "nunber": "number",
        "elligible": "eligible",
        "reffered": "referred",
        "peopel": "people",
        "pyschosocial": "psychosocial",
        "descrimination": "discrimination",
        "coodinators": "coordinators",
        "perforamance": "performance",
        "identifies needs": "identified needs",
        "field visists": "field visits",
        "virsual presentations": "visual presentations",
        "redness": "redress",
        "pwids": "pwid",
        "number of number of people": "number of people",
        "number of number of": "number of",
        "number of number": "number",
        "number eligible": "number of people eligible",
        "number referred": "number of people referred",
        "number linked": "number of people linked",
        "number provided with": "number of people provided with",
        "number of peoples": "number of people",
        "people living with hiv": "plwh",
        "persons living with hiv": "plwh",
        "total number of": "number of",
        "  ": " ",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    return re.sub(r"\s+", " ", normalized).strip()


def resolve_sheet_organization(sheet_name, organizations):
    normalized_sheet = normalize(sheet_name)
    aliases = [SHEET_ORG_ALIASES.get(normalized_sheet), sheet_name]

    for candidate in aliases:
        if not candidate:
            continue
        normalized_candidate = normalize(candidate)
        for organization in organizations:
            if normalize(organization.name) == normalized_candidate:
                return organization

    for organization in organizations:
        normalized_org = normalize(organization.name)
        if normalized_sheet and (
            normalized_sheet in normalized_org or normalized_org in normalized_sheet
        ):
            return organization

    return None


def resolve_sheet_organization_with_overrides(sheet_name, organizations, sheet_org_overrides=None):
    overrides = sheet_org_overrides or {}
    normalized_sheet = normalize(sheet_name)
    override_value = overrides.get(normalized_sheet)
    if override_value is not None:
        text_value = str(override_value).strip()
        if text_value:
            for organization in organizations:
                if text_value.isdigit() and int(text_value) == int(organization.id):
                    return organization
                if normalize(getattr(organization, "name", "")) == normalize(text_value):
                    return organization
                if normalize(getattr(organization, "code", "")) == normalize(text_value):
                    return organization

    return resolve_sheet_organization(sheet_name, organizations)


def get_code_resolution_overrides(code, title):
    code_token = str(code or "").strip().lower()
    title_key = canonical_indicator_key(title)
    matches = []
    for entry in CODE_RESOLUTION_OVERRIDES:
        if entry["code"] != code_token:
            continue
        if canonical_indicator_key(entry["source_title"]) != title_key:
            continue
        matches.extend(entry["indicator_names"])
    return matches


def get_title_resolution_aliases(title):
    return [
        *TITLE_RESOLUTION_ALIASES.get(canonical_indicator_key(title), []),
        *canonical_resolution_aliases(title),
    ]


def get_indicator_resolution_priority(indicator):
    code = str(getattr(indicator, "code", "") or "").upper()
    name = getattr(indicator, "name", "") or ""
    cleaned_name = clean_title(name)
    normalized_name = normalize_indicator_title(name)
    indicator_key = canonical_indicator_key(name)
    return (
        preferred_duplicate_rank(int(getattr(indicator, "id", 0) or 0), indicator_key),
        0 if getattr(indicator, "is_active", False) else 1,
        0 if code.startswith(("NAHPA", "SC_", "HIVPM_")) else 1,
        0 if not code.startswith("AUTO_") else 1,
        0 if cleaned_name == normalized_name else 1,
        -int(getattr(indicator, "id", 0) or 0),
    )


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def number_to_json(value):
    if value is None:
        return None
    if float(value).is_integer():
        return int(value)
    return float(value)


def first_numeric(values):
    for value in values:
        parsed = parse_number(value)
        if parsed is not None:
            return parsed
    return None


def last_numeric(values):
    for value in reversed(list(values)):
        parsed = parse_number(value)
        if parsed is not None:
            return parsed
    return None


def normalize_kp_label(value):
    label = clean_title(value)
    token = normalize(label)
    aliases = {
        "general population": "GENERAL POP.",
        "general pop": "GENERAL POP.",
        "general pop.": "GENERAL POP.",
        "general population ": "GENERAL POP.",
        "pwids": "PWID",
        "pwid": "PWID",
        "pwud": "PWUD",
        "fsm": "FSW",
        "fsw": "FSW",
        "msm": "MSM",
        "pwd": "PWD",
        "lgbtqi": "LGBTQI+",
        "lgbtqi+": "LGBTQI+",
    }
    return aliases.get(token, label or "GENERAL POP.")


def parse_sex(value):
    token = normalize(value)
    if token in {"male", "m"}:
        return "Male"
    if token in {"female", "f"}:
        return "Female"
    if token in {"all", "*"}:
        return "All"
    return None


def is_indicator_start(code):
    token = str(code or "").strip()
    return bool(re.fullmatch(r"\d+[a-z]?", token, flags=re.IGNORECASE))


def infer_matrix_first_dimension(disaggregates):
    labels = [clean_title(value) for value in disaggregates.keys() if clean_title(value)]
    if not labels:
        return "Disaggregate"

    normalized = {normalize(value) for value in labels}
    known_kp_tokens = {
        normalize("GENERAL POP."),
        normalize("PWID"),
        normalize("PWUD"),
        normalize("FSW"),
        normalize("MSM"),
        normalize("PWD"),
        normalize("LGBTQI+"),
    }
    if normalized and normalized.issubset(known_kp_tokens):
        return "Key Population"

    family_planning_keywords = (
        "iud",
        "injectable",
        "implant",
        "contraceptive",
        "pill",
        "emergency",
        "family planning",
    )
    if len(normalized) > 1 and all(
        any(keyword in token for keyword in family_planning_keywords)
        for token in normalized
    ):
        return "Family Planning"

    if len(labels) == 1:
        return labels[0]
    return "Disaggregate"


def should_force_all_second_dimension(code, title):
    _ = code  # code kept for future scoped overrides if needed
    title_key = canonical_indicator_key(title)
    return title_key in {
        canonical_indicator_key(
            "Number of people engaged with NCD prevention and control messages through social media"
        ),
        canonical_indicator_key(
            "Number of people engaged with NCD prevention and control messages through social media."
        ),
    }


def collapse_matrix_second_dimension_to_all(value):
    if not isinstance(value, dict):
        return value

    raw_disaggregates = value.get("disaggregates")
    if not isinstance(raw_disaggregates, dict):
        return value

    collapsed_disaggregates = {}
    computed_total = 0.0

    for primary, secondary_map in raw_disaggregates.items():
        if not isinstance(secondary_map, dict):
            continue

        band_totals = {}
        for band_map in secondary_map.values():
            if not isinstance(band_map, dict):
                continue
            for raw_band, raw_value in band_map.items():
                band = clean_title(raw_band) or "Value"
                number = parse_number(raw_value)
                if number is None:
                    continue
                existing = parse_number(band_totals.get(band))
                band_totals[band] = number_to_json((existing or 0.0) + number)
                computed_total += number

        if band_totals:
            collapsed_disaggregates[clean_title(primary) or "Disaggregate"] = {"All": band_totals}

    output = dict(value)
    output.pop("male", None)
    output.pop("female", None)
    output["disaggregates"] = collapsed_disaggregates

    summary_total = parse_number(value.get("total"))
    output["total"] = number_to_json(summary_total if summary_total is not None else computed_total)
    return output


def resolve_indicator(title, indicator_by_key, code=None, indicator_overrides=None, indicator_by_id=None):
    cleaned = normalize_indicator_title(title)
    override_key = canonical_indicator_key(cleaned)
    if indicator_overrides and indicator_by_id and override_key:
        override_value = indicator_overrides.get(override_key)
        if override_value is not None:
            try:
                override_id = int(override_value)
            except (TypeError, ValueError):
                override_id = None
            if override_id is not None and override_id in indicator_by_id:
                return indicator_by_id[override_id]

    candidates = [
        cleaned,
        re.sub(r"\s+", " ", cleaned.replace("  ", " ")),
        cleaned.replace("with with", "with"),
        cleaned.replace("PLWH reached", "people reached"),
        cleaned.replace("PLWH", "people"),
        cleaned.replace("KVP", "Key and Vulnerable Populations"),
        cleaned.replace("KVPs", "Key and Vulnerable Populations"),
        re.sub(
            r"^Number of PLWH reached with ",
            "Total Number of People Reached with ",
            cleaned,
            flags=re.IGNORECASE,
        ),
        re.sub(
            r"^Total Number of Key and Vulnerable Populations Reached with ",
            "Total Number of People Reached with ",
            cleaned,
            flags=re.IGNORECASE,
        ),
    ]

    for override in get_code_resolution_overrides(code, cleaned):
        candidates.insert(0, override)

    for alias in get_title_resolution_aliases(cleaned):
        candidates.insert(0, alias)

    keys = []
    seen_keys = set()
    for candidate in candidates:
        if not candidate:
            continue
        for key in (
            canonical_indicator_key(candidate),
            canonical_indicator_key(f"Number of {candidate}"),
            canonical_indicator_key(candidate.replace("Total Number of ", "Number of ")),
        ):
            if key and key not in seen_keys:
                seen_keys.add(key)
                keys.append(key)

    for key in keys:
        indicator = indicator_by_key.get(key)
        if indicator:
            return indicator
    return None


def find_matrix_header_row(rows, start_index):
    for index in range(max(0, start_index - 20), min(len(rows), start_index + 2)):
        row = rows[index]
        header_token = normalize(row[5] if len(row) > 5 else "")
        if header_token in {AGE_SEX_HEADER_TOKEN, AGE_ONLY_HEADER_TOKEN}:
            return index
    return None


def extract_age_bands(header_row):
    bands = []
    for cell in header_row[6:]:
        token = clean_title(cell)
        if not token:
            continue
        normalized = normalize(token)
        if normalized in TOTAL_TOKENS:
            break
        bands.append(token)
    return bands


def build_total_payload(row):
    numeric_candidates = [parse_number(cell) for cell in row[4:]]
    numeric_candidates = [value for value in numeric_candidates if value is not None]
    total = numeric_candidates[-1] if numeric_candidates else 0
    return {"total": number_to_json(total)}


def parse_matrix_block(rows, start_index):
    header_index = find_matrix_header_row(rows, start_index)
    if header_index is None:
        return None

    age_bands = extract_age_bands(rows[header_index])
    if not age_bands:
        return None

    disaggregates = {}
    male_total = 0.0
    female_total = 0.0
    male_total_from_rows = False
    female_total_from_rows = False
    summary_male_total = None
    summary_female_total = None
    summary_total = None
    current_kp = None
    row_cursor = start_index

    while row_cursor < len(rows):
        row = rows[row_cursor]
        code = clean_title(row[1] if len(row) > 1 else "")
        if row_cursor > start_index and is_indicator_start(code):
            break

        kp_cell = clean_title(row[4] if len(row) > 4 else "")
        sex = parse_sex(row[5] if len(row) > 5 else "")

        if kp_cell:
            current_kp = kp_cell

        if kp_cell and normalize(kp_cell) == "total male":
            summary_male_total = first_numeric(row[6:])
            row_cursor += 1
            continue

        if kp_cell and normalize(kp_cell) == "total female":
            summary_female_total = first_numeric(row[6:])
            row_cursor += 1
            continue

        if kp_cell and normalize(kp_cell) in {"sub total", "subtotal", "total"}:
            subtotal_index = 6 + len(age_bands)
            summary_total = parse_number(row[subtotal_index] if len(row) > subtotal_index else None)
            if summary_total is None:
                summary_total = parse_number(row[subtotal_index + 1] if len(row) > subtotal_index + 1 else None)
            if summary_total is None:
                summary_total = first_numeric(row[6:])
            summary_row_value = first_numeric(row[6:])
            if sex == "Male" and summary_row_value is not None:
                summary_male_total = summary_row_value
            if sex == "Female" and summary_row_value is not None:
                summary_female_total = summary_row_value
            row_cursor += 1
            continue

        if (
            not kp_cell
            and sex
            and current_kp
            and normalize(current_kp) in {"sub total", "subtotal", "total"}
        ):
            summary_row_value = first_numeric(row[6:])
            if sex == "Male" and summary_row_value is not None:
                summary_male_total = summary_row_value
            if sex == "Female" and summary_row_value is not None:
                summary_female_total = summary_row_value
            row_cursor += 1
            continue

        # Some indicator matrices only have AGE/SEX rows and no explicit
        # first-dimension label in the disaggregate column.
        if not current_kp and sex:
            current_kp = "All"

        # Some sheets encode matrix rows as age-only without a clear sex token.
        # When we can see numeric age-band values, keep the row as a valid
        # matrix line under an "All" bucket instead of dropping it.
        if not sex and current_kp:
            has_numeric_band_values = False
            for band_offset in range(len(age_bands)):
                if parse_number(row[6 + band_offset] if len(row) > 6 + band_offset else None) is not None:
                    has_numeric_band_values = True
                    break
            if has_numeric_band_values:
                sex = "All"

        if not current_kp or not sex:
            row_cursor += 1
            continue

        normalized_kp = normalize_kp_label(current_kp)
        disaggregates.setdefault(normalized_kp, {})
        disaggregates[normalized_kp].setdefault(sex, {})

        subtotal_index = 6 + len(age_bands)
        subtotal_value = parse_number(row[subtotal_index] if len(row) > subtotal_index else None)
        running_total = 0
        for band_offset, band in enumerate(age_bands):
            value = parse_number(row[6 + band_offset] if len(row) > 6 + band_offset else None)
            if value is None:
                continue
            disaggregates[normalized_kp][sex][band] = number_to_json(value)
            running_total += value

        if subtotal_value is not None and not disaggregates[normalized_kp][sex]:
            disaggregates[normalized_kp][sex]["Value"] = number_to_json(subtotal_value)
            running_total = subtotal_value

        if sex == "Male":
            male_total += running_total
            male_total_from_rows = True
        if sex == "Female":
            female_total += running_total
            female_total_from_rows = True

        row_cursor += 1

    has_matrix = any(
        any(bands for bands in sex_map.values())
        for sex_map in disaggregates.values()
    )
    if not has_matrix:
        return None

    if not male_total_from_rows and summary_male_total is not None:
        male_total = summary_male_total
    if not female_total_from_rows and summary_female_total is not None:
        female_total = summary_female_total

    computed_total = male_total + female_total
    if summary_total is not None and (
        computed_total == 0 or abs(summary_total - computed_total) <= 1e-9
    ):
        total = summary_total
    else:
        total = computed_total

    return {
        "value": {
            "male": number_to_json(male_total),
            "female": number_to_json(female_total),
            "total": number_to_json(total),
            "disaggregates": disaggregates,
        },
        "next_index": row_cursor,
        "is_matrix": True,
        "sub_labels": [infer_matrix_first_dimension(disaggregates), "Sex", "Age Range"],
    }


def parse_sheet(ws):
    rows = [
        [cell for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]

    parsed = []
    row_index = 0
    while row_index < len(rows):
        row = rows[row_index]
        code = clean_title(row[1] if len(row) > 1 else "")
        title = normalize_indicator_title(row[2] if len(row) > 2 else "")
        if not is_indicator_start(code) or not title:
            row_index += 1
            continue

        if normalize(title) in SECTION_HEADER_TITLES:
            row_index += 1
            continue

        matrix_result = parse_matrix_block(rows, row_index)
        if matrix_result:
            matrix_value = matrix_result["value"]
            matrix_sub_labels = matrix_result["sub_labels"]
            if should_force_all_second_dimension(code, title):
                matrix_value = collapse_matrix_second_dimension_to_all(matrix_value)
                if len(matrix_sub_labels) >= 3:
                    matrix_sub_labels = [matrix_sub_labels[0], matrix_sub_labels[2]]

            parsed.append(
                {
                    "code": code,
                    "title": title,
                    "value": matrix_value,
                    "is_matrix": True,
                    "sub_labels": matrix_sub_labels,
                }
            )
            row_index = matrix_result["next_index"]
            continue

        parsed.append(
            {
                "code": code,
                "title": title,
                "value": build_total_payload(row),
                "is_matrix": False,
                "sub_labels": [],
            }
        )
        row_index += 1

    deduped = {}
    for item in parsed:
        deduped[item["code"]] = item
    return list(deduped.values())


def update_indicator_disaggregation(indicator, sub_labels, apply_changes):
    def canonical_sub_label(value):
        token = normalize(value)
        aliases = {
            "key population": "Key Population",
            "key populations": "Key Population",
            "service category": "Key Population",
            "disaggregate": "Disaggregate",
            "disaggregation": "Disaggregate",
            "sex": "Sex",
            "gender": "Sex",
            "age range": "Age Range",
            "age": "Age Range",
            "age group": "Age Range",
        }
        if not token:
            return None
        return aliases.get(token, clean_title(value))

    existing = list(indicator.sub_labels or [])
    def canonicalize_labels(values):
        output = []
        for value in values:
            canonical = canonical_sub_label(value)
            if canonical and canonical not in output:
                output.append(canonical)
        return output

    canonical_target = canonicalize_labels(sub_labels)
    canonical_existing = canonicalize_labels(existing)

    # If parser defaults to Key Population but indicator already has a clear
    # custom first dimension, preserve the existing semantic dimension.
    if (
        canonical_target == ["Key Population", "Sex", "Age Range"]
        and len(canonical_existing) == 3
        and canonical_existing[1:] == ["Sex", "Age Range"]
        and canonical_existing[0] not in {"Key Population", "Disaggregate"}
    ):
        canonical_target = canonical_existing

    changed_fields = []
    if canonical_target and canonical_target != existing:
        indicator.sub_labels = canonical_target
        changed_fields.append("sub_labels")

    if indicator and changed_fields and apply_changes:
        indicator.save(update_fields=changed_fields)
    return bool(changed_fields)


def build_args():
    parser = argparse.ArgumentParser(
        description="Import selected quarterly workbook sheets into aggregates without creating indicators."
    )
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    parser.add_argument("--project-id", type=int, default=0)
    parser.add_argument("--project-code", default=DEFAULT_PROJECT_CODE)
    parser.add_argument("--period-start", default=DEFAULT_PERIOD_START)
    parser.add_argument("--period-end", default=DEFAULT_PERIOD_END)
    parser.add_argument("--sheets", nargs="+", default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report-path", default="")
    parser.add_argument("--indicator-overrides-path", default="")
    parser.add_argument("--sheet-org-overrides-path", default="")
    return parser.parse_args()


def main():
    args = build_args()
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    project = None
    if args.project_id:
        project = Project.objects.only("id", "code", "name").filter(id=args.project_id).first()
    if project is None and args.project_code:
        project = Project.objects.only("id", "code", "name").filter(code__iexact=args.project_code).first()
    if project is None:
        raise SystemExit("Project not found.")

    workbook = load_workbook(workbook_path, data_only=True)
    organizations = list(Organization.objects.only("id", "name", "code"))

    indicator_overrides = {}
    if args.indicator_overrides_path:
        indicator_overrides_path = Path(args.indicator_overrides_path)
        if indicator_overrides_path.exists():
            try:
                raw_indicator_overrides = json.loads(indicator_overrides_path.read_text(encoding="utf-8"))
            except Exception:
                raw_indicator_overrides = {}
            if isinstance(raw_indicator_overrides, dict):
                for raw_title, raw_value in raw_indicator_overrides.items():
                    key = canonical_indicator_key(raw_title)
                    if not key:
                        continue
                    indicator_overrides[key] = raw_value

    sheet_org_overrides = {}
    if args.sheet_org_overrides_path:
        sheet_org_overrides_path = Path(args.sheet_org_overrides_path)
        if sheet_org_overrides_path.exists():
            try:
                raw_sheet_org_overrides = json.loads(sheet_org_overrides_path.read_text(encoding="utf-8"))
            except Exception:
                raw_sheet_org_overrides = {}
            if isinstance(raw_sheet_org_overrides, dict):
                for raw_sheet_name, raw_org in raw_sheet_org_overrides.items():
                    key = normalize(raw_sheet_name)
                    if not key:
                        continue
                    sheet_org_overrides[key] = raw_org

    candidate_sheets = args.sheets or [
        sheet_name
        for sheet_name in workbook.sheetnames
        if resolve_sheet_organization_with_overrides(sheet_name, organizations, sheet_org_overrides) is not None
    ]
    selected_sheets = []
    missing_sheets = []
    missing_orgs = []
    for sheet_name in candidate_sheets:
        if sheet_name not in workbook.sheetnames:
            missing_sheets.append(sheet_name)
            continue
        if resolve_sheet_organization_with_overrides(sheet_name, organizations, sheet_org_overrides) is None:
            missing_orgs.append(sheet_name)
            continue
        selected_sheets.append(sheet_name)

    if args.sheets and missing_sheets:
        raise SystemExit(f"Sheets not found in workbook: {', '.join(missing_sheets)}")
    if args.sheets and missing_orgs:
        raise SystemExit(f"Organizations not found for sheets: {', '.join(missing_orgs)}")
    if not selected_sheets:
        raise SystemExit("No organization sheets found in workbook.")

    indicators = list(Indicator.objects.only("id", "name", "sub_labels", "is_active", "code"))
    indicator_by_id = {indicator.id: indicator for indicator in indicators}
    indicator_by_key = {}
    for indicator in indicators:
        for candidate in [indicator.name, *canonical_resolution_aliases(indicator.name)]:
            key = canonical_indicator_key(candidate)
            if not key:
                continue
            existing = indicator_by_key.get(key)
            if existing is None or get_indicator_resolution_priority(indicator) < get_indicator_resolution_priority(existing):
                indicator_by_key[key] = indicator
    for alias in IndicatorAlias.objects.select_related("indicator").filter(is_active=True):
        indicator = alias.indicator
        for candidate in [alias.name, *canonical_resolution_aliases(alias.name)]:
            key = canonical_indicator_key(candidate)
            if not key:
                continue
            existing = indicator_by_key.get(key)
            if existing is None or get_indicator_resolution_priority(indicator) < get_indicator_resolution_priority(existing):
                indicator_by_key[key] = indicator

    report = {
        "workbook": str(workbook_path),
        "project": {"id": project.id, "name": project.name, "code": project.code},
        "period_start": args.period_start,
        "period_end": args.period_end,
        "sheets": {},
        "summary": {},
    }

    matched = 0
    unknown = Counter()
    updated_indicators = []
    upserted_aggregates = []
    created_project_assignments = 0

    for sheet_name in selected_sheets:
        organization = resolve_sheet_organization_with_overrides(
            sheet_name,
            organizations,
            sheet_org_overrides,
        )
        if organization is None:
            raise SystemExit(f"Organization not found for sheet: {sheet_name}")
        parsed_rows = parse_sheet(workbook[sheet_name])
        sheet_report = {
            "organization_id": organization.id,
            "parsed_rows": len(parsed_rows),
            "matched_rows": [],
            "unknown_rows": [],
        }

        for item in parsed_rows:
            indicator = resolve_indicator(
                item["title"],
                indicator_by_key,
                code=item["code"],
                indicator_overrides=indicator_overrides,
                indicator_by_id=indicator_by_id,
            )
            if indicator is None:
                unknown[item["title"]] += 1
                sheet_report["unknown_rows"].append(item["title"])
                continue

            matched += 1
            changed = False
            if item["is_matrix"]:
                changed = update_indicator_disaggregation(
                    indicator,
                    item["sub_labels"],
                    apply_changes=not args.dry_run,
                ) or changed

            project_assignment_created = False
            if not args.dry_run:
                indicator.organizations.add(organization)
                project_assignment_created = ensure_project_indicator_link(project, indicator)
            else:
                project_assignment_created = not ProjectIndicator.objects.filter(
                    project=project,
                    indicator=indicator,
                ).exists()

            if project_assignment_created:
                created_project_assignments += 1

            aggregate_defaults = {
                "value": item["value"],
                "notes": f"Imported from {workbook_path.name} | sheet={sheet_name} | code={item['code']}",
                "status": "pending",
                "reviewed_at": None,
                "reviewed_by": None,
            }

            aggregate_action = "would_update"
            existing = Aggregate.objects.only("id").filter(
                indicator=indicator,
                project=project,
                organization=organization,
                period_start=args.period_start,
                period_end=args.period_end,
            ).first()
            if existing is None:
                aggregate_action = "would_create"
            if not args.dry_run:
                aggregate, created = Aggregate.objects.update_or_create(
                    indicator=indicator,
                    project=project,
                    organization=organization,
                    period_start=args.period_start,
                    period_end=args.period_end,
                    defaults=aggregate_defaults,
                )
                aggregate_action = "created" if created else "updated"
                upserted_aggregates.append(aggregate.id)

            if changed:
                updated_indicators.append(indicator.id)

            sheet_report["matched_rows"].append(
                {
                    "code": item["code"],
                    "title": item["title"],
                    "indicator_id": indicator.id,
                    "indicator_name": indicator.name,
                    "matrix": item["is_matrix"],
                    "aggregate_action": aggregate_action,
                }
            )

        report["sheets"][sheet_name] = sheet_report

    report["summary"] = {
        "matched_rows": matched,
        "unknown_rows": sum(unknown.values()),
        "unique_unknown_titles": len(unknown),
        "updated_indicator_count": len(set(updated_indicators)),
        "project_assignments_created": created_project_assignments,
        "aggregate_upserts": len(upserted_aggregates),
        "dry_run": args.dry_run,
        "unknown_titles": dict(unknown.most_common()),
        "indicator_override_count": len(indicator_overrides),
        "sheet_org_override_count": len(sheet_org_overrides),
    }

    if args.report_path:
        report_path = Path(args.report_path)
    else:
        report_path = Path.cwd() / "reports" / "selected-q3-workbook-import-report.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(json.dumps(report["summary"], indent=2))
    print(f"Saved report: {report_path}")


if __name__ == "__main__":
    main()
