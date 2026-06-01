import argparse
import difflib
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


BACKEND_ROOT = Path(os.environ.get("BONASO_DJANGO_ROOT", r"C:\Projects\django_backend"))
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django  # noqa: E402

django.setup()

from indicators.models import Indicator  # noqa: E402
from organizations.models import Organization  # noqa: E402
from projects.models import Project, ProjectIndicator, ProjectIndicatorOrganizationTarget  # noqa: E402


DEFAULT_WORKBOOK = (
    r"C:\Users\dekok\Downloads\Social Contracting 2025-Final Targets & indicators 5 Clusters 14 AUGUST 202526.xlsx"
)
DEFAULT_PROJECT_CODE = "NAHPA2025/26"
DEFAULT_SKIP_SHEETS = ["INSTRUCTIONS", "CLUSTER 6-BBCA"]

KNOWN_SHEET_ORG_ALIASES = {
    "cluster 1 bonela": ["BONELA"],
    "cluster 2 tebelopele": ["TEBELOPELE"],
    "cluster 3 bonepwa": ["BONEPWA"],
    "cluster 3 bonepwa ": ["BONEPWA"],
    "cluster 3 bonepwa+": ["BONEPWA"],
    "cluster 4 men boys": ["MBGE", "Men for Health and Gender Justice Org."],
    "cluster 5 makgabaneng": ["MAKGABANENG"],
    "cluster 7 bonaso": ["BONASO"],
}

HEADER_SCAN_ROWS = 12
HEADER_SCAN_COLS = 40
MIN_FUZZY_RATIO = 0.9
TITLE_RESOLUTION_ALIASES = {
    "number of people who reported collecting condoms for the repeated time": [
        "Number of people who reported collecting condoms for a repeated time.",
    ],
    "number of people who reported collecting condoms for the repeated time by age and sex": [
        "Number of people who reported collecting condoms for a repeated time.",
    ],
    "number of people who repeated collecting condoms for the repeated time": [
        "Number of people who reported collecting condoms for a repeated time.",
    ],
    "number of male condoms distributed": [
        "Number of condoms distributed",
    ],
    "number of stis referrals completed": [
        "Number of STI cases referrals completed",
    ],
    "number of plwhs provided with treatment literacy": [
        "Total Number of People Reached with ARV Based Prevention Messages",
    ],
    "number of plwh who tested positive for tb and are on treatment": [
        "Number of PLWH who tested positive for TB and are on treatment",
        "Number of people living with HIV who tested positive for TB and are on treatment",
    ],
}


def normalize(value):
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())).strip()


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip(" .")


def canonical_indicator_key(value):
    normalized = normalize(value)
    replacements = {
        "nunber": "number",
        "refferals": "referrals",
        "coodinator": "coordinator",
        "coodinators": "coordinators",
        "coodinatio": "coordination",
        "coodinating": "coordinating",
        "visists": "visits",
        "perforamance": "performance",
        "identifies": "identified",
        "identifies needs": "identified needs",
        "elligible": "eligible",
        "reffered": "referred",
        "pyschosocial": "psychosocial",
        "descrimination": "discrimination",
        "field visists": "field visits",
        "virsual presentations": "visual presentations",
        "redness": "redress",
        "targetted": "targeted",
        "number of number of people": "number of people",
        "number of number of": "number of",
        "total number of": "number of",
        "people living with hiv": "plwh",
        "persons living with hiv": "plwh",
        "people engaged with": "people reached with",
        "  ": " ",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    return re.sub(r"\s+", " ", normalized).strip()


def get_title_resolution_aliases(title):
    return TITLE_RESOLUTION_ALIASES.get(canonical_indicator_key(title), [])


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def number_to_storage(value):
    if value is None:
        return 0
    numeric = float(value)
    if numeric.is_integer():
        return int(numeric)
    return numeric


def build_indicator_lookup():
    indicators = list(Indicator.objects.only("id", "name", "code", "is_active"))
    by_key = {}
    existing_codes = set()

    def lookup_priority(indicator):
        code = clean_text(getattr(indicator, "code", "")).upper()
        return (
            0 if getattr(indicator, "is_active", False) else 1,
            0 if code.startswith(("NAHPA", "SC_", "HIVPM_")) else 1,
            0 if not code.startswith("AUTO_") else 1,
            indicator.id,
        )

    for indicator in indicators:
        code_value = clean_text(getattr(indicator, "code", ""))
        if code_value:
            existing_codes.add(normalize(code_value))
        names = [indicator.name, getattr(indicator, "short_name", None)]
        for name in names:
            key = canonical_indicator_key(name)
            if not key:
                continue
            existing = by_key.get(key)
            if existing is None or lookup_priority(indicator) < lookup_priority(existing):
                by_key[key] = indicator
    return indicators, by_key, existing_codes


def make_indicator_code(name, existing_codes):
    max_length = Indicator._meta.get_field("code").max_length or 50
    base = re.sub(r"[^A-Z0-9]+", "_", clean_text(name).upper()).strip("_")
    if not base:
        base = "INDICATOR"

    prefix = "SC_"
    base_room = max(1, max_length - len(prefix))
    candidate = f"{prefix}{base[:base_room]}"
    suffix_index = 1
    while normalize(candidate) in existing_codes:
        suffix = f"_{suffix_index}"
        base_room = max(1, max_length - len(prefix) - len(suffix))
        candidate = f"{prefix}{base[:base_room]}{suffix}"
        suffix_index += 1

    existing_codes.add(normalize(candidate))
    return candidate


def create_missing_indicator(row_title, organization, indicator_by_key, existing_codes, dry_run=False):
    clean_name = clean_text(row_title)
    key = canonical_indicator_key(clean_name)
    if not clean_name or not key:
        return None, None

    if dry_run:
        return None, "would_create_indicator"

    indicator = Indicator.objects.create(
        name=clean_name,
        code=make_indicator_code(clean_name, existing_codes),
        category="hiv_prevention",
        type="number",
        aggregation_method="sum",
        is_active=True,
    )
    indicator.organizations.add(organization)

    indicator_by_key[key] = indicator
    return indicator, "created_indicator"


def resolve_indicator(title, indicator_by_key):
    cleaned = clean_text(title)
    if not cleaned:
        return None, None

    candidates = [
        cleaned,
        cleaned.replace("with with", "with"),
        cleaned.replace("people engaged with", "people reached with"),
        re.sub(r"^total number of ", "number of ", cleaned, flags=re.IGNORECASE),
        re.sub(r"^number of ", "total number of ", cleaned, flags=re.IGNORECASE),
    ]
    for alias in get_title_resolution_aliases(cleaned):
        candidates.insert(0, alias)

    candidate_keys = []
    for candidate in candidates:
        key = canonical_indicator_key(candidate)
        if key and key not in candidate_keys:
            candidate_keys.append(key)
        alt = canonical_indicator_key(f"number of {candidate}")
        if alt and alt not in candidate_keys:
            candidate_keys.append(alt)

    for key in candidate_keys:
        indicator = indicator_by_key.get(key)
        if indicator:
            return indicator, "exact"

    best_match = None
    best_ratio = 0.0
    for candidate_key in candidate_keys:
        for known_key, indicator in indicator_by_key.items():
            ratio = difflib.SequenceMatcher(None, candidate_key, known_key).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_match = indicator

    if best_match and best_ratio >= MIN_FUZZY_RATIO:
        return best_match, f"fuzzy:{best_ratio:.3f}"

    return None, None


def detect_header(ws):
    for row_index in range(1, min(ws.max_row, HEADER_SCAN_ROWS) + 1):
        token_by_col = {}
        for col_index in range(1, min(ws.max_column, HEADER_SCAN_COLS) + 1):
            token = normalize(ws.cell(row_index, col_index).value)
            if token:
                token_by_col[col_index] = token

        q1_col = next((col for col, token in token_by_col.items() if token.startswith("q1")), None)
        q2_col = next((col for col, token in token_by_col.items() if token.startswith("q2")), None)
        q3_col = next((col for col, token in token_by_col.items() if token.startswith("q3")), None)
        q4_col = next((col for col, token in token_by_col.items() if token.startswith("q4")), None)
        annual_col = next(
            (col for col, token in token_by_col.items() if token.startswith("annual")),
            None,
        )
        indicator_col = next((col for col, token in token_by_col.items() if token == "indicator"), None)
        if indicator_col is None:
            indicator_col = next(
                (col for col, token in token_by_col.items() if "indicator" in token),
                None,
            )

        if q1_col and q2_col and q3_col and q4_col and indicator_col:
            return {
                "row_index": row_index,
                "q1_col": q1_col,
                "q2_col": q2_col,
                "q3_col": q3_col,
                "q4_col": q4_col,
                "annual_col": annual_col,
                "indicator_col": indicator_col,
            }
    return None


def resolve_organization_for_sheet(sheet_name, organizations_by_normalized):
    sheet_token = normalize(sheet_name)
    alias_candidates = KNOWN_SHEET_ORG_ALIASES.get(sheet_token, [])
    for alias in alias_candidates:
        organization = organizations_by_normalized.get(normalize(alias))
        if organization:
            return organization

    for fragment in ["cluster 1", "cluster 2", "cluster 3", "cluster 4", "cluster 5", "cluster 6", "cluster 7"]:
        sheet_token = sheet_token.replace(fragment, "").strip()

    if sheet_token.startswith("cluster"):
        sheet_token = sheet_token.replace("cluster", "").strip()

    for org_key, organization in organizations_by_normalized.items():
        if not org_key:
            continue
        if sheet_token == org_key or sheet_token in org_key or org_key in sheet_token:
            return organization

    return None


def is_meta_or_empty_indicator(value):
    token = normalize(value)
    if not token:
        return True
    return token in {"indicator", "numerator", "denominator"}


def parse_sheet_rows(ws, header):
    rows = []
    for row_index in range(header["row_index"] + 1, ws.max_row + 1):
        indicator_value = ws.cell(row_index, header["indicator_col"]).value
        indicator_text = clean_text(indicator_value)
        if is_meta_or_empty_indicator(indicator_text):
            continue

        q1 = parse_number(ws.cell(row_index, header["q1_col"]).value)
        q2 = parse_number(ws.cell(row_index, header["q2_col"]).value)
        q3 = parse_number(ws.cell(row_index, header["q3_col"]).value)
        q4 = parse_number(ws.cell(row_index, header["q4_col"]).value)
        annual = (
            parse_number(ws.cell(row_index, header["annual_col"]).value)
            if header["annual_col"]
            else None
        )

        if all(value is None for value in [q1, q2, q3, q4, annual]):
            continue

        rows.append(
            {
                "row_index": row_index,
                "indicator_text": indicator_text,
                "q1_target": number_to_storage(q1),
                "q2_target": number_to_storage(q2),
                "q3_target": number_to_storage(q3),
                "q4_target": number_to_storage(q4),
                "annual_target": number_to_storage(
                    annual if annual is not None else (q1 or 0) + (q2 or 0) + (q3 or 0) + (q4 or 0)
                ),
            }
        )

    deduped = {}
    for row in rows:
        deduped[row["indicator_text"]] = row
    return list(deduped.values())


def build_args():
    parser = argparse.ArgumentParser(
        description="Import Social Contracting workbook quarterly indicator targets into project organization targets."
    )
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    parser.add_argument("--project-code", default=DEFAULT_PROJECT_CODE)
    parser.add_argument("--project-id", type=int, default=0)
    parser.add_argument("--skip-sheets", nargs="*", default=DEFAULT_SKIP_SHEETS)
    parser.add_argument("--create-missing-indicators", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report-path", default="")
    return parser.parse_args()


def main():
    args = build_args()
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    project = None
    if args.project_id:
        project = Project.objects.filter(id=args.project_id).first()
    if project is None and args.project_code:
        project = Project.objects.filter(code__iexact=args.project_code).first()
    if project is None:
        raise SystemExit("Project not found.")

    organizations = list(Organization.objects.all())
    organizations_by_normalized = {normalize(org.name): org for org in organizations}
    indicators, indicator_by_key, existing_codes = build_indicator_lookup()

    wb = load_workbook(workbook_path, data_only=True)
    skip_sheet_tokens = {normalize(name) for name in args.skip_sheets}

    report = {
        "workbook": str(workbook_path),
        "project": {"id": project.id, "name": project.name, "code": project.code},
        "dry_run": args.dry_run,
        "skipped_sheets": [],
        "missing_org_sheets": [],
        "missing_header_sheets": [],
        "sheets": {},
        "summary": {},
    }

    created_targets = 0
    updated_targets = 0
    unchanged_targets = 0
    unmatched_indicator_rows = 0
    fuzzy_matches = 0
    exact_matches = 0
    created_matches = 0
    processed_rows = 0
    created_project_links = 0
    created_indicators = 0

    for sheet_name in wb.sheetnames:
        if normalize(sheet_name) in skip_sheet_tokens:
            report["skipped_sheets"].append(sheet_name)
            continue

        organization = resolve_organization_for_sheet(sheet_name, organizations_by_normalized)
        if not organization:
            report["missing_org_sheets"].append(sheet_name)
            continue

        ws = wb[sheet_name]
        header = detect_header(ws)
        if not header:
            report["missing_header_sheets"].append(sheet_name)
            continue

        parsed_rows = parse_sheet_rows(ws, header)
        sheet_report = {
            "organization_id": organization.id,
            "organization_name": organization.name,
            "parsed_rows": len(parsed_rows),
            "matched_rows": [],
            "unmatched_rows": [],
        }

        for row in parsed_rows:
            processed_rows += 1
            indicator, match_mode = resolve_indicator(row["indicator_text"], indicator_by_key)
            if indicator is None:
                if args.create_missing_indicators:
                    indicator, create_mode = create_missing_indicator(
                        row["indicator_text"],
                        organization,
                        indicator_by_key,
                        existing_codes,
                        dry_run=args.dry_run,
                    )
                    if indicator is not None:
                        created_indicators += 1
                        match_mode = create_mode

                if indicator is None:
                    unmatched_indicator_rows += 1
                    sheet_report["unmatched_rows"].append(
                        {
                            "row_index": row["row_index"],
                            "indicator_text": row["indicator_text"],
                        }
                    )
                    continue

            if match_mode and "create_indicator" in match_mode:
                created_matches += 1
            elif match_mode and match_mode.startswith("fuzzy"):
                fuzzy_matches += 1
            else:
                exact_matches += 1

            if not args.dry_run:
                indicator.organizations.add(organization)

            project_indicator, project_link_created = ProjectIndicator.objects.get_or_create(
                project=project,
                indicator=indicator,
            )
            if project_link_created:
                created_project_links += 1

            desired_q1 = row["q1_target"]
            desired_q2 = row["q2_target"]
            desired_q3 = row["q3_target"]
            desired_q4 = row["q4_target"]
            desired_total = row["annual_target"]

            target, target_created = ProjectIndicatorOrganizationTarget.objects.get_or_create(
                project_indicator=project_indicator,
                organization=organization,
                defaults={
                    "q1_target": desired_q1,
                    "q2_target": desired_q2,
                    "q3_target": desired_q3,
                    "q4_target": desired_q4,
                    "target_value": desired_total,
                    "current_value": 0,
                    "baseline_value": 0,
                },
            )

            action = "unchanged"
            changed_fields = []
            if target_created:
                action = "created"
                created_targets += 1
            else:
                field_updates = {
                    "q1_target": desired_q1,
                    "q2_target": desired_q2,
                    "q3_target": desired_q3,
                    "q4_target": desired_q4,
                    "target_value": desired_total,
                }
                for field_name, next_value in field_updates.items():
                    current_value = getattr(target, field_name)
                    if current_value != next_value:
                        setattr(target, field_name, next_value)
                        changed_fields.append(field_name)

                if changed_fields:
                    action = "updated"
                    updated_targets += 1
                    if not args.dry_run:
                        target.save(update_fields=changed_fields)
                else:
                    unchanged_targets += 1

            sheet_report["matched_rows"].append(
                {
                    "row_index": row["row_index"],
                    "indicator_id": indicator.id,
                    "indicator_name": indicator.name,
                    "match_mode": match_mode,
                    "q1_target": desired_q1,
                    "q2_target": desired_q2,
                    "q3_target": desired_q3,
                    "q4_target": desired_q4,
                    "annual_target": desired_total,
                    "action": action if not args.dry_run else f"would_{action}",
                }
            )

        report["sheets"][sheet_name] = sheet_report

    report["summary"] = {
        "project_id": project.id,
        "total_indicators_in_system": len(indicators),
        "processed_rows": processed_rows,
        "exact_matches": exact_matches,
        "fuzzy_matches": fuzzy_matches,
        "created_matches": created_matches,
        "unmatched_indicator_rows": unmatched_indicator_rows,
        "created_project_links": created_project_links,
        "created_indicators": created_indicators,
        "created_targets": created_targets,
        "updated_targets": updated_targets,
        "unchanged_targets": unchanged_targets,
        "missing_org_sheet_count": len(report["missing_org_sheets"]),
        "missing_header_sheet_count": len(report["missing_header_sheets"]),
        "skipped_sheet_count": len(report["skipped_sheets"]),
        "dry_run": args.dry_run,
    }

    if args.report_path:
        report_path = Path(args.report_path)
    else:
        report_path = Path.cwd() / "reports" / "social-contracting-quarter-targets-import-report.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(json.dumps(report["summary"], indent=2))
    print(f"Saved report: {report_path}")


if __name__ == "__main__":
    main()
