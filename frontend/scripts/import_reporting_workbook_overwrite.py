import argparse
import json
import os
import sys
from pathlib import Path


BACKEND_ROOT = Path(os.environ.get("BONASO_DJANGO_ROOT", "/app"))
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django  # noqa: E402

django.setup()

from django.core.management.base import CommandError  # noqa: E402
from django.db import transaction  # noqa: E402

from aggregates.models import Aggregate  # noqa: E402
from indicators.models import Indicator  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from organizations.models import Organization  # noqa: E402
from projects.models import Project, ProjectIndicator  # noqa: E402
from uploads.management.commands.import_reporting_workbook_overwrite import (  # noqa: E402
    IndicatorResolver,
    OrganizationResolver,
    build_disaggregation_config,
    build_ordered_sub_labels,
    canonical_indicator_name,
    extract_section_value,
    find_matrix_sheet_name,
    get_age_band_mapping,
    is_skipped_sheet,
    merge_json_values,
    merge_disaggregation_configs,
    parse_matrix_sheet,
    parse_sections,
    unique_indicator_code,
)


def build_args():
    parser = argparse.ArgumentParser(
        description=(
            "Import a structured quarterly reporting workbook and overwrite existing "
            "aggregate values for the selected period instead of merging into them."
        )
    )
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--project-id", type=int, required=True)
    parser.add_argument("--coordinator-id", type=int)
    parser.add_argument("--period-start", required=True)
    parser.add_argument("--period-end", required=True)
    parser.add_argument("--category", default="ncd")
    parser.add_argument("--report-path", default="")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def ensure_project_indicator(project: Project, indicator: Indicator):
    project_indicator, _ = ProjectIndicator.objects.get_or_create(
        project=project,
        indicator=indicator,
    )
    return project_indicator


def main():
    args = build_args()
    workbook_path = Path(args.workbook).expanduser()
    if not workbook_path.exists():
        raise CommandError(f"Workbook not found: {workbook_path}")

    project = Project.objects.filter(id=args.project_id).first()
    if not project:
        raise CommandError(f"Project not found: {args.project_id}")

    coordinator = None
    if args.coordinator_id:
        coordinator = Organization.objects.filter(id=args.coordinator_id).first()
        if not coordinator:
            raise CommandError(f"Coordinator organization not found: {args.coordinator_id}")

    workbook = load_workbook(workbook_path, data_only=True)
    matrix_sheet_name = find_matrix_sheet_name(list(workbook.sheetnames))
    if not matrix_sheet_name:
        raise CommandError("Workbook is missing the Indicator matrix sheet.")

    matrix_assignments = parse_matrix_sheet(workbook[matrix_sheet_name])
    organization_resolver = OrganizationResolver()
    indicator_resolver = IndicatorResolver(project=project)

    sheet_payloads = []
    for sheet_name in workbook.sheetnames:
        if is_skipped_sheet(sheet_name):
            continue

        organization = organization_resolver.resolve(sheet_name)
        if not organization:
            continue

        sections = parse_sections(workbook[sheet_name])
        age_band_by_column = get_age_band_mapping(organization, coordinator)
        parsed_sections = []
        for section in sections:
            value, disaggregations = extract_section_value(
                section,
                age_band_by_column=age_band_by_column,
            )
            parsed_sections.append(
                {
                    "title": section["title"],
                    "index": section["index"],
                    "value": value,
                    "disaggregations": disaggregations,
                }
            )

        sheet_payloads.append(
            {
                "sheet_name": sheet_name,
                "organization": organization,
                "sections": parsed_sections,
            }
        )

    report = {
        "workbook": str(workbook_path),
        "project_id": project.id,
        "project_name": project.name,
        "coordinator_id": coordinator.id if coordinator else None,
        "coordinator_name": coordinator.name if coordinator else None,
        "period_start": args.period_start,
        "period_end": args.period_end,
        "dry_run": args.dry_run,
        "summary": {
            "sheets_processed": len(sheet_payloads),
            "created_indicators": 0,
            "updated_targets": 0,
            "created_aggregates": 0,
            "updated_aggregates": 0,
            "unchanged_aggregates": 0,
            "coordinator_rollups_created": 0,
            "coordinator_rollups_updated": 0,
            "coordinator_rollups_unchanged": 0,
        },
        "sheets": [],
    }

    with transaction.atomic():
        for payload in sheet_payloads:
            organization = payload["organization"]
            sheet_report = {
                "sheet_name": payload["sheet_name"],
                "organization_id": organization.id,
                "organization_name": organization.name,
                "sections": [],
            }
            project.organizations.add(organization)

            for section in payload["sections"]:
                title = section["title"]
                indicator = indicator_resolver.resolve(title, section["index"])
                desired_sub_labels = build_ordered_sub_labels(
                    section["disaggregations"],
                    section["value"],
                )
                desired_config = build_disaggregation_config(
                    section["value"],
                    desired_sub_labels,
                )

                if not indicator:
                    indicator = Indicator.objects.create(
                        name=title,
                        code=unique_indicator_code(f"{project.code}-{section['index']}-{title}"),
                        type="number",
                        category=args.category,
                        unit="people",
                        sub_labels=desired_sub_labels,
                        aggregate_disaggregation_config=desired_config,
                    )
                    indicator_resolver.remember(indicator)
                    report["summary"]["created_indicators"] += 1
                else:
                    update_fields = []
                    if desired_sub_labels and list(indicator.sub_labels or []) != desired_sub_labels:
                        indicator.sub_labels = desired_sub_labels
                        update_fields.append("sub_labels")

                    merged_config = merge_disaggregation_configs(
                        indicator.aggregate_disaggregation_config,
                        desired_config,
                    )
                    if merged_config != dict(indicator.aggregate_disaggregation_config or {}):
                        indicator.aggregate_disaggregation_config = merged_config
                        update_fields.append("aggregate_disaggregation_config")

                    if update_fields and not args.dry_run:
                        indicator.save(update_fields=update_fields)

                indicator.organizations.add(organization)
                ensure_project_indicator(project, indicator)

                assignment_bundle = matrix_assignments.get(canonical_indicator_name(title))
                if assignment_bundle:
                    for assignment in assignment_bundle["assignments"]:
                        assigned_organization = organization_resolver.resolve(
                            assignment["organization_name"]
                        )
                        if not assigned_organization or assigned_organization.id != organization.id:
                            continue

                        project.organizations.add(assigned_organization)
                        indicator.organizations.add(assigned_organization)

                aggregate, created = Aggregate.objects.get_or_create(
                    indicator=indicator,
                    project=project,
                    organization=organization,
                    period_start=args.period_start,
                    period_end=args.period_end,
                    defaults={"value": section["value"]},
                )

                notes = (
                    f"Imported from {workbook_path.name} | sheet={payload['sheet_name']} | "
                    f"index={section['index']}"
                )
                if created:
                    action = "created"
                    if hasattr(aggregate, "notes") and not args.dry_run:
                        aggregate.notes = notes
                        aggregate.save(update_fields=["notes"])
                    report["summary"]["created_aggregates"] += 1
                else:
                    update_fields = []
                    if aggregate.value != section["value"]:
                        aggregate.value = section["value"]
                        update_fields.append("value")
                    if hasattr(aggregate, "notes") and getattr(aggregate, "notes", None) != notes:
                        aggregate.notes = notes
                        update_fields.append("notes")

                    if update_fields and not args.dry_run:
                        aggregate.save(update_fields=update_fields)
                        action = "updated"
                        report["summary"]["updated_aggregates"] += 1
                    else:
                        action = "unchanged"
                        report["summary"]["unchanged_aggregates"] += 1

                sheet_report["sections"].append(
                    {
                        "index": section["index"],
                        "title": title,
                        "indicator_id": indicator.id,
                        "aggregate_id": aggregate.id,
                        "action": action,
                    }
                )

            report["sheets"].append(sheet_report)

        if coordinator:
            normalized_rollups: dict[int, dict[str, object]] = {}
            for payload in sheet_payloads:
                for section in payload["sections"]:
                    indicator = indicator_resolver.resolve(section["title"], section["index"])
                    if not indicator:
                        continue
                    if indicator.id not in normalized_rollups:
                        normalized_rollups[indicator.id] = {
                            "indicator": indicator,
                            "value": section["value"],
                        }
                    else:
                        normalized_rollups[indicator.id]["value"] = merge_json_values(
                            normalized_rollups[indicator.id]["value"],
                            section["value"],
                        )

            for rollup in normalized_rollups.values():
                aggregate, created = Aggregate.objects.update_or_create(
                    indicator=rollup["indicator"],
                    project=project,
                    organization=coordinator,
                    period_start=args.period_start,
                    period_end=args.period_end,
                    defaults={"value": rollup["value"]},
                )
                if created:
                    report["summary"]["coordinator_rollups_created"] += 1
                else:
                    report["summary"]["coordinator_rollups_updated"] += 1

                if hasattr(aggregate, "notes") and not args.dry_run:
                    aggregate.notes = (
                        f"Rolled up from workbook overwrite import {workbook_path.name} | "
                        f"coordinator={coordinator.name}"
                    )
                    aggregate.save(update_fields=["notes"])

        if args.dry_run:
            transaction.set_rollback(True)

    report_path = (
        Path(args.report_path)
        if args.report_path
        else Path.cwd() / "reports" / f"{workbook_path.stem}-overwrite-import.json"
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(json.dumps(report["summary"], indent=2))
    print(f"Saved report: {report_path}")


if __name__ == "__main__":
    main()
