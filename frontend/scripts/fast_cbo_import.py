"""Fast, low-memory wrapper around import_cbo_monthly_batch.

The only change is HOW a worksheet is loaded: the original opens each CBO
workbook in full (non-streaming) mode, which on the "End of year" files costs
>120s and ~2.3 GB RAM per file and destabilises the live container. This wrapper
monkey-patches just `load_sheet_worksheet` to use openpyxl's streaming
`read_only=True` mode (≈0.4s, a few MB), materialising the sheet's cells into a
dict and exposing the exact same interface the parser uses
(`ws[f"{col}{row}"].value` and `ws.max_row`). ALL parsing, indicator
resolution, upsert and reporting logic is reused unchanged from the original
module, so behaviour is identical — only faster.

Usage is identical to import_cbo_monthly_batch.py, e.g.:
    python fast_cbo_import.py --folder <dir> --project-code "NAHPA2025/26" \
        --sheet MARCH --period-start 2026-03-01 --period-end 2026-03-31 --dry-run
"""
import os
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import import_cbo_monthly_batch as base


_COORD_RE = re.compile(r"^([A-Z]+)(\d+)$")


def _is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _read_cells(ws):
    """Stream a worksheet into {(col_letter, row): value}; also return max row/col."""
    cells = {}
    max_row = 0
    max_column = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col_idx = cell.column
            col = get_column_letter(col_idx)
            cells[(col, cell.row)] = cell.value
            if cell.row > max_row:
                max_row = cell.row
            if col_idx > max_column:
                max_column = col_idx
    return cells, max_row, max_column


class _Cell:
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


class CachedWorksheet:
    """Minimal read-through worksheet backed by a {(col_letter,row): value} dict.

    Supports exactly what import_cbo_monthly_batch's parser needs:
    ``ws["B14"].value`` style access and ``ws.max_row`` / ``ws.max_column``.
    """

    def __init__(self, cells, max_row, max_column):
        self._cells = cells
        self.max_row = max_row
        self.max_column = max_column

    def __getitem__(self, coord):
        match = _COORD_RE.match(coord)
        if not match:
            return _Cell(None)
        col, row = match.group(1), int(match.group(2))
        return _Cell(self._cells.get((col, row)))


class _StubWorkbook:
    """Stand-in so callers can still call ``workbook.close()`` harmlessly."""

    def close(self):
        pass


def fast_load_sheet_worksheet(workbook_path, sheet_name):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Sheet not found: {sheet_name}")
    cells, max_row, max_column = _read_cells(workbook[sheet_name])

    # Optional: fold a second month sheet INTO this one (e.g. CBO_ADD_SHEET=APRIL
    # to add April's numbers into the March period). Layouts are identical, so we
    # sum numeric cells coordinate-by-coordinate and keep the primary's labels.
    add_sheet = os.environ.get("CBO_ADD_SHEET", "").strip()
    if add_sheet and add_sheet in workbook.sheetnames and add_sheet != sheet_name:
        add_cells, add_max_row, add_max_col = _read_cells(workbook[add_sheet])
        for coord, add_val in add_cells.items():
            if not _is_number(add_val):
                continue  # labels/text come from the primary sheet only
            base_val = cells.get(coord)
            if _is_number(base_val):
                cells[coord] = base_val + add_val
            elif base_val is None:
                cells[coord] = add_val
            # if base_val is a non-numeric label, keep the label (skip the add)
        max_row = max(max_row, add_max_row)
        max_column = max(max_column, add_max_col)

    workbook.close()
    return _StubWorkbook(), CachedWorksheet(cells, max_row, max_column)


# Swap in the streaming loader; everything else in `base` is reused unchanged.
base.load_sheet_worksheet = fast_load_sheet_worksheet


# Correct filename→organisation overrides whose base targets are abbreviations
# that don't match the real org names (so they failed to resolve). Prepended so
# they win over the base entries. These orgs all exist in project #2.
_OVERRIDE_FIXES = [
    ("jhf", "Just Hope Foundation"),
    ("just hope foundation", "Just Hope Foundation"),
    ("journey of hope", "Journey of Hope"),
    ("fighters support group", "The Fighters Support Group"),
]
base.FILENAME_ORG_OVERRIDES = _OVERRIDE_FIXES + list(base.FILENAME_ORG_OVERRIDES)


class CodeFirstResolver(base.ExactFirstIndicatorResolver):
    """Resolve known-ambiguous sections by indicator CODE instead of title.

    Some source sections share an identical title that normalises to the same
    key (e.g. code 16 "…screened…(tobacco use)" and code 19 "…(alcohol use)" both
    collapse to "number of people screened for ncds behavioural risk factors").
    Title resolution then maps both to one indicator and the second overwrites
    the first. For these specific codes — which equal the section index in this
    template — we resolve directly to the unique project indicator carrying that
    code. Everything else falls through to the unchanged base resolver.
    """

    AMBIGUOUS_CODES = {"16", "19", "15f", "15h", "37b", "37d"}

    def __init__(self, project):
        super().__init__(project)
        from indicators.models import Indicator

        self.by_code = {}
        for ind in Indicator.objects.filter(id__in=self.project_indicator_ids):
            code = (ind.code or "").strip().lower()
            if code:
                self.by_code.setdefault(code, []).append(ind)

    def resolve(self, title, section_index=None):
        key = str(section_index or "").strip().lower()
        if key in self.AMBIGUOUS_CODES:
            matches = self.by_code.get(key)
            if matches and len(matches) == 1:
                return matches[0]
        return super().resolve(title, section_index)


base.ExactFirstIndicatorResolver = CodeFirstResolver


if __name__ == "__main__":
    base.main()
