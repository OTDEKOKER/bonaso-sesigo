import argparse
import json
import os
import sys
from pathlib import Path


BACKEND_ROOT = Path(os.environ.get("BONASO_DJANGO_ROOT", r"C:\Projects\django_backend"))
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django  # noqa: E402

django.setup()

from django.db import transaction  # noqa: E402

from aggregates.models import Aggregate  # noqa: E402
from indicators.models import Indicator  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from organizations.models import Organization  # noqa: E402
from projects.models import Project, ProjectIndicator  # noqa: E402
from uploads.management.commands.import_reporting_workbook_overwrite import (  # noqa: E402
    IndicatorResolver,
    build_disaggregation_config,
    build_ordered_sub_labels,
    canonical_indicator_name,
    extract_section_value,
    get_age_band_mapping,
    merge_disaggregation_configs,
    normalize_text,
    parse_sections,
    unique_indicator_code,
)


def is_importable_section(title: str) -> bool:
    normalized = normalize_text(title)
    return normalized.startswith("number") or normalized.startswith("total number")


def indicator_sort_key(indicator: Indicator, project_indicator_ids: set[int]) -> tuple:
    return (
        0 if indicator.id in project_indicator_ids else 1,
        0 if not indicator.code.upper().startswith("AUTO_") else 1,
        indicator.name.lower(),
    )


class ExactFirstIndicatorResolver:
    def __init__(self, project: Project):
        self.delegate = IndicatorResolver(project=project)
        self.project_indicator_ids = set(
            ProjectIndicator.objects.filter(project=project).values_list("indicator_id", flat=True)
        )
        self.exact_index: dict[str, list[Indicator]] = {}
        for indicator in Indicator.objects.all():
            key = canonical_indicator_name(indicator.name)
            self.exact_index.setdefault(key, []).append(indicator)

    def resolve(self, title: str, section_index: str | None = None) -> Indicator | None:
        key = canonical_indicator_name(title)
        exact_matches = self.exact_index.get(key, [])
        if exact_matches:
            return sorted(
                exact_matches,
                key=lambda item: indicator_sort_key(item, self.project_indicator_ids),
            )[0]
        return self.delegate.resolve(title, section_index)

    def remember(self, indicator: Indicator):
        key = canonical_indicator_name(indicator.name)
        self.exact_index.setdefault(key, []).append(indicator)
        self.delegate.remember(indicator)


def build_args():
    parser = argparse.ArgumentParser(
        description="Import a single-organization monthly workbook using the TOTALS sheet."
    )
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--organization", required=True)
    parser.add_argument("--project-id", type=int, default=0)
    parser.add_argument("--project-code", default="")
    parser.add_argument("--sheet", default="TOTALS")
    parser.add_argument("--period-start", required=True)
    parser.add_argument("--period-end", required=True)
    parser.add_argument("--category", default="hiv_prevention")
    parser.add_argument("--skip-indexes", nargs="*", default=[])
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report-path", default="")
    return parser.parse_args()


def main():
    args = build_args()
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    project = None
    if args.project_id:
        project = Project.objects.filter(id=args.project_id).first()
    if project is None and args.project_code:
        project = Project.objects.filter(code__iexact=args.project_code).first()
    if project is None:
        raise SystemExit("Project not found.")

    organization = Organization.objects.filter(name__iexact=args.organization).first()
    if organization is None:
        raise SystemExit(f"Organization not found: {args.organization}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True, keep_links=False)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}")

    skip_indexes = {str(value).strip().lower() for value in args.skip_indexes if str(value).strip()}
    age_band_by_column = get_age_band_mapping(organization, None)
    resolver = ExactFirstIndicatorResolver(project=project)
    sections = parse_sections(workbook[args.sheet])

    report = {
        "workbook": str(workbook_path),
        "sheet": args.sheet,
        "project": {"id": project.id, "code": project.code, "name": project.name},
        "organization": {"id": organization.id, "name": organization.name},
        "period_start": args.period_start,
        "period_end": args.period_end,
        "dry_run": args.dry_run,
        "summary": {
            "sections_seen": len(sections),
            "sections_importable": 0,
            "sections_skipped": 0,
            "missing_indicators": 0,
            "created_indicators": 0,
            "created_aggregates": 0,
            "updated_aggregates": 0,
            "unchanged_aggregates": 0,
        },
        "skipped_sections": [],
        "missing_indicator_sections": [],
        "imported_sections": [],
    }

    with transaction.atomic():
        for section in sections:
            index = str(section["index"] or "").strip()
            title = str(section["title"] or "").strip()
            normalized_index = index.lower()

            if normalized_index in skip_indexes:
                report["summary"]["sections_skipped"] += 1
                report["skipped_sections"].append(
                    {"index": index, "title": title, "reason": "explicitly_skipped"}
                )
                continue

            if not is_importable_section(title):
                report["summary"]["sections_skipped"] += 1
                report["skipped_sections"].append(
                    {"index": index, "title": title, "reason": "non_indicator_header"}
                )
                continue

            report["summary"]["sections_importable"] += 1
            value, disaggregations = extract_section_value(section, age_band_by_column)

            indicator = resolver.resolve(title, index)
            created_indicator = False

            desired_sub_labels = build_ordered_sub_labels(disaggregations, value)
            desired_config = build_disaggregation_config(value, desired_sub_labels)

            if indicator is None:
                report["summary"]["missing_indicators"] += 1
                report["missing_indicator_sections"].append(
                    {"index": index, "title": title, "value": value}
                )
                if args.dry_run:
                    continue

                indicator_payload = dict(
                    name=title,
                    code=unique_indicator_code(f"{project.code}-{index}-{title}"),
                    type="number",
                    category=args.category,
                    unit="people",
                    sub_labels=desired_sub_labels,
                )
                if hasattr(Indicator, "aggregate_disaggregation_config"):
                    indicator_payload["aggregate_disaggregation_config"] = desired_config

                indicator = Indicator.objects.create(**indicator_payload)
                resolver.remember(indicator)
                created_indicator = True
                report["summary"]["created_indicators"] += 1
            else:
                update_fields = []
                if desired_sub_labels and not list(indicator.sub_labels or []):
                    indicator.sub_labels = desired_sub_labels
                    update_fields.append("sub_labels")
                if hasattr(indicator, "aggregate_disaggregation_config"):
                    merged_config = merge_disaggregation_configs(
                        indicator.aggregate_disaggregation_config,
                        desired_config,
                    )
                    if merged_config != dict(indicator.aggregate_disaggregation_config or {}):
                        indicator.aggregate_disaggregation_config = merged_config
                        update_fields.append("aggregate_disaggregation_config")
                if update_fields and not args.dry_run:
                    indicator.save(update_fields=update_fields)

            if args.dry_run:
                report["imported_sections"].append(
                    {
                        "index": index,
                        "title": title,
                        "indicator_id": indicator.id if indicator else None,
                        "indicator_name": indicator.name if indicator else None,
                        "action": "would_create_indicator" if created_indicator else "would_import",
                        "value": value,
                    }
                )
                continue

            project.organizations.add(organization)
            indicator.organizations.add(organization)
            ProjectIndicator.objects.get_or_create(project=project, indicator=indicator)

            aggregate, created_aggregate = Aggregate.objects.get_or_create(
                indicator=indicator,
                project=project,
                organization=organization,
                period_start=args.period_start,
                period_end=args.period_end,
                defaults={"value": value},
            )

            action = "created"
            if created_aggregate:
                report["summary"]["created_aggregates"] += 1
            else:
                update_fields = []
                if aggregate.value != value:
                    aggregate.value = value
                    update_fields.append("value")
                if hasattr(aggregate, "notes"):
                    notes = f"Imported from {workbook_path.name} | sheet={args.sheet} | index={index}"
                    if getattr(aggregate, "notes", None) != notes:
                        aggregate.notes = notes
                        update_fields.append("notes")
                if update_fields:
                    aggregate.save(update_fields=update_fields)
                    action = "updated"
                    report["summary"]["updated_aggregates"] += 1
                else:
                    action = "unchanged"
                    report["summary"]["unchanged_aggregates"] += 1

            if created_aggregate and hasattr(aggregate, "notes"):
                aggregate.notes = f"Imported from {workbook_path.name} | sheet={args.sheet} | index={index}"
                aggregate.save(update_fields=["notes"])

            report["imported_sections"].append(
                {
                    "index": index,
                    "title": title,
                    "indicator_id": indicator.id,
                    "indicator_name": indicator.name,
                    "aggregate_id": aggregate.id,
                    "action": action,
                    "value": value,
                }
            )

        if args.dry_run:
            transaction.set_rollback(True)

    report_path = (
        Path(args.report_path)
        if args.report_path
        else Path.cwd() / "reports" / "single-org-reporting-import.json"
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(json.dumps(report["summary"], indent=2))
    print(f"Saved report: {report_path}")


if __name__ == "__main__":
    main()
